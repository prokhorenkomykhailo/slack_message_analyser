#!/usr/bin/env python3
"""
Check token costs across all Excel files in step2_client_analysis
"""

import os
import glob
from openpyxl import load_workbook

def extract_token_info_from_excel(excel_path):
    """Extract token information from Excel file"""
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        model_name = None
        input_tokens = None
        output_tokens = None
        total_tokens = None
        cost_per_input = None
        cost_per_output = None
        token_cost = None
        
        # Scan through rows to find token information
        for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
            if row[0] == 'INPUT_TOKENS':
                input_tokens = row[1]
            elif row[0] == 'OUTPUT_TOKENS':
                output_tokens = row[1]
            elif row[0] == 'TOTAL_TOKENS':
                total_tokens = row[1]
            elif row[0] == 'COST_PER_INPUT_TOKEN':
                cost_per_input = row[1]
            elif row[0] == 'COST_PER_OUTPUT_TOKEN':
                cost_per_output = row[1]
            elif row[0] == 'TOKEN_COST':
                token_cost = row[1]
            elif row[0] and 'Model:' in str(row[0]):
                model_name = str(row[0]).replace('Model:', '').strip()
        
        # Extract model name from filename if not found
        if not model_name:
            basename = os.path.basename(excel_path)
            model_name = basename.replace('_step2_FULLY_EXPLAINED.xlsx', '')
        
        return {
            'file': os.path.basename(excel_path),
            'model': model_name,
            'input_tokens': input_tokens,
            'output_tokens': output_tokens,
            'total_tokens': total_tokens,
            'cost_per_input': cost_per_input,
            'cost_per_output': cost_per_output,
            'token_cost': token_cost
        }
    except Exception as e:
        print(f"Error reading {excel_path}: {e}")
        return None

def main():
    output_dir = 'output/step2_client_analysis'
    excel_files = glob.glob(os.path.join(output_dir, '*_step2_FULLY_EXPLAINED.xlsx'))
    
    print("=" * 100)
    print("TOKEN COST COMPARISON ACROSS ALL EXCEL FILES")
    print("=" * 100)
    print()
    
    all_data = []
    for excel_file in sorted(excel_files):
        data = extract_token_info_from_excel(excel_file)
        if data:
            all_data.append(data)
    
    # Print in a table format
    print(f"{'Model':<40} {'Input':<10} {'Output':<10} {'Total':<10} {'Cost/In':<12} {'Cost/Out':<12} {'Total Cost':<15}")
    print("-" * 120)
    
    for data in all_data:
        model = data['model'][:38]
        input_tok = str(data['input_tokens']) if data['input_tokens'] else 'N/A'
        output_tok = str(data['output_tokens']) if data['output_tokens'] else 'N/A'
        total_tok = str(data['total_tokens']) if data['total_tokens'] else 'N/A'
        cost_in = str(data['cost_per_input']) if data['cost_per_input'] else 'N/A'
        cost_out = str(data['cost_per_output']) if data['cost_per_output'] else 'N/A'
        cost_total = str(data['token_cost']) if data['token_cost'] else 'N/A'
        
        print(f"{model:<40} {input_tok:<10} {output_tok:<10} {total_tok:<10} {cost_in:<12} {cost_out:<12} {cost_total:<15}")
    
    print()
    print("=" * 100)
    print(f"Total files analyzed: {len(all_data)}")
    print("=" * 100)

if __name__ == '__main__':
    main()

