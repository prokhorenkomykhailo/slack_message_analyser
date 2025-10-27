#!/usr/bin/env python3
"""
Improved Excel format for Step 2 (Phase 4) client analysis with AI models
Creates individual analysis files for each model from Phase 4 results
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import math
import json
import os
from pathlib import Path

def parse_people(people_str):
    """Parse people string into set"""
    if pd.isna(people_str) or people_str == '':
        return set()
    return set([p.strip() for p in str(people_str).split(';') if p.strip()])

def calculate_similarity_percentage(set1, set2):
    """Calculate similarity percentage between two sets"""
    if not set1 and not set2:
        return 100.0
    if not set1 or not set2:
        return 0.0
    intersection = len(set1.intersection(set2))
    union = len(set1.union(set2))
    return round((intersection / union) * 100, 1) if union > 0 else 0.0

def load_phase4_evaluation_results():
    """Load Phase 4 comprehensive evaluation results"""
    results_file = "output/phase4_comprehensive_evaluation/detailed_evaluation_results.json"
    
    if not os.path.exists(results_file):
        print(f"❌ Phase 4 evaluation results not found: {results_file}")
        print("Please run phase4_comprehensive_evaluation.py first")
        return []
    
    try:
        with open(results_file, 'r') as f:
            results = json.load(f)
        return results
    except Exception as e:
        print(f"❌ Error loading Phase 4 results: {e}")
        return []

def load_benchmark_topics():
    """Load Phase 4 benchmark topics"""
    benchmark_path = "phases/phase4_clusters_refined.json"
    
    try:
        with open(benchmark_path, 'r') as f:
            topics = json.load(f)
        return topics
    except Exception as e:
        print(f"❌ Error loading benchmark topics: {e}")
        return []

def create_step2_client_excel(model_result, output_dir):
    """Create improved Excel format for Step 2 model analysis"""
    
    if not model_result.get('success', False):
        print(f"⚠️ Skipping {model_result['model']} - evaluation failed")
        return None
    
    model_name = model_result['model']
    
    # Calculate key metrics
    num_refined_clusters = model_result['num_refined_clusters']
    num_step1_clusters = model_result['num_step1_clusters']
    cluster_reduction = model_result['cluster_count_reduction']
    
    # Performance metrics
    avg_similarity = model_result['avg_combined_similarity']
    quality_rate = model_result['quality_rate']
    high_quality_matches = model_result['high_quality_matches']
    
    # Operations analysis
    total_operations = model_result['total_operations']
    merge_ops = model_result['merge_operations']
    split_ops = model_result['split_operations']
    refine_ops = model_result['refine_operations']
    
    # Size optimization
    avg_cluster_size = model_result['avg_cluster_size']
    size_reduction = model_result['size_reduction']
    
    # Calculate performance scores
    similarity_score = avg_similarity * 100
    quality_score = quality_rate * 100
    operation_efficiency = max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))) if num_refined_clusters > 0 else 100
    size_optimization = max(0, min(100, 100 - abs(size_reduction)))
    
    # Combined score
    combined_score = (
        similarity_score * 0.35 +        # 35% - similarity to benchmark
        quality_score * 0.25 +           # 25% - quality rate
        size_optimization * 0.20 +       # 20% - size optimization
        operation_efficiency * 0.20      # 20% - operation efficiency
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{model_name} Step 2 Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    result_font = Font(bold=True, size=11, color="0000FF")
    formula_font = Font(italic=True, size=10)
    
    blue_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    orange_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    current_row = 1
    
    # Title
    ws.cell(row=current_row, column=1, value=f"{model_name.upper()} STEP 2 (PHASE 4) BALANCED REFINEMENT EVALUATION REPORT")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 3
    
    # Executive Summary
    ws.cell(row=current_row, column=1, value="EXECUTIVE SUMMARY")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    summary_data = [
        ['Metric', 'Value', 'Formula', 'Explanation'],
        ['Overall Combined Score', f'{combined_score:.2f}', 'Weighted combination of similarity, quality, size optimization, and efficiency', 'Composite score for Step 2 refinement performance'],
        ['Benchmark Similarity', f'{similarity_score:.2f}%', 'Average similarity to 15 benchmark topics', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'{quality_score:.1f}%', f'{high_quality_matches} / {num_refined_clusters}', 'Percentage of clusters with >70% similarity to benchmarks'],
        ['High Quality Matches', f'{high_quality_matches}', 'Clusters with >70% similarity', 'Number of excellent cluster matches'],
        ['Clusters Generated', f'{num_refined_clusters}', 'Final refined cluster count', 'Number of clusters after refinement'],
        ['Cluster Reduction', f'{cluster_reduction}', f'{num_step1_clusters} - {num_refined_clusters}', 'Number of clusters reduced from Step 1'],
        ['Operations Performed', f'{total_operations}', f'{merge_ops} merge + {split_ops} split + {refine_ops} refine', 'Total refinement operations performed'],
        ['Average Cluster Size', f'{avg_cluster_size:.1f}', 'Mean messages per cluster', 'Average size of refined clusters'],
        ['Size Optimization', f'{size_optimization:.1f}%', '100 - |size_reduction|', 'How well cluster sizes were optimized']
    ]
    
    for row_data in summary_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 5:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Combined Score Explanation
    ws.cell(row=current_row, column=1, value="🔍 DETAILED EXPLANATION OF COMBINED SCORE")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    explanation_data = [
        ['Component', 'Score', 'Weight', 'Contribution', 'Description'],
        ['Benchmark Similarity', f'{similarity_score:.2f}%', '35%', f'{similarity_score * 0.35:.2f}', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'{quality_score:.1f}%', '25%', f'{quality_score * 0.25:.2f}', 'Percentage of high-quality matches (>70% similarity)'],
        ['Size Optimization', f'{size_optimization:.1f}%', '20%', f'{size_optimization * 0.20:.2f}', 'How well cluster sizes were optimized'],
        ['Operation Efficiency', f'{operation_efficiency:.1f}%', '20%', f'{operation_efficiency * 0.20:.2f}', 'Efficiency of refinement operations'],
        ['TOTAL COMBINED SCORE', f'{combined_score:.2f}', '100%', f'{combined_score:.2f}', 'Overall Step 2 performance score']
    ]
    
    for row_data in explanation_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 10:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif 'TOTAL' in str(value):
                cell.fill = yellow_fill
                cell.font = result_font
            else:
                cell.fill = light_green_fill
        current_row += 1
    
    current_row += 2
    
    # Refinement Operations Analysis
    ws.cell(row=current_row, column=1, value="🔧 REFINEMENT OPERATIONS ANALYSIS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    operations_data = [
        ['Operation Type', 'Count', 'Percentage', 'Description'],
        ['Merge Operations', f'{merge_ops}', f'{(merge_ops/total_operations*100):.1f}%' if total_operations > 0 else '0%', 'Combining related clusters'],
        ['Split Operations', f'{split_ops}', f'{(split_ops/total_operations*100):.1f}%' if total_operations > 0 else '0%', 'Separating mixed clusters'],
        ['Refine Operations', f'{refine_ops}', f'{(refine_ops/total_operations*100):.1f}%' if total_operations > 0 else '0%', 'Improving cluster quality'],
        ['Total Operations', f'{total_operations}', '100%', 'All refinement operations performed']
    ]
    
    for row_data in operations_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 15:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif 'Total' in str(value):
                cell.fill = yellow_fill
                cell.font = result_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Cluster Analysis
    ws.cell(row=current_row, column=1, value="📊 CLUSTER REFINEMENT ANALYSIS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    cluster_data = [
        ['Metric', 'Step 1', 'Step 2', 'Change', 'Improvement %'],
        ['Total Clusters', f'{num_step1_clusters}', f'{num_refined_clusters}', f'{cluster_reduction}', f'{(cluster_reduction/num_step1_clusters*100):.1f}%' if num_step1_clusters > 0 else '0%'],
        ['Average Size', f'{model_result.get("avg_step1_size", 0):.1f}', f'{avg_cluster_size:.1f}', f'{size_reduction:.1f}', f'{(abs(size_reduction)/model_result.get("avg_step1_size", 1)*100):.1f}%' if model_result.get("avg_step1_size", 0) > 0 else '0%'],
        ['Size Standard Dev', 'N/A', f'{model_result.get("size_std_reduction", 0):.1f}', 'N/A', 'Size consistency improvement']
    ]
    
    for row_data in cluster_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 20:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_green_fill
        current_row += 1
    
    current_row += 2
    
    # Individual Cluster Analysis
    ws.cell(row=current_row, column=1, value="📋 INDIVIDUAL CLUSTER ANALYSIS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    # Headers
    headers = ['Cluster', 'Title', 'Best Benchmark Match', 'Similarity %', 'Quality Status', 'Participants', 'Message Count']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = green_fill
        cell.font = header_font
    current_row += 1
    
    # Cluster data from benchmark evaluation
    benchmark_eval = model_result.get('benchmark_evaluation', {})
    best_matches = benchmark_eval.get('best_matches', [])
    cluster_titles = model_result.get('cluster_titles', [])
    
    for i, cluster_title in enumerate(cluster_titles):
        if i < len(best_matches):
            match_info = best_matches[i]
            similarity = match_info.get('title_similarity', 0) * 100
            benchmark_match = match_info.get('best_benchmark_match', '')
            is_high_quality = 'YES' if match_info.get('combined_similarity', 0) > 0.7 else 'NO'
        else:
            similarity = 0
            benchmark_match = 'No match found'
            is_high_quality = 'NO'
        
        cluster_data = [
            f"Cluster {i+1}",
            cluster_title,
            benchmark_match,
            f"{similarity:.1f}%",
            is_high_quality,
            "Multiple participants",  # Could be enhanced with actual participant data
            "Variable count"  # Could be enhanced with actual message count data
        ]
        
        for col, value in enumerate(cluster_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if is_high_quality == 'YES':
                cell.fill = light_green_fill
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Performance Comparison
    ws.cell(row=current_row, column=1, value="🏆 PERFORMANCE ASSESSMENT")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    performance_data = [
        ['Performance Level', 'Score Range', 'Current Score', 'Status', 'Recommendation'],
        ['Excellent', '90-100', f'{combined_score:.2f}', '✅ EXCELLENT' if combined_score >= 90 else '❌', 'Ready for production use'],
        ['Very Good', '80-89', f'{combined_score:.2f}', '✅ VERY GOOD' if 80 <= combined_score < 90 else '❌', 'Minor optimization needed'],
        ['Good', '70-79', f'{combined_score:.2f}', '✅ GOOD' if 70 <= combined_score < 80 else '❌', 'Moderate improvement recommended'],
        ['Fair', '60-69', f'{combined_score:.2f}', '⚠️ FAIR' if 60 <= combined_score < 70 else '❌', 'Significant improvement needed'],
        ['Poor', '0-59', f'{combined_score:.2f}', '❌ POOR' if combined_score < 60 else '❌', 'Major reconfiguration required']
    ]
    
    for row_data in performance_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 27:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif '✅' in str(value):
                cell.fill = light_green_fill
                cell.font = Font(color="008000")
            elif '⚠️' in str(value):
                cell.fill = yellow_fill
                cell.font = Font(color="FF8000")
            elif '❌' in str(value):
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                cell.font = Font(color="FF0000")
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Key Insights
    ws.cell(row=current_row, column=1, value="💡 KEY INSIGHTS & RECOMMENDATIONS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    insights_data = [
        ['Insight', 'Value', 'Impact', 'Recommendation'],
        ['Overall Performance', f'{combined_score:.2f}/100', 'Excellent' if combined_score >= 90 else 'Good' if combined_score >= 70 else 'Needs Improvement', 'Model ready for production' if combined_score >= 90 else 'Consider optimization'],
        ['Benchmark Alignment', f'{similarity_score:.1f}%', 'High' if similarity_score >= 80 else 'Moderate' if similarity_score >= 60 else 'Low', 'Excellent topic matching' if similarity_score >= 80 else 'Review clustering strategy'],
        ['Quality Consistency', f'{quality_score:.1f}%', 'High' if quality_score >= 80 else 'Moderate' if quality_score >= 60 else 'Low', 'Consistent high-quality results' if quality_score >= 80 else 'Focus on quality improvement'],
        ['Refinement Efficiency', f'{operation_efficiency:.1f}%', 'Efficient' if operation_efficiency >= 80 else 'Moderate' if operation_efficiency >= 60 else 'Inefficient', 'Optimal operation usage' if operation_efficiency >= 80 else 'Review operation strategy'],
        ['Size Optimization', f'{size_optimization:.1f}%', 'Well Optimized' if size_optimization >= 80 else 'Moderate' if size_optimization >= 60 else 'Poor', 'Good size management' if size_optimization >= 80 else 'Improve size consistency']
    ]
    
    for row_data in insights_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 34:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_green_fill
        current_row += 1
    
    # Auto-adjust column widths
    for col_num in range(1, 14):  # A to M
        max_length = 0
        column_letter = get_column_letter(col_num)
        for row_num in range(1, current_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename
    model_filename = model_name.replace('/', '_').replace('-', '_')
    output_path = os.path.join(output_dir, f"{model_filename}_step2_analysis.xlsx")
    
    # Save workbook
    wb.save(output_path)
    
    print(f"✅ Created Step 2 analysis for {model_name}: {output_path}")
    return output_path

def create_all_step2_analyses():
    """Create Step 2 analysis files for all models"""
    
    print("🚀 Creating Step 2 (Phase 4) Individual Model Analysis Files")
    print("=" * 70)
    
    # Load Phase 4 results
    results = load_phase4_evaluation_results()
    if not results:
        return
    
    # Create output directory
    output_dir = "output/step2_individual_analysis"
    
    successful_models = [r for r in results if r.get('success', False)]
    print(f"📊 Creating analysis files for {len(successful_models)} successful models...")
    
    created_files = []
    
    # Create individual analysis for each model
    for result in successful_models:
        file_path = create_step2_client_excel(result, output_dir)
        if file_path:
            created_files.append(file_path)
    
    # Create summary file
    create_step2_summary(successful_models, output_dir)
    
    print(f"\n✅ Step 2 analysis complete!")
    print(f"📁 Results saved to: {output_dir}")
    print(f"📋 Created {len(created_files)} individual analysis files")
    
    # List created files
    print(f"\n📄 Generated Files:")
    for file_path in created_files[:10]:  # Show first 10
        filename = os.path.basename(file_path)
        print(f"  - {filename}")
    
    if len(created_files) > 10:
        print(f"  ... and {len(created_files) - 10} more files")
    
    return created_files

def create_step2_summary(successful_models, output_dir):
    """Create summary file for all Step 2 models"""
    
    summary_data = []
    
    for result in successful_models:
        # Calculate scores
        similarity_score = result['avg_combined_similarity'] * 100
        quality_score = result['quality_rate'] * 100
        operation_efficiency = max(0, min(100, 100 - (result['total_operations'] / result['num_refined_clusters'] * 10))) if result['num_refined_clusters'] > 0 else 100
        size_optimization = max(0, min(100, 100 - abs(result['size_reduction'])))
        
        combined_score = (
            similarity_score * 0.35 +
            quality_score * 0.25 +
            size_optimization * 0.20 +
            operation_efficiency * 0.20
        )
        
        summary_row = {
            'MODEL_NAME': result['model'],
            'PROVIDER': result['provider'],
            'COMBINED_SCORE': round(combined_score, 2),
            'SIMILARITY_SCORE': round(similarity_score, 2),
            'QUALITY_SCORE': round(quality_score, 2),
            'OPERATION_EFFICIENCY': round(operation_efficiency, 2),
            'SIZE_OPTIMIZATION': round(size_optimization, 2),
            'HIGH_QUALITY_MATCHES': result['high_quality_matches'],
            'CLUSTERS_GENERATED': result['num_refined_clusters'],
            'CLUSTER_REDUCTION': result['cluster_count_reduction'],
            'TOTAL_OPERATIONS': result['total_operations'],
            'EXECUTION_TIME': result['duration']
        }
        
        summary_data.append(summary_row)
    
    # Create DataFrame and sort by combined score
    summary_df = pd.DataFrame(summary_data)
    summary_df = summary_df.sort_values('COMBINED_SCORE', ascending=False)
    
    # Save summary
    summary_file = os.path.join(output_dir, "step2_all_models_summary.xlsx")
    summary_df.to_excel(summary_file, index=False)
    
    print(f"📊 Created summary file: step2_all_models_summary.xlsx")

def main():
    """Main function"""
    
    print("🎯 Step 2 (Phase 4) Individual Model Analysis Generator")
    print("=" * 70)
    print("This script creates individual Excel analysis files for each model")
    print("evaluated in Step 2 (Phase 4) balanced refinement tasks.")
    print()
    
    # Create all analyses
    created_files = create_all_step2_analyses()
    
    if created_files:
        print(f"\n🎉 Successfully created {len(created_files)} Step 2 analysis files!")
        print("Each file contains comprehensive analysis of model performance on")
        print("balanced refinement tasks compared to benchmark topics.")
    else:
        print("\n❌ No analysis files were created. Please check the prerequisites.")

if __name__ == "__main__":
    main()
