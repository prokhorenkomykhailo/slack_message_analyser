#!/usr/bin/env python3
"""
Create FINAL CORRECT Step 2 analysis with proper token costs using actual pricing
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import glob

# Actual pricing data provided by user
PRICING_DATA = [
    {"MODEL": "gemini-1.5-flash", "COST_PER_INPUT_TOKEN": 0.000075, "COST_PER_OUTPUT_TOKEN": 0.0003},
    {"MODEL": "gemini-1.5-flash-002", "COST_PER_INPUT_TOKEN": 0.000075, "COST_PER_OUTPUT_TOKEN": 0.0003},
    {"MODEL": "gemini-1.5-flash-8b", "COST_PER_INPUT_TOKEN": 0.000075, "COST_PER_OUTPUT_TOKEN": 0.0003},
    {"MODEL": "gemini-1.5-flash-latest", "COST_PER_INPUT_TOKEN": 0.000075, "COST_PER_OUTPUT_TOKEN": 0.0003},
    {"MODEL": "gemini-1.5-pro", "COST_PER_INPUT_TOKEN": 0.00125, "COST_PER_OUTPUT_TOKEN": 0.005},
    {"MODEL": "gemini-1.5-pro-002", "COST_PER_INPUT_TOKEN": 0.00125, "COST_PER_OUTPUT_TOKEN": 0.005},
    {"MODEL": "gemini-1.5-pro-latest", "COST_PER_INPUT_TOKEN": 0.00125, "COST_PER_OUTPUT_TOKEN": 0.005},
    {"MODEL": "gemini-2.0-flash", "COST_PER_INPUT_TOKEN": 0.0001, "COST_PER_OUTPUT_TOKEN": 0.0004},
    {"MODEL": "gemini-2.0-flash-001", "COST_PER_INPUT_TOKEN": 0.0001, "COST_PER_OUTPUT_TOKEN": 0.0004},
    {"MODEL": "gemini-2.0-flash-lite", "COST_PER_INPUT_TOKEN": 0.000075, "COST_PER_OUTPUT_TOKEN": 0.0003},
    {"MODEL": "gemini-2.0-flash-lite-001", "COST_PER_INPUT_TOKEN": 0.000075, "COST_PER_OUTPUT_TOKEN": 0.0003},
    {"MODEL": "gemini-2.5-flash-lite", "COST_PER_INPUT_TOKEN": 0.0001, "COST_PER_OUTPUT_TOKEN": 0.0004},
    {"MODEL": "gemini-2.5-pro", "COST_PER_INPUT_TOKEN": 0.00125, "COST_PER_OUTPUT_TOKEN": 0.0025},
    {"MODEL": "gemma-3-12b-it", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0},
    {"MODEL": "gemma-3-1b-it", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0},
    {"MODEL": "gemma-3-27b-it", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0},
    {"MODEL": "gemma-3-4b-it", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0},
    {"MODEL": "gemma-3n-e2b-it", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0},
    {"MODEL": "gemma-3n-e4b-it", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0},
    {"MODEL": "gpt-4o", "COST_PER_INPUT_TOKEN": 0.0025, "COST_PER_OUTPUT_TOKEN": 0.01},
    {"MODEL": "gpt-5", "COST_PER_INPUT_TOKEN": 0.00125, "COST_PER_OUTPUT_TOKEN": 0.01},
    {"MODEL": "grok-2-1212", "COST_PER_INPUT_TOKEN": 0.002, "COST_PER_OUTPUT_TOKEN": 0.01},
    {"MODEL": "grok-2-vision-1212", "COST_PER_INPUT_TOKEN": 0.002, "COST_PER_OUTPUT_TOKEN": 0.01},
    {"MODEL": "grok-3", "COST_PER_INPUT_TOKEN": 0.002, "COST_PER_OUTPUT_TOKEN": 0.01},
    {"MODEL": "gpt-3.5-turbo", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0},
    {"MODEL": "gpt-4", "COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0}
]

# Convert to dict for easy lookup
PRICING_MAP = {item['MODEL']: item for item in PRICING_DATA}

def get_model_pricing(model_name):
    """Get pricing for a model by extracting model name from full identifier"""
    # Extract just the model name (e.g., "google_gemini-2.0-flash-001" -> "gemini-2.0-flash-001")
    if '_' in model_name:
        model_short = model_name.split('_', 1)[1]
    else:
        model_short = model_name
    
    # Try exact match first
    if model_short in PRICING_MAP:
        return PRICING_MAP[model_short]
    
    # Try partial match (for models with versions)
    for pricing_model, pricing in PRICING_MAP.items():
        if pricing_model in model_short:
            return pricing
    
    # Default to 0 if not found
    return {"COST_PER_INPUT_TOKEN": 0, "COST_PER_OUTPUT_TOKEN": 0}

def load_benchmark_clusters():
    """Load benchmark clusters"""
    with open("phases/phase4_clusters_refined.json", 'r') as f:
        return json.load(f)

def calculate_metrics_with_correct_pricing(model_data, benchmark_data, model_name):
    """Calculate metrics with correct pricing"""
    
    model_clusters = model_data.get('refined_clusters', [])
    if not model_clusters:
        return None
    
    benchmark_clusters = benchmark_data
    
    # Message ID sets
    model_cluster_messages = {c['cluster_id']: set(c.get('message_ids', [])) for c in model_clusters}
    benchmark_cluster_messages = {c['cluster_id']: set(c['message_ids']) for c in benchmark_clusters}
    
    # Calculate detailed results
    detailed_results = []
    for model_cluster in model_clusters:
        model_id = model_cluster['cluster_id']
        model_messages = model_cluster_messages[model_id]
        
        # Find best match
        best_match, best_overlap, best_benchmark_id = None, 0, None
        for benchmark_cluster in benchmark_clusters:
            benchmark_id = benchmark_cluster['cluster_id']
            overlap = len(model_messages.intersection(benchmark_cluster_messages[benchmark_id]))
            if overlap > best_overlap:
                best_overlap, best_match, best_benchmark_id = overlap, benchmark_cluster, benchmark_id
        
        if best_match:
            benchmark_messages = benchmark_cluster_messages[best_benchmark_id]
            matched = len(model_messages.intersection(benchmark_messages))
            missing = len(benchmark_messages - model_messages)
            extra = len(model_messages - benchmark_messages)
            
            precision = (matched / len(model_messages)) * 100 if len(model_messages) > 0 else 0
            coverage = (matched / len(benchmark_messages)) * 100 if len(benchmark_messages) > 0 else 0
            deviation = abs(len(model_messages) - len(benchmark_messages)) / len(benchmark_messages) * 100 if len(benchmark_messages) > 0 else 0
            
            detailed_results.append({
                'model_id': model_id,
                'model_title': model_cluster.get('draft_title', 'N/A'),
                'model_messages': len(model_messages),
                'benchmark_id': best_benchmark_id,
                'benchmark_title': best_match['draft_title'],
                'benchmark_messages': len(benchmark_messages),
                'matched_messages': matched,
                'missing_messages': missing,
                'extra_messages': extra,
                'precision': precision,
                'coverage': coverage,
                'deviation': deviation,
                'perfect_match': matched == len(benchmark_messages) and len(model_messages) == len(benchmark_messages)
            })
    
    # Unmatched benchmarks
    matched_ids = {r['benchmark_id'] for r in detailed_results}
    for bc in benchmark_clusters:
        if bc['cluster_id'] not in matched_ids:
            detailed_results.append({
                'model_id': "No Model Match", 'model_title': "No Model Match", 'model_messages': 0,
                'benchmark_id': bc['cluster_id'], 'benchmark_title': bc['draft_title'],
                'benchmark_messages': len(bc['message_ids']), 'matched_messages': 0,
                'missing_messages': len(bc['message_ids']), 'extra_messages': 0,
                'precision': 0, 'coverage': 0, 'deviation': 100, 'perfect_match': False
            })
    
    # Overall metrics
    total_benchmark = sum(len(c['message_ids']) for c in benchmark_clusters)
    total_model = sum(len(c.get('message_ids', [])) for c in model_clusters)
    total_matched = sum(r['matched_messages'] for r in detailed_results)
    
    cluster_score = min(len(model_clusters), len(benchmark_clusters)) / max(len(model_clusters), len(benchmark_clusters)) * 100
    coverage_score = (total_matched / total_benchmark) * 100
    precision_score = (total_matched / total_model) * 100 if total_model > 0 else 0
    avg_deviation = sum(r['deviation'] for r in detailed_results) / len(detailed_results) if detailed_results else 0
    deviation_score = max(0, 100 - avg_deviation)
    final_score = (cluster_score + coverage_score + precision_score + deviation_score) / 4
    
    # Token data from JSON
    usage = model_data.get('usage', {})
    input_tokens = usage.get('prompt_tokens', 0)
    output_tokens = usage.get('completion_tokens', 0)
    reported_total = usage.get('total_tokens', 0)
    
    # Calculate the correct total (use calculated total, not reported if inconsistent)
    calculated_total = input_tokens + output_tokens
    total_tokens = calculated_total  # Use calculated total as the authoritative value
    
    # Get correct pricing
    pricing = get_model_pricing(model_name)
    cost_per_input = pricing['COST_PER_INPUT_TOKEN']
    cost_per_output = pricing['COST_PER_OUTPUT_TOKEN']
    
    # Calculate TOKEN_COST using correct formula
    # TOKEN_COST = (INPUT_TOKENS × COST_PER_INPUT_TOKEN) + (OUTPUT_TOKENS × COST_PER_OUTPUT_TOKEN)
    token_cost = (input_tokens * cost_per_input) + (output_tokens * cost_per_output)
    
    return {
        'model_name': model_name,
        'model_cluster_count': len(model_clusters),
        'final_score': final_score,
        'cluster_count_score': cluster_score,
        'coverage_score': coverage_score,
        'precision_score': precision_score,
        'deviation_score': deviation_score,
        'avg_deviation': avg_deviation,
        'detailed_results': detailed_results,
        'merge_operations': len(model_data.get('ai_operations', {}).get('merge_operations', [])),
        'split_operations': len(model_data.get('ai_operations', {}).get('split_operations', [])),
        'duration': model_data.get('duration', 0),
        # Correct token data
        'INPUT_TOKENS': input_tokens,
        'OUTPUT_TOKENS': output_tokens,
        'TOTAL_TOKENS': total_tokens,
        'REPORTED_TOTAL': reported_total,
        'CALCULATED_TOTAL': calculated_total,
        'COST_PER_INPUT_TOKEN': cost_per_input,
        'COST_PER_OUTPUT_TOKEN': cost_per_output,
        'TOKEN_COST': token_cost
    }

def create_excel_final(metrics, output_dir):
    """Create Excel with correct token format"""
    
    if not metrics:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Step 2 Analysis"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="000000")
    data_font = Font(size=10)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    light_green_fill = PatternFill(start_color="E7F7E7", end_color="E7F7E7", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Step 2 Analysis: {metrics['model_name']}"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_align
    row += 2
    
    # Performance Summary
    ws[f'A{row}'] = "PERFORMANCE SUMMARY"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    summary = [
        ["FINAL SCORE", f"{metrics['final_score']:.2f}"],
        ["Model Clusters", metrics['model_cluster_count']],
        ["Coverage Score", f"{metrics['coverage_score']:.2f}%"],
        ["Precision Score", f"{metrics['precision_score']:.2f}%"],
        ["Deviation Score", f"{metrics['deviation_score']:.2f}%"]
    ]
    
    for item in summary:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].fill = light_blue_fill if col == 'A' else light_green_fill
            ws[f'{col}{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # TOKEN COST CALCULATION
    ws[f'A{row}'] = "TOKEN COST CALCULATION"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Headers
    token_headers = ["INPUT_TOKENS", "OUTPUT_TOKENS", "COST_PER_INPUT_TOKEN", "COST_PER_OUTPUT_TOKEN", "TOKEN_COST"]
    for i, header in enumerate(token_headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_align
    row += 1
    
    # Values
    token_values = [
        metrics['INPUT_TOKENS'],
        metrics['OUTPUT_TOKENS'],
        metrics['COST_PER_INPUT_TOKEN'],
        metrics['COST_PER_OUTPUT_TOKEN'],
        f"{metrics['TOKEN_COST']:.10f}"
    ]
    
    for i, value in enumerate(token_values):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = value
        ws[f'{col}{row}'].font = data_font
        ws[f'{col}{row}'].fill = light_yellow_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_align
    row += 1
    
    # Formula and verification
    ws[f'A{row}'] = f"Formula: TOKEN_COST = (INPUT_TOKENS × COST_PER_INPUT_TOKEN) + (OUTPUT_TOKENS × COST_PER_OUTPUT_TOKEN)"
    ws[f'A{row}'].font = Font(size=9, italic=True)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = f"Verification: ({metrics['INPUT_TOKENS']} × {metrics['COST_PER_INPUT_TOKEN']}) + ({metrics['OUTPUT_TOKENS']} × {metrics['COST_PER_OUTPUT_TOKEN']}) = ${metrics['TOKEN_COST']:.10f}"
    ws[f'A{row}'].font = Font(size=9, italic=True)
    ws[f'A{row}'].fill = light_blue_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = f"Total Tokens Check: {metrics['INPUT_TOKENS']} + {metrics['OUTPUT_TOKENS']} = {metrics['CALCULATED_TOTAL']} (Reported: {metrics['REPORTED_TOTAL']})"
    ws[f'A{row}'].font = Font(size=9, italic=True)
    ws[f'A{row}'].fill = light_green_fill if metrics['CALCULATED_TOTAL'] == metrics['REPORTED_TOTAL'] else light_red_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 2
    
    # Individual Topic Performance (condensed)
    ws[f'A{row}'] = "INDIVIDUAL TOPIC PERFORMANCE"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:L{row}')
    row += 1
    
    headers = ["#", "Model Title", "Model", "Benchmark Title", "Bench", 
               "Matched", "Missing", "Extra", "Coverage", "Precision", "Deviation", "Perfect"]
    
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_align
    row += 1
    
    for i, result in enumerate(metrics['detailed_results'], 1):
        row_data = [
            i, result['model_title'][:25], result['model_messages'],
            result['benchmark_title'][:25], result['benchmark_messages'],
            result['matched_messages'], result['missing_messages'], result['extra_messages'],
            f"{result['coverage']:.1f}%", f"{result['precision']:.1f}%",
            f"{result['deviation']:.1f}%", "Y" if result['perfect_match'] else "N"
        ]
        
        for j, val in enumerate(row_data):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = val
            ws[f'{col}{row}'].font = Font(size=9)
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].fill = light_green_fill if result['perfect_match'] else (
                light_red_fill if result['model_messages'] == 0 else light_yellow_fill)
            ws[f'{col}{row}'].alignment = center_align if j in [0, 2, 4, 5, 6, 7] else left_align
        row += 1
    
    # Column widths
    widths = {'A': 6, 'B': 28, 'C': 8, 'D': 28, 'E': 8, 'F': 9, 'G': 9, 'H': 9, 'I': 10, 'J': 10, 'K': 10, 'L': 8}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Save
    model_safe = metrics['model_name'].replace('/', '_').replace(' ', '_')
    output_file = f"{output_dir}/{model_safe}_step2_analysis.xlsx"
    wb.save(output_file)
    print(f"  Created: {output_file}")

def main():
    """Process all models with correct pricing"""
    
    print("=== CREATING FINAL CORRECT STEP 2 ANALYSES ===\n")
    
    benchmark_data = load_benchmark_clusters()
    model_files = sorted([f for f in glob.glob("output/phase4_balanced_refinement/*_balanced.json") 
                         if not f.endswith('comprehensive_results.json')])
    
    output_dir = "output/step2_client_analysis"
    os.makedirs(output_dir, exist_ok=True)
    
    all_metrics = []
    for model_file in model_files:
        model_name = os.path.basename(model_file).replace('_balanced.json', '')
        print(f"Processing: {model_name}")
        
        try:
            with open(model_file, 'r') as f:
                model_data = json.load(f)
            
            metrics = calculate_metrics_with_correct_pricing(model_data, benchmark_data, model_name)
            if metrics:
                all_metrics.append(metrics)
                create_excel_final(metrics, output_dir)
        except Exception as e:
            print(f"  ERROR: {e}")
    
    # Create CSV
    print(f"\n=== CREATING FINAL CSV ===\n")
    
    comparison_data = []
    for m in all_metrics:
        comparison_data.append({
            'Rank': 0,
            'Model Name': m['model_name'],
            'Final Score': round(m['final_score'], 2),
            'INPUT_TOKENS': m['INPUT_TOKENS'],
            'OUTPUT_TOKENS': m['OUTPUT_TOKENS'],
            'TOTAL_TOKENS': m['TOTAL_TOKENS'],
            'COST_PER_INPUT_TOKEN': m['COST_PER_INPUT_TOKEN'],
            'COST_PER_OUTPUT_TOKEN': m['COST_PER_OUTPUT_TOKEN'],
            'TOKEN_COST': f"{m['TOKEN_COST']:.10f}",
            'Duration (sec)': round(m['duration'], 2)
        })
    
    comparison_data.sort(key=lambda x: x['Final Score'], reverse=True)
    for i, data in enumerate(comparison_data, 1):
        data['Rank'] = i
    
    df = pd.DataFrame(comparison_data)
    df = df[['Rank', 'Model Name', 'Final Score', 'INPUT_TOKENS', 'OUTPUT_TOKENS', 'TOTAL_TOKENS',
             'COST_PER_INPUT_TOKEN', 'COST_PER_OUTPUT_TOKEN', 'TOKEN_COST', 'Duration (sec)']]
    
    csv_file = f"{output_dir}/step2_all_models_FINAL_CORRECT.csv"
    df.to_csv(csv_file, index=False)
    print(f"Created: {csv_file}\n")
    
    print("=== TOP 5 MODELS ===")
    for i, row in df.head(5).iterrows():
        print(f"{row['Rank']}. {row['Model Name']}: {row['Final Score']} (${row['TOKEN_COST']}, {row['TOTAL_TOKENS']} tokens)")
    
    print(f"\n=== COMPLETE ===")
    print(f"Generated {len(all_metrics)} Excel files with CORRECT pricing")

if __name__ == "__main__":
    main()


