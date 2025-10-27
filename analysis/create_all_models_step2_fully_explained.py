#!/usr/bin/env python3
"""
Create Step 2 Analysis for ALL models in FULLY_EXPLAINED format
Matches the format of gemini_2.0_flash_001_step2_analysis_FULLY_EXPLAINED.xlsx

Data sources:
- Benchmark: phases/phase4_clusters_refined.json
- Model outputs: output/phase4_balanced_refinement/*_balanced.json
- Step 1 outputs: output/phase3_topic_clustering/*.json (for token data)
"""

import json
import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_benchmark_clusters():
    """Load Phase 4 refined benchmark clusters (15 clusters)"""
    file_path = "phases/phase4_clusters_refined.json"
    with open(file_path, 'r') as f:
        data = json.load(f)
    print(f"✅ Loaded {len(data)} benchmark clusters from {file_path}")
    return data

def load_model_step2_output(model_file):
    """Load model's Step 2 output from phase4_balanced_refinement"""
    with open(model_file, 'r') as f:
        data = json.load(f)
    return data

def load_model_step1_output(model_name):
    """Load model's Step 1 output for token data"""
    # Try to find the corresponding Step 1 file
    step1_dir = "output/phase3_topic_clustering"
    step1_file = os.path.join(step1_dir, f"{model_name}.json")
    
    if os.path.exists(step1_file):
        with open(step1_file, 'r') as f:
            return json.load(f)
    return None

def extract_model_name_from_file(file_path):
    """Extract model name from file path"""
    # e.g., google_gemini-2.0-flash-001_balanced.json -> google_gemini-2.0-flash-001
    basename = os.path.basename(file_path)
    return basename.replace('_balanced.json', '')

def get_token_usage(step1_data, step2_data):
    """Extract and combine token usage from Step 1 and Step 2"""
    step1_usage = step1_data.get('usage', {}) if step1_data else {}
    step2_usage = step2_data.get('usage', {})
    
    # Step 1 tokens
    step1_prompt = step1_usage.get('prompt_tokens', 0)
    step1_completion = step1_usage.get('completion_tokens', 0)
    step1_total = step1_usage.get('total_tokens', 0)
    
    # Step 2 tokens
    step2_prompt = step2_usage.get('prompt_tokens', 0)
    step2_completion = step2_usage.get('completion_tokens', 0)
    step2_total = step2_usage.get('total_tokens', 0)
    
    # Handle special cases like GPT-5 where prompt_tokens and completion_tokens are 0
    # but total_tokens has the actual value
    if step1_prompt == 0 and step1_completion == 0 and step1_total > 0:
        # Use total_tokens as authoritative source
        step1_prompt = step1_total
        step1_completion = 0
    
    if step2_prompt == 0 and step2_completion == 0 and step2_total > 0:
        # Use total_tokens as authoritative source
        step2_prompt = step2_total
        step2_completion = 0
    
    # Combined
    total_input = step1_prompt + step2_prompt
    total_output = step1_completion + step2_completion
    total_tokens = total_input + total_output
    
    return {
        'INPUT_TOKENS': total_input,
        'OUTPUT_TOKENS': total_output,
        'TOTAL_TOKENS': total_tokens,
        'step1_input': step1_prompt,
        'step1_output': step1_completion,
        'step2_input': step2_prompt,
        'step2_output': step2_completion
    }

def get_pricing(model_name):
    """Get pricing for model - returns cost per 1000 tokens (divide by 1000 from per 1M pricing)"""
    # Official Google AI pricing as of October 2024
    # Source: https://ai.google.dev/pricing
    pricing_map = {
        # Gemini 1.5 models
        'gemini-1.5-flash': (0.000075, 0.0003),           # $0.075/$0.30 per 1M tokens
        'gemini-1.5-flash-002': (0.000075, 0.0003),
        'gemini-1.5-flash-8b': (0.0000375, 0.00015),      # $0.0375/$0.15 per 1M tokens
        'gemini-1.5-flash-latest': (0.000075, 0.0003),
        'gemini-1.5-pro': (0.00125, 0.005),               # $1.25/$5.00 per 1M tokens
        'gemini-1.5-pro-002': (0.00125, 0.005),
        'gemini-1.5-pro-latest': (0.00125, 0.005),
        
        # Gemini 2.0 models
        'gemini-2.0-flash': (0.0001, 0.0004),             # $0.10/$0.40 per 1M tokens
        'gemini-2.0-flash-001': (0.0001, 0.0004),
        'gemini-2.0-flash-lite': (0.000075, 0.0003),      # $0.075/$0.30 per 1M tokens
        'gemini-2.0-flash-lite-001': (0.000075, 0.0003),
        
        # Gemini 2.5 models - UPDATED PRICING (order matters: check specific names first!)
        'gemini-2.5-flash-lite': (0.0001, 0.0004),        # $0.10/$0.40 per 1M tokens
        'gemini-2.5-flash': (0.0003, 0.0025),             # $0.30/$2.50 per 1M tokens (CORRECTED)
        'gemini-2.5-pro': (0.00125, 0.01),                # $1.25/$10.00 per 1M tokens (CORRECTED)
        
        # Gemma models (free)
        'gemma-3-1b-it': (0.0, 0.0),
        'gemma-3-4b-it': (0.0, 0.0),
        'gemma-3-12b-it': (0.0, 0.0),
        'gemma-3-27b-it': (0.0, 0.0),
        'gemma-3n-e2b-it': (0.0, 0.0),
        'gemma-3n-e4b-it': (0.0, 0.0),
        
        # OpenAI models
        'gpt-4o': (0.0025, 0.01),
        'gpt-5': (0.00125, 0.01),                         # $1.25/$10.00 per 1M tokens
        
        # xAI models
        'grok-2-1212': (0.002, 0.01),
        'grok-2-vision-1212': (0.002, 0.01),
        'grok-3': (0.002, 0.01),
    }
    
    # Try to match model name - check most specific names first (longer strings)
    model_lower = model_name.lower()
    sorted_keys = sorted(pricing_map.keys(), key=len, reverse=True)
    
    for key in sorted_keys:
        if key in model_lower:
            # Return cost per 1000 tokens
            return pricing_map[key]
    
    return (0.0, 0.0)

def calculate_metrics(model_clusters, benchmark_clusters):
    """Calculate detailed metrics comparing model vs benchmark"""
    
    # Build message ID mappings
    model_msg_map = {}
    for cluster in model_clusters:
        cluster_id = cluster.get('cluster_id', 'unknown')
        msg_ids = set(cluster.get('message_ids', []))
        model_msg_map[cluster_id] = {
            'messages': msg_ids,
            'title': cluster.get('draft_title', 'N/A'),
            'size': len(msg_ids)
        }
    
    benchmark_msg_map = {}
    for cluster in benchmark_clusters:
        cluster_id = cluster['cluster_id']
        msg_ids = set(cluster['message_ids'])
        benchmark_msg_map[cluster_id] = {
            'messages': msg_ids,
            'title': cluster['draft_title'],
            'size': len(msg_ids)
        }
    
    # Match each model cluster to best benchmark cluster
    detailed_results = []
    matched_benchmark_ids = set()
    
    for model_id, model_data in model_msg_map.items():
        model_msgs = model_data['messages']
        
        # Find best match
        best_match_id = None
        best_overlap = 0
        
        for bench_id, bench_data in benchmark_msg_map.items():
            bench_msgs = bench_data['messages']
            overlap = len(model_msgs & bench_msgs)
            
            if overlap > best_overlap:
                best_overlap = overlap
                best_match_id = bench_id
        
        # Calculate metrics for best match
        if best_match_id:
            matched_benchmark_ids.add(best_match_id)
            bench_data = benchmark_msg_map[best_match_id]
            bench_msgs = bench_data['messages']
            
            matched = len(model_msgs & bench_msgs)
            missing = len(bench_msgs - model_msgs)
            extra = len(model_msgs - bench_msgs)
            
            coverage = (matched / len(bench_msgs) * 100) if len(bench_msgs) > 0 else 0
            precision = (matched / len(model_msgs) * 100) if len(model_msgs) > 0 else 0
            deviation = abs(len(model_msgs) - len(bench_msgs)) / len(bench_msgs) * 100 if len(bench_msgs) > 0 else 0
            
            detailed_results.append({
                'model_id': model_id,
                'model_title': model_data['title'],
                'model_size': model_data['size'],
                'benchmark_id': best_match_id,
                'benchmark_title': bench_data['title'],
                'benchmark_size': bench_data['size'],
                'matched': matched,
                'missing': missing,
                'extra': extra,
                'coverage': coverage,
                'precision': precision,
                'deviation': deviation,
                'perfect_match': 'YES' if (matched == len(bench_msgs) and len(model_msgs) == len(bench_msgs)) else 'NO'
            })
        else:
            # No match found
            detailed_results.append({
                'model_id': model_id,
                'model_title': model_data['title'],
                'model_size': model_data['size'],
                'benchmark_id': 'No Match',
                'benchmark_title': 'No Match',
                'benchmark_size': 0,
                'matched': 0,
                'missing': 0,
                'extra': model_data['size'],
                'coverage': 0,
                'precision': 0,
                'deviation': 100,
                'perfect_match': 'NO'
            })
    
    # Add unmatched benchmark clusters
    for bench_id, bench_data in benchmark_msg_map.items():
        if bench_id not in matched_benchmark_ids:
            detailed_results.append({
                'model_id': 'Missing',
                'model_title': 'Missing',
                'model_size': 0,
                'benchmark_id': bench_id,
                'benchmark_title': bench_data['title'],
                'benchmark_size': bench_data['size'],
                'matched': 0,
                'missing': bench_data['size'],
                'extra': 0,
                'coverage': 0,
                'precision': 0,
                'deviation': 100,
                'perfect_match': 'NO'
            })
    
    # Calculate overall scores
    total_benchmark_msgs = sum(len(d['messages']) for d in benchmark_msg_map.values())
    total_model_msgs = sum(len(d['messages']) for d in model_msg_map.values())
    total_matched = sum(r['matched'] for r in detailed_results)
    
    overall_coverage = (total_matched / total_benchmark_msgs * 100) if total_benchmark_msgs > 0 else 0
    overall_precision = (total_matched / total_model_msgs * 100) if total_model_msgs > 0 else 0
    avg_deviation = sum(r['deviation'] for r in detailed_results) / len(detailed_results) if detailed_results else 0
    
    cluster_count_score = min(len(model_clusters), len(benchmark_clusters)) / max(len(model_clusters), len(benchmark_clusters)) * 100
    deviation_score = max(0, 100 - avg_deviation)
    
    combined_score = (
        overall_coverage * 0.25 +
        overall_precision * 0.25 +
        cluster_count_score * 0.25 +
        deviation_score * 0.25
    )
    
    return {
        'model_cluster_count': len(model_clusters),
        'benchmark_cluster_count': len(benchmark_clusters),
        'total_benchmark_messages': total_benchmark_msgs,
        'total_model_messages': total_model_msgs,
        'total_matched_messages': total_matched,
        'overall_coverage': overall_coverage,
        'overall_precision': overall_precision,
        'avg_deviation': avg_deviation,
        'cluster_count_score': cluster_count_score,
        'deviation_score': deviation_score,
        'combined_score': combined_score,
        'detailed_results': detailed_results
    }

def create_excel_file(model_name, metrics, token_data, output_path):
    """Create fully explained Excel file for a model"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Step 2 Analysis"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=11)
    data_font = Font(size=10)
    formula_font = Font(size=9, italic=True, color="666666")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = f"STEP 2 ANALYSIS: {model_name}"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_align
    row += 2
    
    # Data Sources
    ws[f'A{row}'] = "DATA SOURCES"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = "Benchmark File:"
    ws[f'B{row}'] = "phases/phase4_clusters_refined.json"
    ws[f'A{row}'].font = data_font
    ws[f'B{row}'].font = data_font
    row += 1
    
    ws[f'A{row}'] = "Model Output File:"
    ws[f'B{row}'] = f"output/phase4_balanced_refinement/{model_name}_balanced.json"
    ws[f'A{row}'].font = data_font
    ws[f'B{row}'].font = data_font
    row += 2
    
    # Token Usage Summary
    ws[f'A{row}'] = "TOKEN USAGE & COST ANALYSIS"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    cost_per_input = get_pricing(model_name)[0]
    cost_per_output = get_pricing(model_name)[1]
    input_tokens = token_data['INPUT_TOKENS']
    output_tokens = token_data['OUTPUT_TOKENS']
    total_tokens = token_data['TOTAL_TOKENS']
    # Pricing is per 1000 tokens, so divide by 1000
    token_cost = (input_tokens / 1000 * cost_per_input) + (output_tokens / 1000 * cost_per_output)
    
    token_info = [
        ["INPUT_TOKENS", f"{input_tokens:,}", f"Step 1: {token_data['step1_input']:,} + Step 2: {token_data['step2_input']:,}"],
        ["OUTPUT_TOKENS", f"{output_tokens:,}", f"Step 1: {token_data['step1_output']:,} + Step 2: {token_data['step2_output']:,}"],
        ["TOTAL_TOKENS", f"{total_tokens:,}", f"{input_tokens:,} + {output_tokens:,}"],
        ["COST_PER_INPUT_TOKEN", f"${cost_per_input:.10f}", "Per 1000 tokens"],
        ["COST_PER_OUTPUT_TOKEN", f"${cost_per_output:.10f}", "Per 1000 tokens"],
        ["TOKEN_COST", f"${token_cost:.10f}", f"({input_tokens:,}/1000 × ${cost_per_input:.10f}) + ({output_tokens:,}/1000 × ${cost_per_output:.10f})"]
    ]
    
    for item in token_info:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'C{row}'] = item[2]
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = data_font
        ws[f'C{row}'].font = formula_font
        ws[f'A{row}'].fill = blue_fill
        ws[f'B{row}'].fill = green_fill if "COST" in item[0] else blue_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # Metrics Summary
    ws[f'A{row}'] = "PERFORMANCE METRICS"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    metrics_info = [
        ["Model Clusters", metrics['model_cluster_count']],
        ["Benchmark Clusters", metrics['benchmark_cluster_count']],
        ["Total Benchmark Messages", metrics['total_benchmark_messages']],
        ["Total Model Messages", metrics['total_model_messages']],
        ["Total Matched Messages", metrics['total_matched_messages']],
        ["Overall Coverage", f"{metrics['overall_coverage']:.2f}%"],
        ["Overall Precision", f"{metrics['overall_precision']:.2f}%"],
        ["Average Deviation", f"{metrics['avg_deviation']:.2f}%"],
        ["Cluster Count Score", f"{metrics['cluster_count_score']:.2f}%"],
        ["Deviation Score", f"{metrics['deviation_score']:.2f}%"],
        ["COMBINED SCORE", f"{metrics['combined_score']:.2f}"]
    ]
    
    for item in metrics_info:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = Font(bold=True, size=11) if "COMBINED" in item[0] else data_font
        ws[f'A{row}'].fill = blue_fill
        ws[f'B{row}'].fill = green_fill if "COMBINED" in item[0] else yellow_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # Formula Explanation
    ws[f'A{row}'] = "SCORE CALCULATION FORMULAS"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    formulas = [
        ["Coverage", f"= {metrics['total_matched_messages']} / {metrics['total_benchmark_messages']} × 100 = {metrics['overall_coverage']:.2f}%"],
        ["Precision", f"= {metrics['total_matched_messages']} / {metrics['total_model_messages']} × 100 = {metrics['overall_precision']:.2f}%"],
        ["Cluster Count Score", f"= min({metrics['model_cluster_count']}, {metrics['benchmark_cluster_count']}) / max({metrics['model_cluster_count']}, {metrics['benchmark_cluster_count']}) × 100 = {metrics['cluster_count_score']:.2f}%"],
        ["Deviation Score", f"= 100 - {metrics['avg_deviation']:.2f} = {metrics['deviation_score']:.2f}%"],
        ["COMBINED SCORE", f"= (Coverage × 0.25) + (Precision × 0.25) + (Cluster Count × 0.25) + (Deviation × 0.25)"]
    ]
    
    for item in formulas:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = formula_font
        ws[f'A{row}'].fill = blue_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws.merge_cells(f'B{row}:D{row}')
        row += 1
    
    row += 2
    
    # Detailed Cluster Comparison Table
    ws[f'A{row}'] = "DETAILED CLUSTER COMPARISON"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:K{row}')
    row += 1
    
    # Table headers
    headers = ["Model ID", "Model Title", "Model Msgs", "Benchmark ID", "Benchmark Title", 
               "Benchmark Msgs", "Matched", "Missing", "Extra", "Coverage %", "Precision %", 
               "Deviation %", "Perfect Match"]
    
    for col_idx, header in enumerate(headers):
        col = get_column_letter(col_idx + 1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].border = thin_border
    
    row += 1
    
    # Table data
    for result in metrics['detailed_results']:
        data_row = [
            result['model_id'],
            result['model_title'],
            result['model_size'],
            result['benchmark_id'],
            result['benchmark_title'],
            result['benchmark_size'],
            result['matched'],
            result['missing'],
            result['extra'],
            f"{result['coverage']:.2f}%",
            f"{result['precision']:.2f}%",
            f"{result['deviation']:.2f}%",
            result['perfect_match']
        ]
        
        for col_idx, value in enumerate(data_row):
            col = get_column_letter(col_idx + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].alignment = center_align if col_idx in [0, 3, 6, 7, 8, 9, 10, 11, 12] else left_align
            ws[f'{col}{row}'].border = thin_border
            
            # Color coding
            if result['perfect_match'] == 'YES':
                ws[f'{col}{row}'].fill = green_fill
            elif result['model_id'] == 'Missing':
                ws[f'{col}{row}'].fill = red_fill
            elif result['coverage'] < 50:
                ws[f'{col}{row}'].fill = yellow_fill
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    
    # Save
    wb.save(output_path)
    print(f"  ✅ Created: {output_path}")

def main():
    """Generate Step 2 analysis for all models"""
    
    print("=" * 80)
    print("CREATING STEP 2 ANALYSIS FOR ALL MODELS (FULLY EXPLAINED FORMAT)")
    print("=" * 80)
    print()
    
    # Load benchmark
    print("📂 Loading benchmark clusters...")
    benchmark_clusters = load_benchmark_clusters()
    print()
    
    # Find all model output files
    print("📂 Finding model output files...")
    model_files = glob.glob("output/phase4_balanced_refinement/*_balanced.json")
    model_files = [f for f in model_files if os.path.basename(f) != 'comprehensive_results.json']
    model_files.sort()
    print(f"   Found {len(model_files)} model output files")
    print()
    
    # Create output directory
    output_dir = "output/step2_client_analysis"
    os.makedirs(output_dir, exist_ok=True)
    
    # Process each model
    all_results = []
    
    for i, model_file in enumerate(model_files, 1):
        model_name = extract_model_name_from_file(model_file)
        print(f"[{i}/{len(model_files)}] Processing: {model_name}")
        
        try:
            # Load model data
            step2_data = load_model_step2_output(model_file)
            step1_data = load_model_step1_output(model_name)
            
            # Check if model has refined clusters
            model_clusters = step2_data.get('refined_clusters', [])
            if not model_clusters:
                print(f"  ⚠️  No refined_clusters found in {model_file}, skipping")
                continue
            
            # Get token usage
            token_data = get_token_usage(step1_data, step2_data)
            
            # Calculate metrics
            metrics = calculate_metrics(model_clusters, benchmark_clusters)
            
            # Create Excel file
            safe_name = model_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            output_file = os.path.join(output_dir, f"{safe_name}_step2_FULLY_EXPLAINED.xlsx")
            create_excel_file(model_name, metrics, token_data, output_file)
            
            # Store for comparison CSV
            cost_per_input, cost_per_output = get_pricing(model_name)
            # Pricing is per 1000 tokens, so divide by 1000
            token_cost = (token_data['INPUT_TOKENS'] / 1000 * cost_per_input) + (token_data['OUTPUT_TOKENS'] / 1000 * cost_per_output)
            
            all_results.append({
                'Model Name': model_name,
                'Combined Score': f"{metrics['combined_score']:.2f}",
                'Coverage %': f"{metrics['overall_coverage']:.2f}",
                'Precision %': f"{metrics['overall_precision']:.2f}",
                'Cluster Count Score %': f"{metrics['cluster_count_score']:.2f}",
                'Deviation Score %': f"{metrics['deviation_score']:.2f}",
                'Model Clusters': metrics['model_cluster_count'],
                'Benchmark Clusters': metrics['benchmark_cluster_count'],
                'Matched Messages': metrics['total_matched_messages'],
                'INPUT_TOKENS': token_data['INPUT_TOKENS'],
                'OUTPUT_TOKENS': token_data['OUTPUT_TOKENS'],
                'TOTAL_TOKENS': token_data['TOTAL_TOKENS'],
                'COST_PER_INPUT_TOKEN': cost_per_input,
                'COST_PER_OUTPUT_TOKEN': cost_per_output,
                'TOKEN_COST': f"{token_cost:.10f}"
            })
            
        except Exception as e:
            print(f"  ❌ Error processing {model_name}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print()
    print("=" * 80)
    print("CREATING COMPARISON CSV")
    print("=" * 80)
    
    # Create comparison CSV
    if all_results:
        df = pd.DataFrame(all_results)
        df = df.sort_values('Combined Score', ascending=False)
        csv_file = os.path.join(output_dir, "step2_all_models_comparison_FULLY_EXPLAINED.csv")
        df.to_csv(csv_file, index=False)
        print(f"✅ Created comparison CSV: {csv_file}")
        print()
        
        # Show top 5
        print("🏆 TOP 5 MODELS:")
        for i, row in df.head(5).iterrows():
            print(f"  {i+1}. {row['Model Name']}: {row['Combined Score']} (Coverage: {row['Coverage %']}%, Cost: ${row['TOKEN_COST']})")
    
    print()
    print("=" * 80)
    print(f"✅ COMPLETE! Generated {len(all_results)} Excel files")
    print(f"   Output directory: {output_dir}")
    print("=" * 80)

if __name__ == "__main__":
    main()

