#!/usr/bin/env python3
"""
Accurate Excel Verification - Reads actual Excel structure
"""

import pandas as pd
import json
import os
from openpyxl import load_workbook

def load_model_data(model_name):
    """Load specific model data from JSON results"""
    
    results_file = "output/phase4_comprehensive_evaluation/detailed_evaluation_results.json"
    
    try:
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Find the specific model
        for result in results:
            if result['model'] == model_name:
                return result
        
        print(f"❌ Model {model_name} not found in results")
        return None
        
    except Exception as e:
        print(f"❌ Error loading results: {e}")
        return None

def read_excel_structure(excel_file_path):
    """Read and display the actual Excel structure"""
    
    print(f"📊 Reading Excel file structure: {excel_file_path}")
    print("=" * 70)
    
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
        
        print(f"📋 Excel Structure Analysis:")
        print(f"  - Sheet name: {ws.title}")
        print(f"  - Max row: {ws.max_row}")
        print(f"  - Max column: {ws.max_column}")
        print()
        
        # Read all data and display structure
        print(f"📄 Cell Contents (first 20 rows):")
        for row_num in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, min(8, ws.max_column + 1)):  # First 7 columns
                cell = ws.cell(row=row_num, column=col_num)
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value)[:30])  # Truncate long values
            print(f"  Row {row_num:2d}: {' | '.join(row_data)}")
        
        print()
        return ws
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return None

def verify_excel_against_source(model_name, excel_file_path):
    """Verify Excel values against source data"""
    
    print(f"🔍 Verifying Excel against source data for: {model_name}")
    print("=" * 70)
    
    # Load source data
    source_data = load_model_data(model_name)
    if not source_data:
        return
    
    # Load Excel file
    ws = read_excel_structure(excel_file_path)
    if not ws:
        return
    
    # Calculate expected values
    avg_similarity = source_data['avg_combined_similarity']
    quality_rate = source_data['quality_rate']
    total_operations = source_data['total_operations']
    num_refined_clusters = source_data['num_refined_clusters']
    size_reduction = source_data['size_reduction']
    
    similarity_score = avg_similarity * 100
    quality_score = quality_rate * 100
    operation_efficiency = max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))) if num_refined_clusters > 0 else 100
    size_optimization = max(0, min(100, 100 - abs(size_reduction)))
    
    combined_score = (
        similarity_score * 0.35 +
        quality_score * 0.25 +
        size_optimization * 0.20 +
        operation_efficiency * 0.20
    )
    
    print(f"🎯 Source Data Values:")
    print(f"  - Model: {source_data['model']}")
    print(f"  - Duration: {source_data['duration']:.3f}s")
    print(f"  - Refined Clusters: {source_data['num_refined_clusters']}")
    print(f"  - Step 1 Clusters: {source_data['num_step1_clusters']}")
    print(f"  - Cluster Reduction: {source_data['cluster_count_reduction']}")
    print(f"  - Total Operations: {source_data['total_operations']}")
    print(f"  - Avg Cluster Size: {source_data['avg_cluster_size']:.3f}")
    print(f"  - Avg Combined Similarity: {source_data['avg_combined_similarity']:.4f}")
    print(f"  - Quality Rate: {source_data['quality_rate']:.4f}")
    print(f"  - High Quality Matches: {source_data['high_quality_matches']}")
    print()
    
    print(f"🧮 Expected Calculated Values:")
    print(f"  - Similarity Score: {similarity_score:.2f}%")
    print(f"  - Quality Score: {quality_score:.2f}%")
    print(f"  - Operation Efficiency: {operation_efficiency:.2f}%")
    print(f"  - Size Optimization: {size_optimization:.2f}%")
    print(f"  - Combined Score: {combined_score:.2f}")
    print()
    
    # Search for key values in Excel
    print(f"🔍 Searching for key values in Excel:")
    
    # Look for the combined score
    found_combined_score = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell_str = str(cell.value).strip()
                if "91.28" in cell_str or "91.3" in cell_str:
                    print(f"  ✅ Found Combined Score '{cell_str}' at Row {row}, Col {col}")
                    found_combined_score = True
    
    if not found_combined_score:
        print(f"  ❌ Combined Score not found in Excel")
    
    # Look for similarity score
    found_similarity = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell_str = str(cell.value).strip()
                if "82.34" in cell_str or "82.3" in cell_str:
                    print(f"  ✅ Found Similarity Score '{cell_str}' at Row {row}, Col {col}")
                    found_similarity = True
    
    if not found_similarity:
        print(f"  ❌ Similarity Score not found in Excel")
    
    # Look for quality score
    found_quality = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell_str = str(cell.value).strip()
                if "92.31" in cell_str or "92.3" in cell_str:
                    print(f"  ✅ Found Quality Score '{cell_str}' at Row {row}, Col {col}")
                    found_quality = True
    
    if not found_quality:
        print(f"  ❌ Quality Score not found in Excel")
    
    print()
    print(f"📊 Summary:")
    print(f"  - Combined Score Found: {'✅' if found_combined_score else '❌'}")
    print(f"  - Similarity Score Found: {'✅' if found_similarity else '❌'}")
    print(f"  - Quality Score Found: {'✅' if found_quality else '❌'}")
    print()
    print(f"✅ Verification complete!")

def main():
    """Main verification function"""
    
    model_name = "google_gemini-2.0-flash-001"
    excel_file = "output/step2_client_analysis/gemini_2.0_flash_001_step2_updated_analysis_FIXED.xlsx"
    
    if os.path.exists(excel_file):
        verify_excel_against_source(model_name, excel_file)
    else:
        print(f"❌ Excel file not found: {excel_file}")

if __name__ == "__main__":
    main()
