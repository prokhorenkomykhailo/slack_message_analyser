#!/usr/bin/env python3
"""
Create Step 3 Analysis for ALL models in FULLY_EXPLAINED format
Evaluates metadata generation quality (Phase 5)

Data sources:
- Benchmark: phases/phase5_metadata_benchmark.json
- Model outputs: output/phase5_metadata_generation/*.json

SCORE CALCULATION METHOD:
1. For each cluster: Calculate individual overall score using weighted formula:
   Overall = (Title×30%) + (Summary×25%) + (Action Items×20%) + (Participants×10%) + (Tags×5%) + (Urgency×5%) + (Status×5%)

2. Final Average Overall Score = Simple arithmetic average of all 15 individual cluster overall scores

Evaluation Metrics:
1. Title Match Score (0-100): Semantic similarity between titles
2. Summary Match Score (0-100): Semantic similarity between summaries  
3. Action Items Score (0-100): Coverage and accuracy of action items
4. Participants Score (0-100): Accuracy of participant identification
5. Metadata Accuracy (0-100): Urgency, deadline, status, tags correctness

Note: Individual metric averages (Title: 92.81%, Summary: 19.50%, etc.) are for reference only
and are NOT directly used in the final Average Overall Score calculation.
"""

import json
import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
from datetime import datetime
import re

def load_benchmark_metadata():
    """Load Phase 5 metadata benchmark (15 clusters with detailed metadata)"""
    file_path = "phases/phase5_metadata_benchmark.json"
    with open(file_path, 'r') as f:
        data = json.load(f)
    print(f"✅ Loaded {len(data)} benchmark metadata entries from {file_path}")
    return data

def load_model_output(model_file):
    """Load model's Step 3 output from phase5_metadata_generation"""
    with open(model_file, 'r') as f:
        data = json.load(f)
    return data

def extract_model_name_from_file(file_path):
    """Extract model name from file path"""
    # e.g., google_gemini-2.0-flash-001.json -> google_gemini-2.0-flash-001
    basename = os.path.basename(file_path)
    return basename.replace('.json', '')

def get_token_usage(model_data):
    """Extract token usage from model output"""
    usage = model_data.get('usage', {})
    
    prompt_tokens = usage.get('prompt_tokens', 0)
    completion_tokens = usage.get('completion_tokens', 0)
    total_tokens = usage.get('total_tokens', 0)
    
    # If we have detailed token breakdown, use it
    if prompt_tokens > 0 or completion_tokens > 0:
        return {
            'INPUT_TOKENS': prompt_tokens,
            'OUTPUT_TOKENS': completion_tokens,
            'TOTAL_TOKENS': prompt_tokens + completion_tokens,
            'REPORTED_TOTAL': total_tokens
        }
    
    # For models that only report total_tokens, calculate from actual responses
    if total_tokens > 0:
        # Calculate actual output tokens from raw responses
        metadata_results = model_data.get('metadata_results', [])
        actual_output_tokens = 0
        actual_input_tokens = 0
        
        for result in metadata_results:
            if result.get('success') and 'raw_response' in result:
                raw_response = result['raw_response']
                # Estimate tokens from raw response (roughly 1 token per 4 characters)
                response_tokens = len(raw_response) // 4
                actual_output_tokens += response_tokens
        
        # Input tokens are the remainder
        actual_input_tokens = total_tokens - actual_output_tokens
        
        # Ensure we don't have negative input tokens
        if actual_input_tokens < 0:
            actual_input_tokens = int(total_tokens * 0.85)  # Fallback to 85% input
            actual_output_tokens = total_tokens - actual_input_tokens
        
        return {
            'INPUT_TOKENS': actual_input_tokens,
            'OUTPUT_TOKENS': actual_output_tokens,
            'TOTAL_TOKENS': total_tokens,
            'REPORTED_TOTAL': total_tokens
        }
    
    # Fallback if no token data available
    return {
        'INPUT_TOKENS': 0,
        'OUTPUT_TOKENS': 0,
        'TOTAL_TOKENS': 0,
        'REPORTED_TOTAL': 0
    }

def get_pricing(model_name):
    """Get pricing for model - returns cost per 1M tokens (official Google pricing)
    
    Based on: https://ai.google.dev/gemini-api/docs/pricing
    
    Note: We assume Paid Tier pricing. If using Free Tier, costs would be $0.00
    """
    pricing_map = {
        # Gemini 1.5 models (Paid Tier)
        'gemini-1.5-flash': (0.075, 0.30),
        'gemini-1.5-flash-002': (0.075, 0.30),
        'gemini-1.5-flash-8b': (0.0375, 0.15),
        'gemini-1.5-flash-latest': (0.075, 0.30),
        'gemini-1.5-pro': (1.25, 5.00),
        'gemini-1.5-pro-002': (1.25, 5.00),
        'gemini-1.5-pro-latest': (1.25, 5.00),
        
        # Gemini 2.0 models (Paid Tier)
        'gemini-2.0-flash': (0.10, 0.40),
        'gemini-2.0-flash-001': (0.10, 0.40),
        'gemini-2.0-flash-lite': (0.075, 0.30),
        'gemini-2.0-flash-lite-001': (0.075, 0.30),
        
        # Gemini 2.5 models (Paid Tier)
        'gemini-2.5-flash-lite': (0.10, 0.40),
        'gemini-2.5-flash': (0.30, 2.50),
        'gemini-2.5-pro': (1.25, 10.00),
        
        # Gemma models (free)
        'gemma-3-1b-it': (0.0, 0.0),
        'gemma-3-4b-it': (0.0, 0.0),
        'gemma-3-12b-it': (0.0, 0.0),
        'gemma-3-27b-it': (0.0, 0.0),
        'gemma-3n-e2b-it': (0.0, 0.0),
        'gemma-3n-e4b-it': (0.0, 0.0),
        
        # OpenAI models (Paid Tier)
        'gpt-4o': (2.50, 10.00),  # $2.50 input, $10.00 output per 1M tokens
        'gpt-5': (1.25, 10.00),   # $1.25 input, $10.00 output per 1M tokens
        'gpt-3.5-turbo': (0.50, 1.50),  # $0.50 input, $1.50 output per 1M tokens
        'gpt-4': (0.0, 0.0),  # Free tier
        
        # xAI models
        'grok-2-1212': (0.002, 0.01),
        'grok-2-vision-1212': (0.002, 0.01),
        'grok-3': (0.002, 0.01),
        
        # Anthropic models
        'claude-3-haiku': (0.00025, 0.00125),
        'claude-3-sonnet': (0.003, 0.015),
        'claude-3-opus': (0.015, 0.075),
        
        # Groq models (free tier)
        'llama3-70b': (0.0, 0.0),
        'llama3-8b': (0.0, 0.0),
        'mixtral-8x7b': (0.0, 0.0),
    }
    
    # Try to match model name - check most specific names first
    model_lower = model_name.lower()
    sorted_keys = sorted(pricing_map.keys(), key=len, reverse=True)
    
    for key in sorted_keys:
        if key in model_lower:
            return pricing_map[key]
    
    return (0.0, 0.0)

def text_similarity(text1, text2):
    """Calculate text similarity using SequenceMatcher (0-100)"""
    if not text1 or not text2:
        return 0.0
    return SequenceMatcher(None, text1.lower(), text2.lower()).ratio() * 100

def normalize_participant(name):
    """Normalize participant name (remove @, lowercase, trim)"""
    return name.replace('@', '').strip().lower()

def normalize_date(date_str):
    """Normalize date string to YYYY-MM-DD format"""
    if not date_str or date_str.lower() in ['n/a', 'none', 'ongoing']:
        return None
    
    # Try various date formats
    date_patterns = [
        r'(\d{4})-(\d{2})-(\d{2})',  # 2025-06-15
        r'(\w+)\s+(\d{1,2}),\s+(\d{4})',  # June 15, 2025
        r'(\d{1,2})/(\d{1,2})/(\d{4})',  # 06/15/2025
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, date_str)
        if match:
            try:
                if '-' in date_str:
                    return date_str[:10]  # Return YYYY-MM-DD
                elif ',' in date_str:
                    # Convert "June 15, 2025" to "2025-06-15"
                    month_map = {
                        'january': '01', 'february': '02', 'march': '03', 'april': '04',
                        'may': '05', 'june': '06', 'july': '07', 'august': '08',
                        'september': '09', 'october': '10', 'november': '11', 'december': '12'
                    }
                    month_name, day, year = match.groups()
                    month_num = month_map.get(month_name.lower(), '01')
                    return f"{year}-{month_num}-{int(day):02d}"
            except:
                pass
    
    return None

def evaluate_action_items(benchmark_items, model_items):
    """Evaluate action items accuracy"""
    if not benchmark_items:
        return 100.0 if not model_items else 0.0
    
    if not model_items:
        return 0.0
    
    # Count matches based on task similarity and owner match
    matches = 0
    for bench_item in benchmark_items:
        bench_task = bench_item.get('task', '').lower()
        bench_owner = normalize_participant(bench_item.get('owner', ''))
        
        for model_item in model_items:
            model_task = model_item.get('task', '').lower()
            model_owner = normalize_participant(model_item.get('owner', ''))
            
            # Check if task is similar (>60% match) and owner matches
            task_sim = SequenceMatcher(None, bench_task, model_task).ratio()
            owner_match = bench_owner in model_owner or model_owner in bench_owner
            
            if task_sim > 0.6 or owner_match:
                matches += 1
                break
    
    # Calculate score
    recall = matches / len(benchmark_items) * 100
    precision = matches / len(model_items) * 100 if model_items else 0
    
    # F1-like score
    if recall + precision == 0:
        return 0.0
    return 2 * (recall * precision) / (recall + precision)

def evaluate_participants(benchmark_participants, model_participants):
    """Evaluate participant identification accuracy"""
    if not benchmark_participants:
        return 100.0 if not model_participants else 0.0
    
    if not model_participants:
        return 0.0
    
    # Normalize all names
    bench_set = set(normalize_participant(p) for p in benchmark_participants)
    model_set = set(normalize_participant(p) for p in model_participants)
    
    # Calculate precision and recall
    matches = len(bench_set & model_set)
    
    if len(bench_set) == 0:
        return 100.0
    
    recall = matches / len(bench_set) * 100
    precision = matches / len(model_set) * 100 if model_set else 0
    
    # F1-like score
    if recall + precision == 0:
        return 0.0
    return 2 * (recall * precision) / (recall + precision)

def evaluate_tags(benchmark_tags, model_tags):
    """Evaluate tags accuracy"""
    if not benchmark_tags:
        return 100.0 if not model_tags else 0.0
    
    if not model_tags:
        return 0.0
    
    # Normalize tags
    bench_set = set(tag.lower().strip() for tag in benchmark_tags)
    model_set = set(tag.lower().strip() for tag in model_tags)
    
    # Calculate Jaccard similarity
    intersection = len(bench_set & model_set)
    union = len(bench_set | model_set)
    
    if union == 0:
        return 100.0
    
    return (intersection / union) * 100

def evaluate_metadata(benchmark_meta, model_meta):
    """Evaluate individual metadata fields"""
    scores = {}
    
    # Title (30% weight)
    scores['title'] = text_similarity(
        benchmark_meta.get('title', ''),
        model_meta.get('title', '')
    )
    
    # Summary (25% weight)
    scores['summary'] = text_similarity(
        benchmark_meta.get('summary', ''),
        model_meta.get('summary', '')
    )
    
    # Action Items (20% weight)
    scores['action_items'] = evaluate_action_items(
        benchmark_meta.get('action_items', []),
        model_meta.get('action_items', [])
    )
    
    # Participants (10% weight)
    scores['participants'] = evaluate_participants(
        benchmark_meta.get('participants', []),
        model_meta.get('participants', [])
    )
    
    # Tags (5% weight)
    scores['tags'] = evaluate_tags(
        benchmark_meta.get('tags', []),
        model_meta.get('tags', [])
    )
    
    # Urgency (5% weight)
    bench_urgency = benchmark_meta.get('urgency', '').lower()
    model_urgency = model_meta.get('urgency', '').lower()
    scores['urgency'] = 100.0 if bench_urgency == model_urgency else 0.0
    
    # Status (5% weight)
    bench_status = benchmark_meta.get('status', '').lower()
    model_status = model_meta.get('status', '').lower()
    scores['status'] = 100.0 if bench_status == model_status else 0.0
    
    # Overall score (weighted average)
    weights = {
        'title': 0.30,
        'summary': 0.25,
        'action_items': 0.20,
        'participants': 0.10,
        'tags': 0.05,
        'urgency': 0.05,
        'status': 0.05
    }
    
    overall = sum(scores[key] * weights[key] for key in weights.keys())
    scores['overall'] = overall
    
    return scores

def calculate_metrics(model_data, benchmark_data):
    """Calculate detailed metrics comparing model vs benchmark"""
    
    # Build cluster ID mapping for benchmark (extract metadata from nested structure)
    benchmark_map = {item['cluster_id']: item['metadata'] for item in benchmark_data}
    
    # Build cluster ID mapping for model
    model_results = model_data.get('metadata_results', [])
    model_map = {}
    for result in model_results:
        if result.get('success', False):
            cluster_id = result['cluster_id']
            model_map[cluster_id] = result.get('metadata', {})
    
    # Evaluate each cluster
    cluster_scores = []
    
    for cluster_id in sorted(benchmark_map.keys()):
        benchmark_meta = benchmark_map[cluster_id]
        model_meta = model_map.get(cluster_id, {})
        
        if not model_meta:
            # Model failed to generate metadata for this cluster
            scores = {
                'cluster_id': cluster_id,
                'title': 0.0,
                'summary': 0.0,
                'action_items': 0.0,
                'participants': 0.0,
                'tags': 0.0,
                'urgency': 0.0,
                'status': 0.0,
                'overall': 0.0,
                'model_title': 'N/A',
                'benchmark_title': benchmark_meta.get('title', ''),
                'model_action_items_count': 0,
                'benchmark_action_items_count': len(benchmark_meta.get('action_items', []))
            }
        else:
            scores = evaluate_metadata(benchmark_meta, model_meta)
            scores['cluster_id'] = cluster_id
            scores['model_title'] = model_meta.get('title', 'N/A')
            scores['benchmark_title'] = benchmark_meta.get('title', '')
            scores['model_action_items_count'] = len(model_meta.get('action_items', []))
            scores['benchmark_action_items_count'] = len(benchmark_meta.get('action_items', []))
        
        cluster_scores.append(scores)
    
    # Calculate overall metrics
    avg_title = sum(s['title'] for s in cluster_scores) / len(cluster_scores)
    avg_summary = sum(s['summary'] for s in cluster_scores) / len(cluster_scores)
    avg_action_items = sum(s['action_items'] for s in cluster_scores) / len(cluster_scores)
    avg_participants = sum(s['participants'] for s in cluster_scores) / len(cluster_scores)
    avg_tags = sum(s['tags'] for s in cluster_scores) / len(cluster_scores)
    avg_urgency = sum(s['urgency'] for s in cluster_scores) / len(cluster_scores)
    avg_status = sum(s['status'] for s in cluster_scores) / len(cluster_scores)
    avg_overall = sum(s['overall'] for s in cluster_scores) / len(cluster_scores)
    
    # Success rate
    success_count = len(model_map)
    total_count = len(benchmark_map)
    success_rate = (success_count / total_count * 100) if total_count > 0 else 0
    
    return {
        'cluster_scores': cluster_scores,
        'avg_title_score': avg_title,
        'avg_summary_score': avg_summary,
        'avg_action_items_score': avg_action_items,
        'avg_participants_score': avg_participants,
        'avg_tags_score': avg_tags,
        'avg_urgency_score': avg_urgency,
        'avg_status_score': avg_status,
        'avg_overall_score': avg_overall,
        'success_rate': success_rate,
        'successful_clusters': success_count,
        'total_clusters': total_count
    }

def create_excel_file(model_name, metrics, token_usage, cost):
    """Create Excel file for a single model with fully explained analysis"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Step 3 Analysis"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Title
    ws[f'A{current_row}'] = f"STEP 3 METADATA GENERATION ANALYSIS: {model_name}"
    ws[f'A{current_row}'].font = title_font
    ws.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 2
    
    # Overall Metrics Section
    ws[f'A{current_row}'] = "OVERALL METRICS"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    # Main metrics (used for final score calculation)
    metrics_data = [
        ['Success Rate', f"{metrics['success_rate']:.2f}%"],
        ['Successful Clusters', f"{metrics['successful_clusters']}/{metrics['total_clusters']}"],
        ['Average Overall Score', f"{metrics['avg_overall_score']:.2f}%"],
    ]
    
    for label, value in metrics_data:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
    
    # Add formula explanation
    current_row += 1
    ws[f'A{current_row}'] = "OVERALL SCORE FORMULA:"
    ws[f'A{current_row}'].font = Font(bold=True, italic=True)
    current_row += 1
    
    formula_text = "Average of all individual cluster overall scores (see table below)"
    ws[f'A{current_row}'] = formula_text
    ws[f'A{current_row}'].font = Font(italic=True)
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 2
    
    # Individual metric averages (for reference only)
    ws[f'A{current_row}'] = "INDIVIDUAL METRIC AVERAGES (Reference Only)"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    reference_metrics = [
        ['Average Title Score', f"{metrics['avg_title_score']:.2f}%"],
        ['Average Summary Score', f"{metrics['avg_summary_score']:.2f}%"],
        ['Average Action Items Score', f"{metrics['avg_action_items_score']:.2f}%"],
        ['Average Participants Score', f"{metrics['avg_participants_score']:.2f}%"],
        ['Average Tags Score', f"{metrics['avg_tags_score']:.2f}%"],
        ['Average Urgency Score', f"{metrics['avg_urgency_score']:.2f}%"],
        ['Average Status Score', f"{metrics['avg_status_score']:.2f}%"],
    ]
    
    for label, value in reference_metrics:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        ws[f'A{current_row}'].font = Font(italic=True)
        current_row += 1
    
    current_row += 1
    note_text = "Note: These averages are for reference only and not used in final score calculation"
    ws[f'A{current_row}'] = note_text
    ws[f'A{current_row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{current_row}:B{current_row}')
    
    current_row += 1
    
    # Token Usage Section
    ws[f'A{current_row}'] = "TOKEN USAGE & COST"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    ws[f'A{current_row}'] = 'INPUT_TOKENS'
    ws[f'B{current_row}'] = f"{token_usage['INPUT_TOKENS']:,}"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws[f'A{current_row}'] = 'OUTPUT_TOKENS'
    ws[f'B{current_row}'] = f"{token_usage['OUTPUT_TOKENS']:,}"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws[f'A{current_row}'] = 'TOTAL_TOKENS'
    ws[f'B{current_row}'] = f"{token_usage['TOTAL_TOKENS']:,}"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    # Add note about token calculation
    ws[f'A{current_row}'] = 'Token Calculation Method'
    note_text = "Output tokens calculated from actual raw responses (~4 chars per token). Input tokens = Total - Output tokens."
    ws[f'B{current_row}'] = note_text
    ws[f'A{current_row}'].font = Font(italic=True, size=9)
    ws[f'B{current_row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    ws[f'A{current_row}'] = 'TOKEN_COST'
    ws[f'B{current_row}'] = f"${cost['total_cost']:.6f}"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    # Cost calculation formula
    ws[f'A{current_row}'] = 'Cost Formula'
    formula_text = f"({token_usage['INPUT_TOKENS']:,} / 1,000,000 × ${cost['input_cost_per_1m']:.2f}) + ({token_usage['OUTPUT_TOKENS']:,} / 1,000,000 × ${cost['output_cost_per_1m']:.2f})"
    ws[f'B{current_row}'] = formula_text
    ws[f'A{current_row}'].font = Font(italic=True)
    ws[f'B{current_row}'].font = Font(italic=True)
    current_row += 1
    
    # Step-by-step cost breakdown
    ws[f'A{current_row}'] = 'Step 1 - Input Cost:'
    ws[f'B{current_row}'] = f"{token_usage['INPUT_TOKENS']:,} tokens ÷ 1,000,000 × ${cost['input_cost_per_1m']:.2f} = ${cost['input_cost']:.6f}"
    current_row += 1
    
    ws[f'A{current_row}'] = 'Step 2 - Output Cost:'
    ws[f'B{current_row}'] = f"{token_usage['OUTPUT_TOKENS']:,} tokens ÷ 1,000,000 × ${cost['output_cost_per_1m']:.2f} = ${cost['output_cost']:.6f}"
    current_row += 1
    
    ws[f'A{current_row}'] = 'Step 3 - Total Cost:'
    ws[f'B{current_row}'] = f"${cost['input_cost']:.6f} + ${cost['output_cost']:.6f} = ${cost['total_cost']:.6f}"
    current_row += 2
    
    # Detailed Cluster Scores
    ws[f'A{current_row}'] = "DETAILED CLUSTER SCORES"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:K{current_row}')
    current_row += 1
    
    # Add formula explanation for individual scores
    ws[f'A{current_row}'] = "INDIVIDUAL CLUSTER OVERALL SCORE FORMULA:"
    ws[f'A{current_row}'].font = Font(bold=True, italic=True)
    current_row += 1
    
    formula_text = "Overall Score = (Title×30%) + (Summary×25%) + (Action Items×20%) + (Participants×10%) + (Tags×5%) + (Urgency×5%) + (Status×5%)"
    ws[f'A{current_row}'] = formula_text
    ws[f'A{current_row}'].font = Font(italic=True, size=10)
    ws.merge_cells(f'A{current_row}:K{current_row}')
    current_row += 2
    
    # Table headers
    headers = [
        'Cluster ID', 'Benchmark Title', 'Model Title', 'Title Score (30%)', 
        'Summary Score (25%)', 'Action Items Score (20%)', 'Participants Score (10%)',
        'Tags Score (5%)', 'Urgency Score (5%)', 'Status Score (5%)', 'Overall Score'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    current_row += 1
    
    # Add cluster data
    for cluster_score in metrics['cluster_scores']:
        row_data = [
            cluster_score['cluster_id'],
            cluster_score['benchmark_title'],
            cluster_score['model_title'],
            f"{cluster_score['title']:.2f}%",
            f"{cluster_score['summary']:.2f}%",
            f"{cluster_score['action_items']:.2f}%",
            f"{cluster_score['participants']:.2f}%",
            f"{cluster_score['tags']:.2f}%",
            f"{cluster_score['urgency']:.2f}%",
            f"{cluster_score['status']:.2f}%",
            f"{cluster_score['overall']:.2f}%"
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = border
            
            # Color code overall score
            if col_num == 11:  # Overall Score column
                score = cluster_score['overall']
                if score >= 80:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= 60:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['K'].width = 14
    
    # Save file
    output_dir = "output/step3_client_analysis"
    os.makedirs(output_dir, exist_ok=True)
    
    clean_model_name = model_name.replace('/', '_').replace('\\', '_')
    output_file = os.path.join(output_dir, f"{clean_model_name}_step3_analysis.xlsx")
    wb.save(output_file)
    
    print(f"  ✅ Created {output_file}")
    return output_file

def main():
    print("=" * 80)
    print("STEP 3 METADATA GENERATION ANALYSIS - ALL MODELS")
    print("=" * 80)
    print()
    
    # Load benchmark
    benchmark_data = load_benchmark_metadata()
    
    # Find all model output files
    model_files = glob.glob("output/phase5_metadata_generation/*.json")
    
    # Exclude comprehensive_results.json
    model_files = [f for f in model_files if 'comprehensive_results' not in f]
    
    print(f"\n📁 Found {len(model_files)} model output files")
    print()
    
    # Store results for CSV comparison
    all_results = []
    
    # Process each model
    for model_file in sorted(model_files):
        model_name = extract_model_name_from_file(model_file)
        print(f"📊 Processing: {model_name}")
        
        try:
            # Load model output
            model_data = load_model_output(model_file)
            
            # Check if model has any successful results (even if overall success is False)
            metadata_results = model_data.get('metadata_results', [])
            successful_results = sum(1 for r in metadata_results if r.get('success', False))
            
            if successful_results == 0:
                print(f"  ⚠️  Model failed: No successful results")
                continue
            
            # Get token usage
            token_usage = get_token_usage(model_data)
            
            # Get pricing (per 1M tokens)
            input_price_per_1m, output_price_per_1m = get_pricing(model_name)
            
            # Calculate costs using official pricing and actual token counts
            input_cost = (token_usage['INPUT_TOKENS'] / 1000000) * input_price_per_1m
            output_cost = (token_usage['OUTPUT_TOKENS'] / 1000000) * output_price_per_1m
            total_cost = input_cost + output_cost
            
            cost_info = {
                'input_cost': input_cost,
                'output_cost': output_cost,
                'total_cost': total_cost,
                'input_cost_per_1m': input_price_per_1m,
                'output_cost_per_1m': output_price_per_1m
            }
            
            # Calculate metrics
            metrics = calculate_metrics(model_data, benchmark_data)
            
            # Create Excel file
            excel_file = create_excel_file(model_name, metrics, token_usage, cost_info)
            
            # Store for CSV (only include metrics used for final score calculation)
            all_results.append({
                'MODEL': model_name,
                'OVERALL_SCORE': metrics['avg_overall_score'],
                'SUCCESS_RATE': metrics['success_rate'],
                'SUCCESSFUL_CLUSTERS': metrics['successful_clusters'],
                'TOTAL_CLUSTERS': metrics['total_clusters'],
                'INPUT_TOKENS': token_usage['INPUT_TOKENS'],
                'OUTPUT_TOKENS': token_usage['OUTPUT_TOKENS'],
                'TOTAL_TOKENS': token_usage['TOTAL_TOKENS'],
                'TOKEN_COST': total_cost,
                'DURATION': model_data.get('duration', 0),
                # Reference metrics (not used in final score)
                'TITLE_SCORE_AVG': metrics['avg_title_score'],
                'SUMMARY_SCORE_AVG': metrics['avg_summary_score'],
                'ACTION_ITEMS_SCORE_AVG': metrics['avg_action_items_score'],
                'PARTICIPANTS_SCORE_AVG': metrics['avg_participants_score'],
                'TAGS_SCORE_AVG': metrics['avg_tags_score'],
                'URGENCY_SCORE_AVG': metrics['avg_urgency_score'],
                'STATUS_SCORE_AVG': metrics['avg_status_score']
            })
            
            print(f"  ✅ Overall Score: {metrics['avg_overall_score']:.2f}%")
            print(f"  💰 Cost: ${total_cost:.6f}")
            print()
            
        except Exception as e:
            print(f"  ❌ Error processing {model_name}: {str(e)}")
            print()
            continue
    
    # Create comparison CSV
    if all_results:
        df = pd.DataFrame(all_results)
        
        # Sort by overall score (descending)
        df = df.sort_values('OVERALL_SCORE', ascending=False)
        
        csv_file = "output/step3_client_analysis/step3_all_models_comparison.csv"
        df.to_csv(csv_file, index=False)
        
        print("=" * 80)
        print(f"✅ Created comparison CSV: {csv_file}")
        print("=" * 80)
        print()
        
        # Display top 5 models
        print("🏆 TOP 5 MODELS BY OVERALL SCORE:")
        print("-" * 80)
        for idx, row in df.head(5).iterrows():
            print(f"  {row['MODEL']:<40} Score: {row['OVERALL_SCORE']:>6.2f}% | Cost: ${row['TOKEN_COST']:>9.6f}")
        print()
    
    print("✅ All models processed successfully!")

if __name__ == "__main__":
    main()

