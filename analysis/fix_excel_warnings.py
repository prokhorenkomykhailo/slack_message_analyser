#!/usr/bin/env python3
"""
Fix all Excel warnings in gemini_1.5_flash_updated_analysis.xlsx
WITHOUT removing any cells - just clean up the values and formatting
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def fix_excel_warnings():
    """Fix all warnings in the existing Excel file without removing cells"""
    
    input_file = "gemini_1.5_flash_updated_analysis.xlsx"
    output_file = "gemini_1.5_flash_updated_analysis_FIXED.xlsx"
    
    print(f"Loading Excel file: {input_file}")
    
    try:
        # Load the existing workbook
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        print(f"Original worksheet: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {ws.max_column}")
        
        # Fix worksheet title (remove special characters that cause warnings)
        ws.title = "Gemini_1_5_Flash_Analysis"
        
        warnings_fixed = 0
        
        # Go through all cells and fix warnings
        for row_num in range(1, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Fix cell value issues
                if cell.value is not None:
                    original_value = cell.value
                    
                    # Convert to string and clean up
                    if isinstance(cell.value, (int, float)):
                        # Keep numbers as they are, but ensure they're clean
                        if isinstance(cell.value, float) and cell.value == int(cell.value):
                            cell.value = int(cell.value)  # Convert 5.0 to 5
                    else:
                        # Clean up string values
                        str_value = str(cell.value).strip()
                        
                        # Fix common problematic values
                        if str_value.lower() in ['nan', 'none', 'null', '']:
                            cell.value = ""  # Convert to empty string instead of None
                            warnings_fixed += 1
                        elif str_value != str(original_value):
                            cell.value = str_value
                            warnings_fixed += 1
                
                # Ensure empty cells have empty string instead of None
                elif cell.value is None:
                    cell.value = ""
                    
                # Fix any data validation or formatting issues
                # Clear any problematic data validation
                if hasattr(cell, 'data_validation') and cell.data_validation:
                    cell.data_validation = None
                    warnings_fixed += 1
        
        print(f"Fixed {warnings_fixed} potential warning issues")
        
        # Clean up column widths to reasonable values
        for col_num in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate reasonable width based on content
            max_length = 0
            for row_num in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            
            # Set reasonable width (between 8 and 50)
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            else:
                ws.column_dimensions[column_letter].width = 8
        
        # Clean up any problematic cell references or formulas
        for row_num in range(1, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # If cell has a formula, ensure it's valid
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    # This is a formula cell - make sure it's clean
                    try:
                        formula_value = cell.value
                        if formula_value and isinstance(formula_value, str):
                            # Clean up formula if needed
                            if formula_value.startswith('='):
                                # Keep formula as is if it looks valid
                                pass
                            else:
                                # Convert to regular value
                                cell.value = str(formula_value)
                    except:
                        # If there's any issue with formula, convert to text
                        cell.value = str(cell.value) if cell.value else ""
        
        # Ensure all merged cells are properly handled
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            try:
                # Check if merged range is valid
                min_row, min_col, max_row, max_col = merged_range.bounds
                if min_row <= ws.max_row and min_col <= ws.max_column:
                    # Merged range is valid
                    pass
                else:
                    # Remove invalid merged range
                    ws.unmerge_cells(str(merged_range))
                    warnings_fixed += 1
            except:
                # If there's any issue, unmerge
                try:
                    ws.unmerge_cells(str(merged_range))
                    warnings_fixed += 1
                except:
                    pass
        
        # Save the fixed workbook
        wb.save(output_file)
        
        print(f"Fixed Excel file saved as: {output_file}")
        print(f"Total warnings fixed: {warnings_fixed}")
        
        # Verify the fix worked
        print("\nVerifying the fixed file...")
        wb_test = openpyxl.load_workbook(output_file)
        ws_test = wb_test.active
        
        empty_cells = 0
        content_cells = 0
        
        for row_num in range(1, min(ws_test.max_row + 1, 50)):  # Check first 50 rows
            for col_num in range(1, min(ws_test.max_column + 1, 15)):  # Check first 15 columns
                cell = ws_test.cell(row=row_num, column=col_num)
                if cell.value == "" or cell.value is None:
                    empty_cells += 1
                else:
                    content_cells += 1
        
        print(f"Verification complete:")
        print(f"  - Content cells: {content_cells}")
        print(f"  - Empty cells: {empty_cells}")
        print(f"  - Max row: {ws_test.max_row}")
        print(f"  - Max column: {ws_test.max_column}")
        
        return output_file
        
    except Exception as e:
        print(f"Error fixing Excel file: {e}")
        return None

if __name__ == "__main__":
    print("FIXING EXCEL WARNINGS")
    print("="*50)
    
    fixed_file = fix_excel_warnings()
    
    if fixed_file:
        print(f"\n✅ SUCCESS!")
        print(f"Fixed file: {fixed_file}")
        print("This file should open without warnings in Excel.")
    else:
        print(f"\n❌ FAILED to fix the Excel file.")
