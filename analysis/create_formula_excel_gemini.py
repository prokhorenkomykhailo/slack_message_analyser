#!/usr/bin/env python3
"""
Create Enhanced Excel with All Formulas for Gemini 2.0 Flash 001
Includes all calculations with visible formulas for verification
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

def load_gemini_data():
    """Load specific gemini_2.0_flash_001 data from JSON results"""
    
    results_file = "output/phase4_comprehensive_evaluation/detailed_evaluation_results.json"
    
    try:
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Find the specific model
        for result in results:
            if result['model'] == 'google_gemini-2.0-flash-001':
                return result
        
        print(f"❌ Model google_gemini-2.0-flash-001 not found in results")
        return None
        
    except Exception as e:
        print(f"❌ Error loading results: {e}")
        return None

def create_formula_excel():
    """Create Excel file with all formulas visible"""
    
    print("🎯 Creating Enhanced Excel with All Formulas for Gemini 2.0 Flash 001")
    print("=" * 80)
    
    # Load model data
    model_data = load_gemini_data()
    if not model_data:
        return None
    
    # Create workbook with multiple sheets
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    source_sheet = wb.create_sheet("1_Source_Data")
    calc_sheet = wb.create_sheet("2_Calculations")
    results_sheet = wb.create_sheet("3_Results")
    verification_sheet = wb.create_sheet("4_Verification")
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    formula_font = Font(italic=True, size=10, color="0000FF")
    result_font = Font(bold=True, size=11, color="000000")
    
    blue_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # SHEET 1: SOURCE DATA
    create_source_data_sheet(source_sheet, model_data, header_font, title_font, blue_fill, green_fill, light_blue_fill)
    
    # SHEET 2: CALCULATIONS
    create_calculations_sheet(calc_sheet, model_data, header_font, title_font, formula_font, blue_fill, green_fill, light_blue_fill)
    
    # SHEET 3: RESULTS
    create_results_sheet(results_sheet, model_data, header_font, title_font, result_font, blue_fill, green_fill, light_blue_fill, light_green_fill, yellow_fill)
    
    # SHEET 4: VERIFICATION
    create_verification_sheet(verification_sheet, model_data, header_font, title_font, formula_font, blue_fill, green_fill, light_blue_fill, light_green_fill)
    
    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        auto_adjust_columns(sheet)
    
    # Save workbook
    output_path = "output/step2_client_analysis/gemini_2.0_flash_001_step2_analysis_WITH_FORMULAS.xlsx"
    wb.save(output_path)
    
    print(f"✅ Enhanced Excel with formulas created: {output_path}")
    return output_path

def create_source_data_sheet(ws, data, header_font, title_font, blue_fill, green_fill, light_blue_fill):
    """Create source data sheet with raw values"""
    
    current_row = 1
    
    # Title
    ws.cell(row=current_row, column=1, value="SOURCE DATA - GEMINI 2.0 FLASH 001")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 3
    
    # Basic Model Information
    ws.cell(row=current_row, column=1, value="BASIC MODEL INFORMATION")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 2
    
    basic_info = [
        ['Field', 'Value', 'Source', 'Description'],
        ['Model Name', data['model'], 'JSON: model', 'Full model identifier'],
        ['Provider', data['provider'], 'JSON: provider', 'Model provider'],
        ['Model Short Name', data['model_name'], 'JSON: model_name', 'Short model name'],
        ['Success', data['success'], 'JSON: success', 'Whether evaluation succeeded'],
        ['Duration (seconds)', data['duration'], 'JSON: duration', 'Execution time'],
        ['Timestamp', data['timestamp'], 'JSON: timestamp', 'When evaluation was run']
    ]
    
    for row_data in basic_info:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 6:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Cluster Metrics
    ws.cell(row=current_row, column=1, value="CLUSTER METRICS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 2
    
    cluster_info = [
        ['Field', 'Value', 'Source', 'Description'],
        ['Number of Refined Clusters', data['num_refined_clusters'], 'JSON: num_refined_clusters', 'Final cluster count after refinement'],
        ['Number of Step 1 Clusters', data['num_step1_clusters'], 'JSON: num_step1_clusters', 'Initial cluster count before refinement'],
        ['Cluster Count Reduction', data['cluster_count_reduction'], 'JSON: cluster_count_reduction', 'Clusters reduced from Step 1'],
        ['Average Cluster Size', data['avg_cluster_size'], 'JSON: avg_cluster_size', 'Mean messages per refined cluster'],
        ['Average Step 1 Size', data['avg_step1_size'], 'JSON: avg_step1_size', 'Mean messages per initial cluster'],
        ['Size Reduction', data['size_reduction'], 'JSON: size_reduction', 'Change in average cluster size'],
        ['Size Std Reduction', data['size_std_reduction'], 'JSON: size_std_reduction', 'Change in cluster size standard deviation'],
        ['Total Participants', data['total_participants'], 'JSON: total_participants', 'Unique participants across all clusters']
    ]
    
    for row_data in cluster_info:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 15:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Operations Data
    ws.cell(row=current_row, column=1, value="OPERATIONS DATA")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 2
    
    operations_info = [
        ['Field', 'Value', 'Source', 'Description'],
        ['Total Operations', data['total_operations'], 'JSON: total_operations', 'Total refinement operations performed'],
        ['Merge Operations', data['merge_operations'], 'JSON: merge_operations', 'Number of merge operations'],
        ['Split Operations', data['split_operations'], 'JSON: split_operations', 'Number of split operations'],
        ['Refine Operations', data['refine_operations'], 'JSON: refine_operations', 'Number of refine operations']
    ]
    
    for row_data in operations_info:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 25:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Benchmark Evaluation Data
    benchmark_eval = data.get('benchmark_evaluation', {})
    ws.cell(row=current_row, column=1, value="BENCHMARK EVALUATION DATA")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 2
    
    benchmark_info = [
        ['Field', 'Value', 'Source', 'Description'],
        ['Number of Refined Clusters', benchmark_eval.get('num_refined_clusters', 0), 'JSON: benchmark_evaluation.num_refined_clusters', 'Clusters evaluated against benchmark'],
        ['Number of Benchmark Topics', benchmark_eval.get('num_benchmark_topics', 0), 'JSON: benchmark_evaluation.num_benchmark_topics', 'Total benchmark topics available'],
        ['Average Title Similarity', benchmark_eval.get('avg_title_similarity', 0), 'JSON: benchmark_evaluation.avg_title_similarity', 'Mean similarity between cluster and benchmark titles'],
        ['Average Combined Similarity', benchmark_eval.get('avg_combined_similarity', 0), 'JSON: benchmark_evaluation.avg_combined_similarity', 'Mean combined similarity score'],
        ['High Quality Matches', benchmark_eval.get('high_quality_matches', 0), 'JSON: benchmark_evaluation.high_quality_matches', 'Clusters with >70% similarity to benchmarks'],
        ['Quality Rate', benchmark_eval.get('quality_rate', 0), 'JSON: benchmark_evaluation.quality_rate', 'Percentage of high-quality matches']
    ]
    
    for row_data in benchmark_info:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 33:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1

def create_calculations_sheet(ws, data, header_font, title_font, formula_font, blue_fill, green_fill, light_blue_fill):
    """Create calculations sheet with all formulas"""
    
    current_row = 1
    
    # Title
    ws.cell(row=current_row, column=1, value="CALCULATIONS - ALL FORMULAS FOR GEMINI 2.0 FLASH 001")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 3
    
    # Performance Score Calculations
    ws.cell(row=current_row, column=1, value="PERFORMANCE SCORE CALCULATIONS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2
    
    # Set up calculation references
    # Row references for source data (assuming source data is in rows 6-40)
    source_row_base = 6
    
    calc_data = [
        ['Metric', 'Formula', 'Excel Formula', 'Result', 'Explanation'],
        ['Similarity Score', 'avg_combined_similarity × 100', f'=1_Source_Data.B{source_row_base + 7}*100', '', 'Convert similarity to percentage'],
        ['Quality Score', 'quality_rate × 100', f'=1_Source_Data.B{source_row_base + 8}*100', '', 'Convert quality rate to percentage'],
        ['Operation Efficiency', 'max(0, min(100, 100 - (total_ops / refined_clusters × 10)))', f'=MAX(0,MIN(100,100-(1_Source_Data.B{source_row_base + 1}/1_Source_Data.B{source_row_base}*10)))', '', 'Efficiency based on operations per cluster'],
        ['Size Optimization', 'max(0, min(100, 100 - |size_reduction|))', f'=MAX(0,MIN(100,100-ABS(1_Source_Data.B{source_row_base + 5})))', '', 'How well cluster sizes were optimized'],
        ['Combined Score', 'similarity×0.35 + quality×0.25 + size×0.20 + efficiency×0.20', f'=B{current_row + 1}*0.35+B{current_row + 2}*0.25+B{current_row + 4}*0.20+B{current_row + 3}*0.20', '', 'Weighted combination of all scores']
    ]
    
    for i, row_data in enumerate(calc_data):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 6:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif col == 3:  # Formula column
                cell.font = formula_font
                cell.fill = light_blue_fill
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Detailed Breakdown Calculations
    ws.cell(row=current_row, column=1, value="DETAILED SCORE BREAKDOWN")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2
    
    breakdown_data = [
        ['Component', 'Weight', 'Score', 'Contribution Formula', 'Contribution'],
        ['Benchmark Similarity', '35%', f'=B{current_row - 4}', f'=B{current_row - 4}*0.35', ''],
        ['Quality Rate', '25%', f'=B{current_row - 3}', f'=B{current_row - 3}*0.25', ''],
        ['Size Optimization', '20%', f'=B{current_row - 1}', f'=B{current_row - 1}*0.20', ''],
        ['Operation Efficiency', '20%', f'=B{current_row - 2}', f'=B{current_row - 2}*0.20', ''],
        ['TOTAL COMBINED SCORE', '100%', f'=SUM(E{current_row - 4}:E{current_row - 1})', '=SUM(E{current_row - 4}:E{current_row - 1})', '']
    ]
    
    for i, row_data in enumerate(breakdown_data):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 15:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif col == 4:  # Formula column
                cell.font = formula_font
                cell.fill = light_blue_fill
            elif 'TOTAL' in str(value):
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True, color="0000FF")
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Operations Analysis Calculations
    ws.cell(row=current_row, column=1, value="OPERATIONS ANALYSIS CALCULATIONS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2
    
    ops_calc_data = [
        ['Metric', 'Formula', 'Excel Formula', 'Result', 'Explanation'],
        ['Merge Operations %', 'merge_ops / total_ops × 100', f'=IF(1_Source_Data.B{source_row_base + 1}>0,1_Source_Data.B{source_row_base + 2}/1_Source_Data.B{source_row_base + 1}*100,0)', '', 'Percentage of merge operations'],
        ['Split Operations %', 'split_ops / total_ops × 100', f'=IF(1_Source_Data.B{source_row_base + 1}>0,1_Source_Data.B{source_row_base + 3}/1_Source_Data.B{source_row_base + 1}*100,0)', '', 'Percentage of split operations'],
        ['Refine Operations %', 'refine_ops / total_ops × 100', f'=IF(1_Source_Data.B{source_row_base + 1}>0,1_Source_Data.B{source_row_base + 4}/1_Source_Data.B{source_row_base + 1}*100,0)', '', 'Percentage of refine operations'],
        ['Operations per Cluster', 'total_ops / refined_clusters', f'=1_Source_Data.B{source_row_base + 1}/1_Source_Data.B{source_row_base}', '', 'Average operations per refined cluster']
    ]
    
    for i, row_data in enumerate(ops_calc_data):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 22:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif col == 3:  # Formula column
                cell.font = formula_font
                cell.fill = light_blue_fill
            else:
                cell.fill = light_blue_fill
        current_row += 1

def create_results_sheet(ws, data, header_font, title_font, result_font, blue_fill, green_fill, light_blue_fill, light_green_fill, yellow_fill):
    """Create results sheet with formatted output"""
    
    current_row = 1
    source_row_base = 6  # Base row for source data references
    
    # Title
    ws.cell(row=current_row, column=1, value="GEMINI 2.0 FLASH 001 - STEP 2 ANALYSIS RESULTS")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 3
    
    # Executive Summary
    ws.cell(row=current_row, column=1, value="EXECUTIVE SUMMARY")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    summary_data = [
        ['Metric', 'Value', 'Formula Reference', 'Explanation'],
        ['Overall Combined Score', f'=2_Calculations.B{11}', '=2_Calculations.B11', 'Composite score for Step 2 refinement performance'],
        ['Benchmark Similarity', f'=2_Calculations.B{7}', '=2_Calculations.B7', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'=2_Calculations.B{8}', '=2_Calculations.B8', 'Percentage of clusters with >70% similarity to benchmarks'],
        ['High Quality Matches', f'=1_Source_Data.B{source_row_base + 8}', '=1_Source_Data.B{source_row_base + 8}', 'Number of excellent cluster matches'],
        ['Clusters Generated', f'=1_Source_Data.B{source_row_base}', '=1_Source_Data.B{source_row_base}', 'Number of clusters after refinement'],
        ['Cluster Reduction', f'=1_Source_Data.B{source_row_base + 2}', '=1_Source_Data.B{source_row_base + 2}', 'Number of clusters reduced from Step 1'],
        ['Operations Performed', f'=1_Source_Data.B{source_row_base + 1}', '=1_Source_Data.B{source_row_base + 1}', 'Total refinement operations performed'],
        ['Average Cluster Size', f'=1_Source_Data.B{source_row_base + 3}', '=1_Source_Data.B{source_row_base + 3}', 'Average size of refined clusters'],
        ['Size Optimization', f'=2_Calculations.B{10}', '=2_Calculations.B10', 'How well cluster sizes were optimized']
    ]
    
    for row_data in summary_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 5:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Combined Score Explanation
    ws.cell(row=current_row, column=1, value="🔍 DETAILED EXPLANATION OF COMBINED SCORE")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    explanation_data = [
        ['Component', 'Score', 'Weight', 'Contribution', 'Description'],
        ['Benchmark Similarity', f'=2_Calculations.B{7}', '35%', f'=2_Calculations.E{16}', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'=2_Calculations.B{8}', '25%', f'=2_Calculations.E{17}', 'Percentage of high-quality matches (>70% similarity)'],
        ['Size Optimization', f'=2_Calculations.B{10}', '20%', f'=2_Calculations.E{19}', 'How well cluster sizes were optimized'],
        ['Operation Efficiency', f'=2_Calculations.B{9}', '20%', f'=2_Calculations.E{18}', 'Efficiency of refinement operations'],
        ['TOTAL COMBINED SCORE', f'=2_Calculations.B{11}', '100%', f'=2_Calculations.E{20}', 'Overall Step 2 performance score']
    ]
    
    for row_data in explanation_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 10:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif 'TOTAL' in str(value):
                cell.fill = yellow_fill
                cell.font = result_font
            else:
                cell.fill = light_green_fill
        current_row += 1

def create_verification_sheet(ws, data, header_font, title_font, formula_font, blue_fill, green_fill, light_blue_fill, light_green_fill):
    """Create verification sheet to cross-check all calculations"""
    
    current_row = 1
    
    # Title
    ws.cell(row=current_row, column=1, value="VERIFICATION - CROSS-CHECK ALL CALCULATIONS")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 3
    
    # Manual Calculations
    ws.cell(row=current_row, column=1, value="MANUAL CALCULATIONS (for verification)")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2
    
    # Get actual values for manual calculation
    avg_combined_similarity = data.get('benchmark_evaluation', {}).get('avg_combined_similarity', 0)
    quality_rate = data.get('benchmark_evaluation', {}).get('quality_rate', 0)
    total_operations = data.get('total_operations', 0)
    num_refined_clusters = data.get('num_refined_clusters', 0)
    size_reduction = data.get('size_reduction', 0)
    
    verification_data = [
        ['Calculation', 'Manual Formula', 'Expected Result', 'Excel Result', 'Match?', 'Notes'],
        ['Similarity Score', f'{avg_combined_similarity} × 100', f'{avg_combined_similarity * 100:.2f}', f'=2_Calculations.B7', f'=IF(ABS(C{current_row}-D{current_row})<0.01,"YES","NO")', 'Should match within 0.01'],
        ['Quality Score', f'{quality_rate} × 100', f'{quality_rate * 100:.2f}', f'=2_Calculations.B8', f'=IF(ABS(C{current_row}-D{current_row})<0.01,"YES","NO")', 'Should match within 0.01'],
        ['Operation Efficiency', f'max(0, min(100, 100 - ({total_operations} / {num_refined_clusters} × 10)))', f'{max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))):.2f}', f'=2_Calculations.B9', f'=IF(ABS(C{current_row}-D{current_row})<0.01,"YES","NO")', 'Should match within 0.01'],
        ['Size Optimization', f'max(0, min(100, 100 - |{size_reduction}|))', f'{max(0, min(100, 100 - abs(size_reduction))):.2f}', f'=2_Calculations.B10', f'=IF(ABS(C{current_row}-D{current_row})<0.01,"YES","NO")', 'Should match within 0.01'],
        ['Combined Score', f'{avg_combined_similarity * 100:.2f}×0.35 + {quality_rate * 100:.2f}×0.25 + {max(0, min(100, 100 - abs(size_reduction))):.2f}×0.20 + {max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))):.2f}×0.20', f'{(avg_combined_similarity * 100 * 0.35 + quality_rate * 100 * 0.25 + max(0, min(100, 100 - abs(size_reduction))) * 0.20 + max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))) * 0.20):.2f}', f'=2_Calculations.B11', f'=IF(ABS(C{current_row}-D{current_row})<0.01,"YES","NO")', 'Should match within 0.01']
    ]
    
    for i, row_data in enumerate(verification_data):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 6:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif col == 2 or col == 4:  # Formula columns
                cell.font = formula_font
                cell.fill = light_blue_fill
            elif col == 5:  # Match column
                cell.font = formula_font
                cell.fill = light_green_fill
            else:
                cell.fill = light_blue_fill
        current_row += 1

def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for col_num in range(1, 15):  # A to N
        max_length = 0
        column_letter = get_column_letter(col_num)
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function"""
    
    print("🎯 Creating Enhanced Excel with All Formulas for Gemini 2.0 Flash 001")
    print("=" * 80)
    print("This will create an Excel file with:")
    print("  - Sheet 1: Source Data (raw values from JSON)")
    print("  - Sheet 2: Calculations (all formulas visible)")
    print("  - Sheet 3: Results (formatted output)")
    print("  - Sheet 4: Verification (cross-check calculations)")
    print()
    
    # Create the enhanced Excel file
    result_path = create_formula_excel()
    
    if result_path:
        print(f"\n🎉 Successfully created enhanced Excel file!")
        print(f"📁 File saved to: {result_path}")
        print("\n📋 The file includes:")
        print("  ✅ All source data from JSON")
        print("  ✅ All calculation formulas visible")
        print("  ✅ Cross-references between sheets")
        print("  ✅ Verification formulas to check accuracy")
        print("  ✅ Step-by-step breakdown of all calculations")
    else:
        print("\n❌ Failed to create enhanced Excel file")

if __name__ == "__main__":
    main()
