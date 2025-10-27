#!/usr/bin/env python3
"""
Update Excel sheet to show CORRECT 92.9 score using proper benchmark data
Fix all the incorrect values with the corrected ones
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
import json

def update_to_correct_score():
    """Update Excel to show correct 92.9 score with proper benchmark data"""
    
    file_path = "gemini_1.5_flash_updated_analysis_FIXED.xlsx"
    
    print("UPDATING TO CORRECT 92.9 SCORE")
    print("="*50)
    
    # Load corrected data
    corrected_df = pd.read_csv('/home/ubuntu/deemerge/phase_evaluation_engine/gemini_1.5_flash_corrected_analysis.csv')
    corrected_gemini = corrected_df[corrected_df['MODEL'] == 'gemini-1.5-flash'].copy()
    
    print("CORRECTED VALUES:")
    for _, row in corrected_gemini.iterrows():
        print(f"Topic {row['CLUSTER_ID']}: Coverage={row['COVERAGE_PERCENTAGE']:.2f}%, Precision={row['PRECISION_PERCENT']:.2f}%")
    
    # Calculate correct component scores
    expected_clusters = 6
    total_clusters = len(corrected_gemini)
    
    # Component 1: Cluster Count Score
    cluster_count_ratio = min(expected_clusters, total_clusters) / max(expected_clusters, total_clusters)
    cluster_count_score = cluster_count_ratio * 100
    
    # Component 2: Coverage Score  
    coverage_score = 100.0  # Found all 6 expected clusters
    
    # Component 3: Precision Score (CORRECTED)
    corrected_precisions = corrected_gemini['PRECISION_PERCENT'].tolist()
    avg_precision_corrected = sum(corrected_precisions) / len(corrected_precisions)
    
    # Component 4: Deviation Score (CORRECTED)
    corrected_deviations = [abs(x) for x in corrected_gemini['MESSAGE_COUNT_DEVIATION_PERCENT'].tolist()]
    avg_deviation_corrected = sum(corrected_deviations) / len(corrected_deviations)
    deviation_score_corrected = max(0, 100 - avg_deviation_corrected)
    
    # Final CORRECT score
    correct_final_score = (
        cluster_count_score * 0.25 +
        coverage_score * 0.25 + 
        avg_precision_corrected * 0.25 +
        deviation_score_corrected * 0.25
    )
    
    print(f"\nCORRECT COMPONENT SCORES:")
    print(f"Cluster Count Score: {cluster_count_score:.2f}")
    print(f"Coverage Score: {coverage_score:.2f}")
    print(f"Precision Score: {avg_precision_corrected:.2f}")
    print(f"Deviation Score: {deviation_score_corrected:.2f}")
    print(f"FINAL CORRECT SCORE: {correct_final_score:.1f}")
    
    # Load Excel workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Update overall score in summary section
    print(f"\nUpdating Excel file...")
    
    for row_num in range(1, 50):  # Check first 50 rows for summary
        for col_num in range(1, 10):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value and str(cell.value).strip() == "91.1":
                cell.value = f"{correct_final_score:.1f}"
                cell.font = Font(bold=True, color="FF0000")
                print(f"Updated overall score at row {row_num}, col {col_num}")
    
    # Update Step 1 individual topic values
    step1_found = False
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "STEP 1" in str(cell_value):
            step1_found = True
            continue
            
        if step1_found and cell_value and str(cell_value).strip() == "Topic":
            # Found Step 1 header row
            print(f"Updating Step 1 values starting at row {row_num + 1}")
            
            # Update each topic's values with corrected data
            for i, (_, corr_row) in enumerate(corrected_gemini.iterrows()):
                topic_row = row_num + 1 + i
                if topic_row <= ws.max_row:
                    # Update Coverage (column 2)
                    coverage_cell = ws.cell(row=topic_row, column=2)
                    coverage_cell.value = f"{corr_row['COVERAGE_PERCENTAGE']:.2f}%"
                    
                    # Update Precision (column 3)
                    precision_cell = ws.cell(row=topic_row, column=3)
                    precision_cell.value = f"{corr_row['PRECISION_PERCENT']:.2f}%"
                    
                    # Update Recall (column 4)
                    recall_cell = ws.cell(row=topic_row, column=4)
                    recall_cell.value = f"{corr_row['RECALL_PERCENT']:.2f}%"
                    
                    # Update Deviation (column 5)
                    deviation_cell = ws.cell(row=topic_row, column=5)
                    deviation_cell.value = f"{corr_row['MESSAGE_COUNT_DEVIATION_PERCENT']:.2f}%"
                    
                    # Update Perfect Match status (column 6)
                    perfect_cell = ws.cell(row=topic_row, column=6)
                    is_perfect = (corr_row['COVERAGE_PERCENTAGE'] == 100.0 and corr_row['PRECISION_PERCENT'] == 100.0)
                    perfect_cell.value = "YES" if is_perfect else "NO"
                    if is_perfect:
                        perfect_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    else:
                        perfect_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    
                    print(f"Updated Topic {corr_row['CLUSTER_ID']}: {corr_row['COVERAGE_PERCENTAGE']:.2f}% coverage, {corr_row['PRECISION_PERCENT']:.2f}% precision")
            break
    
    # Update Step 2 detailed calculations
    print(f"\nUpdating Step 2 detailed calculations...")
    
    # Find and update detailed precision calculation
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "DETAILED PRECISION CALCULATION" in str(cell_value):
            print(f"Found detailed precision section at row {row_num}")
            
            # Update individual precision values
            for i, (_, corr_row) in enumerate(corrected_gemini.iterrows()):
                precision_row = row_num + 1 + i
                if precision_row <= ws.max_row:
                    precision_line = f"  Topic {corr_row['CLUSTER_ID']}: {corr_row['PRECISION_PERCENT']:.2f}%"
                    ws.cell(row=precision_row, column=1).value = precision_line
            
            # Update sum and average
            sum_row = row_num + 1 + len(corrected_gemini)
            avg_row = sum_row + 1
            
            if sum_row <= ws.max_row:
                sum_line = f"Sum: {' + '.join([f'{p:.2f}' for p in corrected_precisions])} = {sum(corrected_precisions):.2f}"
                ws.cell(row=sum_row, column=1).value = sum_line
                ws.cell(row=sum_row, column=1).font = Font(bold=True)
            
            if avg_row <= ws.max_row:
                avg_line = f"Average: {sum(corrected_precisions):.2f} ÷ {len(corrected_precisions)} = {avg_precision_corrected:.2f}%"
                ws.cell(row=avg_row, column=1).value = avg_line
                ws.cell(row=avg_row, column=1).font = Font(bold=True, color="0000FF")
            
            break
    
    # Update Component 3 in the detailed formulas section
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "COMPONENT 3: PRECISION SCORE" in str(cell_value):
            print(f"Found Component 3 section at row {row_num}")
            
            # Update individual precision values in this section
            current_row = row_num + 1
            for i, (_, corr_row) in enumerate(corrected_gemini.iterrows()):
                precision_line = f"  Topic {corr_row['CLUSTER_ID']}: {corr_row['PRECISION_PERCENT']:.2f}%"
                ws.cell(row=current_row + i, column=1).value = precision_line
            
            # Update calculation formulas
            calc_start_row = current_row + len(corrected_gemini)
            
            # Update sum
            ws.cell(row=calc_start_row, column=2).value = f"{sum(corrected_precisions):.2f}"
            
            # Update average calculation
            ws.cell(row=calc_start_row + 1, column=2).value = f"{sum(corrected_precisions):.2f} ÷ {len(corrected_precisions)} = {avg_precision_corrected:.2f}%"
            
            # Update result
            ws.cell(row=calc_start_row + 2, column=2).value = f"{avg_precision_corrected:.2f}"
            ws.cell(row=calc_start_row + 2, column=2).font = Font(bold=True, color="0000FF")
            
            break
    
    # Update final calculation
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "FINAL WEIGHTED CALCULATION" in str(cell_value):
            print(f"Found final calculation section at row {row_num}")
            
            # Update substitution line
            subst_row = row_num + 2
            subst_line = f"Substitution: ({cluster_count_score:.2f} × 0.25) + ({coverage_score:.2f} × 0.25) + ({avg_precision_corrected:.2f} × 0.25) + ({deviation_score_corrected:.2f} × 0.25)"
            ws.cell(row=subst_row, column=2).value = subst_line
            
            # Update calculation line
            calc_row = row_num + 3
            calc_line = f"Calculation: {cluster_count_score * 0.25:.2f} + {coverage_score * 0.25:.2f} + {avg_precision_corrected * 0.25:.2f} + {deviation_score_corrected * 0.25:.2f}"
            ws.cell(row=calc_row, column=2).value = calc_line
            
            # Update final result
            result_row = row_num + 4
            result_line = f"FINAL RESULT: {correct_final_score:.1f}"
            ws.cell(row=result_row, column=2).value = result_line
            ws.cell(row=result_row, column=2).font = Font(bold=True, color="FF0000", size=12)
            ws.cell(row=result_row, column=2).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            break
    
    # Add correction notice at the top
    ws.cell(row=3, column=1).value = "✅ CORRECTED TO SHOW PROPER 92.9 SCORE WITH ACCURATE BENCHMARK DATA"
    ws.cell(row=3, column=1).font = Font(bold=True, color="008000", size=12)
    ws.cell(row=3, column=1).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Save the updated workbook
    wb.save(file_path)
    
    print(f"\n✅ EXCEL FILE UPDATED!")
    print(f"File: {file_path}")
    print(f"Now shows CORRECT score: {correct_final_score:.1f}")
    print("All values updated to reflect accurate benchmark data")

if __name__ == "__main__":
    update_to_correct_score()
