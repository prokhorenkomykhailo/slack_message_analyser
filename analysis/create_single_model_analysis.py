#!/usr/bin/env python3
"""
Create Single Model Analysis - Step 2
Creates individual Excel analysis for one model at a time
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
import sys

def load_model_data(model_name):
    """Load specific model data from JSON results"""
    
    results_file = "output/phase4_comprehensive_evaluation/detailed_evaluation_results.json"
    
    try:
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Find the specific model
        for result in results:
            if result['model'] == model_name:
                return result
        
        print(f"❌ Model {model_name} not found in results")
        return None
        
    except Exception as e:
        print(f"❌ Error loading results: {e}")
        return None

def create_single_model_excel(model_name, output_dir):
    """Create Excel analysis for a single model"""
    
    print(f"🎯 Creating Step 2 analysis for: {model_name}")
    print("=" * 60)
    
    # Load model data
    model_result = load_model_data(model_name)
    if not model_result:
        return None
    
    if not model_result.get('success', False):
        print(f"⚠️ Model {model_name} evaluation failed - skipping")
        return None
    
    # Calculate key metrics
    num_refined_clusters = model_result['num_refined_clusters']
    num_step1_clusters = model_result['num_step1_clusters']
    cluster_reduction = model_result['cluster_count_reduction']
    
    # Performance metrics
    avg_similarity = model_result['avg_combined_similarity']
    quality_rate = model_result['quality_rate']
    high_quality_matches = model_result['high_quality_matches']
    
    # Operations analysis
    total_operations = model_result['total_operations']
    merge_ops = model_result['merge_operations']
    split_ops = model_result['split_operations']
    refine_ops = model_result['refine_operations']
    
    # Size optimization
    avg_cluster_size = model_result['avg_cluster_size']
    size_reduction = model_result['size_reduction']
    
    # Calculate performance scores
    similarity_score = avg_similarity * 100
    quality_score = quality_rate * 100
    operation_efficiency = max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))) if num_refined_clusters > 0 else 100
    size_optimization = max(0, min(100, 100 - abs(size_reduction)))
    
    # Combined score
    combined_score = (
        similarity_score * 0.35 +        # 35% - similarity to benchmark
        quality_score * 0.25 +           # 25% - quality rate
        size_optimization * 0.20 +       # 20% - size optimization
        operation_efficiency * 0.20      # 20% - operation efficiency
    )
    
    print(f"📊 Model Performance Summary:")
    print(f"   - Combined Score: {combined_score:.2f}")
    print(f"   - Similarity Score: {similarity_score:.2f}%")
    print(f"   - Quality Score: {quality_score:.2f}%")
    print(f"   - High Quality Matches: {high_quality_matches}/{num_refined_clusters}")
    print(f"   - Clusters: {num_step1_clusters} → {num_refined_clusters} (reduction: {cluster_reduction})")
    print(f"   - Operations: {total_operations} (merge: {merge_ops}, split: {split_ops}, refine: {refine_ops})")
    print()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{model_name} Step 2 Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    result_font = Font(bold=True, size=11, color="0000FF")
    formula_font = Font(italic=True, size=10)
    
    blue_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    orange_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    current_row = 1
    
    # Title
    ws.cell(row=current_row, column=1, value=f"{model_name.upper()} STEP 2 (PHASE 4) BALANCED REFINEMENT EVALUATION REPORT")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 3
    
    # Executive Summary
    ws.cell(row=current_row, column=1, value="EXECUTIVE SUMMARY")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    summary_data = [
        ['Metric', 'Value', 'Formula', 'Explanation'],
        ['Overall Combined Score', f'{combined_score:.2f}', 'Weighted combination of similarity, quality, size optimization, and efficiency', 'Composite score for Step 2 refinement performance'],
        ['Benchmark Similarity', f'{similarity_score:.2f}%', 'Average similarity to 15 benchmark topics', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'{quality_score:.1f}%', f'{high_quality_matches} / {num_refined_clusters}', 'Percentage of clusters with >70% similarity to benchmarks'],
        ['High Quality Matches', f'{high_quality_matches}', 'Clusters with >70% similarity', 'Number of excellent cluster matches'],
        ['Clusters Generated', f'{num_refined_clusters}', 'Final refined cluster count', 'Number of clusters after refinement'],
        ['Cluster Reduction', f'{cluster_reduction}', f'{num_step1_clusters} - {num_refined_clusters}', 'Number of clusters reduced from Step 1'],
        ['Operations Performed', f'{total_operations}', f'{merge_ops} merge + {split_ops} split + {refine_ops} refine', 'Total refinement operations performed'],
        ['Average Cluster Size', f'{avg_cluster_size:.1f}', 'Mean messages per cluster', 'Average size of refined clusters'],
        ['Size Optimization', f'{size_optimization:.1f}%', '100 - |size_reduction|', 'How well cluster sizes were optimized']
    ]
    
    for row_data in summary_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 5:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Combined Score Explanation
    ws.cell(row=current_row, column=1, value="🔍 DETAILED EXPLANATION OF COMBINED SCORE")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    explanation_data = [
        ['Component', 'Score', 'Weight', 'Contribution', 'Description'],
        ['Benchmark Similarity', f'{similarity_score:.2f}%', '35%', f'{similarity_score * 0.35:.2f}', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'{quality_score:.1f}%', '25%', f'{quality_score * 0.25:.2f}', 'Percentage of high-quality matches (>70% similarity)'],
        ['Size Optimization', f'{size_optimization:.1f}%', '20%', f'{size_optimization * 0.20:.2f}', 'How well cluster sizes were optimized'],
        ['Operation Efficiency', f'{operation_efficiency:.1f}%', '20%', f'{operation_efficiency * 0.20:.2f}', 'Efficiency of refinement operations'],
        ['TOTAL COMBINED SCORE', f'{combined_score:.2f}', '100%', f'{combined_score:.2f}', 'Overall Step 2 performance score']
    ]
    
    for row_data in explanation_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 10:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif 'TOTAL' in str(value):
                cell.fill = yellow_fill
                cell.font = result_font
            else:
                cell.fill = light_green_fill
        current_row += 1
    
    # Auto-adjust column widths
    for col_num in range(1, 14):  # A to M
        max_length = 0
        column_letter = get_column_letter(col_num)
        for row_num in range(1, current_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename
    model_filename = model_name.replace('/', '_').replace('-', '_')
    output_path = os.path.join(output_dir, f"{model_filename}_step2_analysis.xlsx")
    
    # Save workbook
    wb.save(output_path)
    
    print(f"✅ Created Step 2 analysis: {output_path}")
    return output_path

def main():
    """Main function"""
    
    if len(sys.argv) != 2:
        print("Usage: python create_single_model_analysis.py <model_name>")
        print("Example: python create_single_model_analysis.py google_gemini-2.0-flash-001")
        print()
        print("Available models:")
        
        # Show available models
        try:
            with open("output/phase4_comprehensive_evaluation/detailed_evaluation_results.json", 'r') as f:
                results = json.load(f)
            
            successful_models = [r for r in results if r.get('success', False)]
            for i, result in enumerate(successful_models[:10], 1):  # Show first 10
                print(f"  {i:2d}. {result['model']}")
            
            if len(successful_models) > 10:
                print(f"     ... and {len(successful_models) - 10} more models")
                
        except Exception as e:
            print(f"Error loading model list: {e}")
        
        return
    
    model_name = sys.argv[1]
    output_dir = "output/step2_individual_analysis"
    
    # Create the analysis
    result_path = create_single_model_excel(model_name, output_dir)
    
    if result_path:
        print(f"\n🎉 Successfully created analysis for {model_name}")
        print(f"📁 File saved to: {result_path}")
    else:
        print(f"\n❌ Failed to create analysis for {model_name}")

if __name__ == "__main__":
    main()
