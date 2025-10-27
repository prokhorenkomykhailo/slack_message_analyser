#!/usr/bin/env python3
"""
Comprehensive Verification Guide for Excel Files
Shows exactly how to verify each value in the Excel file
"""

import json
import os
from openpyxl import load_workbook

def load_model_data(model_name):
    """Load specific model data from JSON results"""
    
    results_file = "output/phase4_comprehensive_evaluation/detailed_evaluation_results.json"
    
    try:
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Find the specific model
        for result in results:
            if result['model'] == model_name:
                return result
        
        return None
        
    except Exception as e:
        print(f"❌ Error loading results: {e}")
        return None

def comprehensive_verification_guide(model_name, excel_file_path):
    """Provide comprehensive verification guide"""
    
    print(f"📋 COMPREHENSIVE VERIFICATION GUIDE")
    print(f"Model: {model_name}")
    print(f"Excel File: {excel_file_path}")
    print("=" * 80)
    
    # Load source data
    source_data = load_model_data(model_name)
    if not source_data:
        print(f"❌ Model data not found: {model_name}")
        return
    
    # Load Excel file
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return
    
    print(f"🎯 SOURCE DATA VERIFICATION")
    print(f"   File: output/phase4_comprehensive_evaluation/detailed_evaluation_results.json")
    print(f"   Model: {source_data['model']}")
    print()
    
    # Calculate expected values
    avg_similarity = source_data['avg_combined_similarity']
    quality_rate = source_data['quality_rate']
    total_operations = source_data['total_operations']
    num_refined_clusters = source_data['num_refined_clusters']
    size_reduction = source_data['size_reduction']
    
    similarity_score = avg_similarity * 100
    quality_score = quality_rate * 100
    operation_efficiency = max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))) if num_refined_clusters > 0 else 100
    size_optimization = max(0, min(100, 100 - abs(size_reduction)))
    
    combined_score = (
        similarity_score * 0.35 +
        quality_score * 0.25 +
        size_optimization * 0.20 +
        operation_efficiency * 0.20
    )
    
    print(f"📊 HOW TO VERIFY EACH VALUE:")
    print(f"   Open the Excel file and check these locations:")
    print()
    
    # Executive Summary Section (Rows 6-15)
    print(f"🔍 EXECUTIVE SUMMARY SECTION (Rows 6-15):")
    print(f"   Row 7, Col 2: Overall Combined Score")
    print(f"     Expected: {combined_score:.2f}")
    print(f"     Excel Value: {ws.cell(row=7, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=7, column=2).value) == f'{combined_score:.2f}' else '❌ CHECK'}")
    print()
    
    print(f"   Row 8, Col 2: Benchmark Similarity")
    print(f"     Expected: {similarity_score:.2f}%")
    print(f"     Excel Value: {ws.cell(row=8, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=8, column=2).value) == f'{similarity_score:.2f}%' else '❌ CHECK'}")
    print()
    
    print(f"   Row 9, Col 2: Quality Rate")
    print(f"     Expected: {quality_score:.1f}%")
    print(f"     Excel Value: {ws.cell(row=9, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=9, column=2).value) == f'{quality_score:.1f}%' else '❌ CHECK'}")
    print()
    
    print(f"   Row 10, Col 2: High Quality Matches")
    print(f"     Expected: {source_data['high_quality_matches']}")
    print(f"     Excel Value: {ws.cell(row=10, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=10, column=2).value) == str(source_data['high_quality_matches']) else '❌ CHECK'}")
    print()
    
    print(f"   Row 11, Col 2: Clusters Generated")
    print(f"     Expected: {source_data['num_refined_clusters']}")
    print(f"     Excel Value: {ws.cell(row=11, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=11, column=2).value) == str(source_data['num_refined_clusters']) else '❌ CHECK'}")
    print()
    
    print(f"   Row 12, Col 2: Cluster Reduction")
    print(f"     Expected: {source_data['cluster_count_reduction']}")
    print(f"     Excel Value: {ws.cell(row=12, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=12, column=2).value) == str(source_data['cluster_count_reduction']) else '❌ CHECK'}")
    print()
    
    print(f"   Row 13, Col 2: Operations Performed")
    print(f"     Expected: {source_data['total_operations']}")
    print(f"     Excel Value: {ws.cell(row=13, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=13, column=2).value) == str(source_data['total_operations']) else '❌ CHECK'}")
    print()
    
    print(f"   Row 14, Col 2: Average Cluster Size")
    print(f"     Expected: {source_data['avg_cluster_size']:.1f}")
    print(f"     Excel Value: {ws.cell(row=14, column=2).value}")
    expected_avg_size = f"{source_data['avg_cluster_size']:.1f}"
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=14, column=2).value) == expected_avg_size else '❌ CHECK'}")
    print()
    
    print(f"   Row 15, Col 2: Size Optimization")
    print(f"     Expected: {size_optimization:.1f}%")
    print(f"     Excel Value: {ws.cell(row=15, column=2).value}")
    print(f"     ✅ {'CORRECT' if str(ws.cell(row=15, column=2).value) == f'{size_optimization:.1f}%' else '❌ CHECK'}")
    print()
    
    # Detailed Explanation Section
    print(f"🔍 DETAILED EXPLANATION SECTION (Rows 20-25):")
    print(f"   Row 21, Col 2: Benchmark Similarity Score")
    print(f"     Expected: {similarity_score:.2f}%")
    print(f"     Excel Value: {ws.cell(row=21, column=2).value}")
    print()
    
    print(f"   Row 22, Col 2: Quality Rate Score")
    print(f"     Expected: {quality_score:.2f}%")
    print(f"     Excel Value: {ws.cell(row=22, column=2).value}")
    print()
    
    print(f"   Row 23, Col 2: Size Optimization Score")
    print(f"     Expected: {size_optimization:.1f}%")
    print(f"     Excel Value: {ws.cell(row=23, column=2).value}")
    print()
    
    print(f"   Row 24, Col 2: Operation Efficiency Score")
    print(f"     Expected: {operation_efficiency:.1f}%")
    print(f"     Excel Value: {ws.cell(row=24, column=2).value}")
    print()
    
    print(f"   Row 25, Col 2: TOTAL COMBINED SCORE")
    print(f"     Expected: {combined_score:.2f}")
    print(f"     Excel Value: {ws.cell(row=25, column=2).value}")
    print()
    
    # Source Data Verification
    print(f"📋 SOURCE DATA VERIFICATION:")
    print(f"   📁 File: output/phase4_comprehensive_evaluation/detailed_evaluation_results.json")
    print(f"   🔍 Search for: \"model\": \"{model_name}\"")
    print(f"   📊 Key values to verify:")
    print(f"     - avg_combined_similarity: {source_data['avg_combined_similarity']:.6f}")
    print(f"     - quality_rate: {source_data['quality_rate']:.6f}")
    print(f"     - high_quality_matches: {source_data['high_quality_matches']}")
    print(f"     - num_refined_clusters: {source_data['num_refined_clusters']}")
    print(f"     - cluster_count_reduction: {source_data['cluster_count_reduction']}")
    print(f"     - total_operations: {source_data['total_operations']}")
    print(f"     - avg_cluster_size: {source_data['avg_cluster_size']:.6f}")
    print(f"     - size_reduction: {source_data['size_reduction']:.6f}")
    print()
    
    # Formula Verification
    print(f"🧮 FORMULA VERIFICATION:")
    print(f"   Similarity Score = avg_combined_similarity × 100")
    print(f"                  = {avg_similarity:.6f} × 100 = {similarity_score:.2f}%")
    print()
    
    print(f"   Quality Score = quality_rate × 100")
    print(f"                = {quality_rate:.6f} × 100 = {quality_score:.2f}%")
    print()
    
    print(f"   Operation Efficiency = max(0, min(100, 100 - (total_operations / num_refined_clusters × 10)))")
    print(f"                       = max(0, min(100, 100 - ({total_operations} / {num_refined_clusters} × 10)))")
    print(f"                       = max(0, min(100, 100 - ({total_operations / num_refined_clusters * 10:.2f})))")
    print(f"                       = {operation_efficiency:.2f}%")
    print()
    
    print(f"   Size Optimization = max(0, min(100, 100 - |size_reduction|))")
    print(f"                    = max(0, min(100, 100 - |{size_reduction:.6f}|))")
    print(f"                    = max(0, min(100, 100 - {abs(size_reduction):.6f}))")
    print(f"                    = {size_optimization:.2f}%")
    print()
    
    print(f"   Combined Score = (similarity_score × 0.35) + (quality_score × 0.25) + (size_optimization × 0.20) + (operation_efficiency × 0.20)")
    print(f"                 = ({similarity_score:.2f} × 0.35) + ({quality_score:.2f} × 0.25) + ({size_optimization:.2f} × 0.20) + ({operation_efficiency:.2f} × 0.20)")
    print(f"                 = {similarity_score * 0.35:.2f} + {quality_score * 0.25:.2f} + {size_optimization * 0.20:.2f} + {operation_efficiency * 0.20:.2f}")
    print(f"                 = {combined_score:.2f}")
    print()
    
    print(f"✅ VERIFICATION COMPLETE!")
    print(f"   All values should match between the Excel file and source JSON data.")

def main():
    """Main verification function"""
    
    model_name = "google_gemini-2.0-flash-001"
    excel_file = "output/step2_client_analysis/gemini_2.0_flash_001_step2_updated_analysis_FIXED.xlsx"
    
    if os.path.exists(excel_file):
        comprehensive_verification_guide(model_name, excel_file)
    else:
        print(f"❌ Excel file not found: {excel_file}")

if __name__ == "__main__":
    main()
