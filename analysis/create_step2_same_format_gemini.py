#!/usr/bin/env python3
"""
Create Step 2 Analysis Using Same Format as Step 1
Uses the same formulas and structure as the Step 1 analysis files
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

def load_gemini_step2_data():
    """Load specific gemini_2.0_flash_001 data from Step 2 JSON results"""
    
    results_file = "output/phase4_comprehensive_evaluation/detailed_evaluation_results.json"
    
    try:
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Find the specific model
        for result in results:
            if result['model'] == 'google_gemini-2.0-flash-001':
                return result
        
        print(f"❌ Model google_gemini-2.0-flash-001 not found in results")
        return None
        
    except Exception as e:
        print(f"❌ Error loading results: {e}")
        return None

def create_step2_same_format_excel():
    """Create Step 2 Excel file using the same format as Step 1"""
    
    print("🎯 Creating Step 2 Analysis Using Same Format as Step 1")
    print("=" * 70)
    
    # Load model data
    model_data = load_gemini_step2_data()
    if not model_data:
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gemini 2.0 Flash 001 Step 2 Analysis"
    
    # Define styles (same as Step 1 format)
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    result_font = Font(bold=True, size=11, color="0000FF")
    formula_font = Font(italic=True, size=10)
    
    blue_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    orange_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    current_row = 1
    
    # Title (same format as Step 1)
    ws.cell(row=current_row, column=1, value="GOOGLE GEMINI-2.0-FLASH-001 STEP 2 (PHASE 4) BALANCED REFINEMENT EVALUATION REPORT")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 3
    
    # Executive Summary (same format as Step 1)
    ws.cell(row=current_row, column=1, value="EXECUTIVE SUMMARY")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    # Extract key metrics (using same calculations as Step 1)
    benchmark_eval = model_data.get('benchmark_evaluation', {})
    
    # Calculate metrics using same formulas as Step 1
    avg_combined_similarity = benchmark_eval.get('avg_combined_similarity', 0)
    quality_rate = benchmark_eval.get('quality_rate', 0)
    high_quality_matches = benchmark_eval.get('high_quality_matches', 0)
    num_refined_clusters = model_data.get('num_refined_clusters', 0)
    num_step1_clusters = model_data.get('num_step1_clusters', 0)
    cluster_count_reduction = model_data.get('cluster_count_reduction', 0)
    total_operations = model_data.get('total_operations', 0)
    avg_cluster_size = model_data.get('avg_cluster_size', 0)
    size_reduction = model_data.get('size_reduction', 0)
    
    # Calculate scores using same formulas as Step 1
    similarity_score = avg_combined_similarity * 100
    quality_score = quality_rate * 100
    operation_efficiency = max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))) if num_refined_clusters > 0 else 100
    size_optimization = max(0, min(100, 100 - abs(size_reduction)))
    
    # Combined score (same weighting as Step 1)
    combined_score = (
        similarity_score * 0.35 +        # 35% - similarity to benchmark
        quality_score * 0.25 +           # 25% - quality rate
        size_optimization * 0.20 +       # 20% - size optimization
        operation_efficiency * 0.20      # 20% - operation efficiency
    )
    
    # Summary data (same format as Step 1)
    summary_data = [
        ['Metric', 'Value', 'Formula', 'Explanation'],
        ['Overall Combined Score', f'{combined_score:.2f}', 'Weighted combination of similarity, quality, size optimization, and efficiency', 'Composite score for Step 2 refinement performance'],
        ['Benchmark Similarity', f'{similarity_score:.2f}%', 'Average similarity to 15 benchmark topics', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'{quality_score:.1f}%', f'{high_quality_matches} / {num_refined_clusters}', 'Percentage of clusters with >70% similarity to benchmarks'],
        ['High Quality Matches', f'{high_quality_matches}', 'Clusters with >70% similarity', 'Number of excellent cluster matches'],
        ['Clusters Generated', f'{num_refined_clusters}', 'Final refined cluster count', 'Number of clusters after refinement'],
        ['Cluster Reduction', f'{cluster_count_reduction}', f'{num_step1_clusters} - {num_refined_clusters}', 'Number of clusters reduced from Step 1'],
        ['Operations Performed', f'{total_operations}', f'{model_data.get("merge_operations", 0)} merge + {model_data.get("split_operations", 0)} split + {model_data.get("refine_operations", 0)} refine', 'Total refinement operations performed'],
        ['Average Cluster Size', f'{avg_cluster_size:.1f}', 'Mean messages per cluster', 'Average size of refined clusters'],
        ['Size Optimization', f'{size_optimization:.1f}%', '100 - |size_reduction|', 'How well cluster sizes were optimized']
    ]
    
    for row_data in summary_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 5:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Combined Score Explanation (same format as Step 1)
    ws.cell(row=current_row, column=1, value="🔍 DETAILED EXPLANATION OF COMBINED SCORE")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    explanation_data = [
        ['Component', 'Score', 'Weight', 'Contribution', 'Description'],
        ['Benchmark Similarity', f'{similarity_score:.2f}%', '35%', f'{similarity_score * 0.35:.2f}', 'How well refined clusters match benchmark topics'],
        ['Quality Rate', f'{quality_score:.1f}%', '25%', f'{quality_score * 0.25:.2f}', 'Percentage of high-quality matches (>70% similarity)'],
        ['Size Optimization', f'{size_optimization:.1f}%', '20%', f'{size_optimization * 0.20:.2f}', 'How well cluster sizes were optimized'],
        ['Operation Efficiency', f'{operation_efficiency:.1f}%', '20%', f'{operation_efficiency * 0.20:.2f}', 'Efficiency of refinement operations'],
        ['TOTAL COMBINED SCORE', f'{combined_score:.2f}', '100%', f'{combined_score:.2f}', 'Overall Step 2 performance score']
    ]
    
    for row_data in explanation_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 10:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif 'TOTAL' in str(value):
                cell.fill = yellow_fill
                cell.font = result_font
            else:
                cell.fill = light_green_fill
        current_row += 1
    
    current_row += 2
    
    # Refinement Operations Analysis (same format as Step 1)
    ws.cell(row=current_row, column=1, value="🔧 REFINEMENT OPERATIONS ANALYSIS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    merge_ops = model_data.get('merge_operations', 0)
    split_ops = model_data.get('split_operations', 0)
    refine_ops = model_data.get('refine_operations', 0)
    
    operations_data = [
        ['Operation Type', 'Count', 'Percentage', 'Description'],
        ['Merge Operations', f'{merge_ops}', f'{(merge_ops/total_operations*100):.1f}%' if total_operations > 0 else '0%', 'Combining related clusters'],
        ['Split Operations', f'{split_ops}', f'{(split_ops/total_operations*100):.1f}%' if total_operations > 0 else '0%', 'Separating mixed clusters'],
        ['Refine Operations', f'{refine_ops}', f'{(refine_ops/total_operations*100):.1f}%' if total_operations > 0 else '0%', 'Improving cluster quality'],
        ['Total Operations', f'{total_operations}', '100%', 'All refinement operations performed']
    ]
    
    for row_data in operations_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 15:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif 'Total' in str(value):
                cell.fill = yellow_fill
                cell.font = result_font
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Cluster Analysis (same format as Step 1)
    ws.cell(row=current_row, column=1, value="📊 CLUSTER REFINEMENT ANALYSIS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    cluster_data = [
        ['Metric', 'Step 1', 'Step 2', 'Change', 'Improvement %'],
        ['Total Clusters', f'{num_step1_clusters}', f'{num_refined_clusters}', f'{cluster_count_reduction}', f'{(cluster_count_reduction/num_step1_clusters*100):.1f}%' if num_step1_clusters > 0 else '0%'],
        ['Average Size', f'{model_data.get("avg_step1_size", 0):.1f}', f'{avg_cluster_size:.1f}', f'{size_reduction:.1f}', f'{(abs(size_reduction)/model_data.get("avg_step1_size", 1)*100):.1f}%' if model_data.get("avg_step1_size", 0) > 0 else '0%'],
        ['Size Standard Dev', 'N/A', f'{model_data.get("size_std_reduction", 0):.1f}', 'N/A', 'Size consistency improvement']
    ]
    
    for row_data in cluster_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 20:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_green_fill
        current_row += 1
    
    current_row += 2
    
    # Individual Cluster Analysis (same format as Step 1)
    ws.cell(row=current_row, column=1, value="📋 INDIVIDUAL CLUSTER ANALYSIS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    # Headers
    headers = ['Cluster', 'Title', 'Best Benchmark Match', 'Similarity %', 'Quality Status', 'Participants', 'Message Count']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = green_fill
        cell.font = header_font
    current_row += 1
    
    # Cluster data from benchmark evaluation
    best_matches = benchmark_eval.get('best_matches', [])
    cluster_titles = model_data.get('cluster_titles', [])
    
    for i, cluster_title in enumerate(cluster_titles):
        if i < len(best_matches):
            match_info = best_matches[i]
            similarity = match_info.get('title_similarity', 0) * 100
            benchmark_match = match_info.get('best_benchmark_match', '')
            is_high_quality = 'YES' if match_info.get('combined_similarity', 0) > 0.7 else 'NO'
        else:
            similarity = 0
            benchmark_match = 'No match found'
            is_high_quality = 'NO'
        
        cluster_data = [
            f"Cluster {i+1}",
            cluster_title,
            benchmark_match,
            f"{similarity:.1f}%",
            is_high_quality,
            "Multiple participants",  # Could be enhanced with actual participant data
            "Variable count"  # Could be enhanced with actual message count data
        ]
        
        for col, value in enumerate(cluster_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if is_high_quality == 'YES':
                cell.fill = light_green_fill
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Performance Comparison (same format as Step 1)
    ws.cell(row=current_row, column=1, value="🏆 PERFORMANCE ASSESSMENT")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    performance_data = [
        ['Performance Level', 'Score Range', 'Current Score', 'Status', 'Recommendation'],
        ['Excellent', '90-100', f'{combined_score:.2f}', '✅ EXCELLENT' if combined_score >= 90 else '❌', 'Ready for production use'],
        ['Very Good', '80-89', f'{combined_score:.2f}', '✅ VERY GOOD' if 80 <= combined_score < 90 else '❌', 'Minor optimization needed'],
        ['Good', '70-79', f'{combined_score:.2f}', '✅ GOOD' if 70 <= combined_score < 80 else '❌', 'Moderate improvement recommended'],
        ['Fair', '60-69', f'{combined_score:.2f}', '⚠️ FAIR' if 60 <= combined_score < 70 else '❌', 'Significant improvement needed'],
        ['Poor', '0-59', f'{combined_score:.2f}', '❌ POOR' if combined_score < 60 else '❌', 'Major reconfiguration required']
    ]
    
    for row_data in performance_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 27:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            elif '✅' in str(value):
                cell.fill = light_green_fill
                cell.font = Font(color="008000")
            elif '⚠️' in str(value):
                cell.fill = yellow_fill
                cell.font = Font(color="FF8000")
            elif '❌' in str(value):
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                cell.font = Font(color="FF0000")
            else:
                cell.fill = light_blue_fill
        current_row += 1
    
    current_row += 2
    
    # Key Insights (same format as Step 1)
    ws.cell(row=current_row, column=1, value="💡 KEY INSIGHTS & RECOMMENDATIONS")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).fill = blue_fill
    ws.merge_cells(f'A{current_row}:M{current_row}')
    current_row += 2
    
    insights_data = [
        ['Insight', 'Value', 'Impact', 'Recommendation'],
        ['Overall Performance', f'{combined_score:.2f}/100', 'Excellent' if combined_score >= 90 else 'Good' if combined_score >= 70 else 'Needs Improvement', 'Model ready for production' if combined_score >= 90 else 'Consider optimization'],
        ['Benchmark Alignment', f'{similarity_score:.1f}%', 'High' if similarity_score >= 80 else 'Moderate' if similarity_score >= 60 else 'Low', 'Excellent topic matching' if similarity_score >= 80 else 'Review clustering strategy'],
        ['Quality Consistency', f'{quality_score:.1f}%', 'High' if quality_score >= 80 else 'Moderate' if quality_score >= 60 else 'Low', 'Consistent high-quality results' if quality_score >= 80 else 'Focus on quality improvement'],
        ['Refinement Efficiency', f'{operation_efficiency:.1f}%', 'Efficient' if operation_efficiency >= 80 else 'Moderate' if operation_efficiency >= 60 else 'Inefficient', 'Optimal operation usage' if operation_efficiency >= 80 else 'Review operation strategy'],
        ['Size Optimization', f'{size_optimization:.1f}%', 'Well Optimized' if size_optimization >= 80 else 'Moderate' if size_optimization >= 60 else 'Poor', 'Good size management' if size_optimization >= 80 else 'Improve size consistency']
    ]
    
    for row_data in insights_data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if current_row == 34:  # Header row
                cell.fill = green_fill
                cell.font = header_font
            else:
                cell.fill = light_green_fill
        current_row += 1
    
    # Auto-adjust column widths
    for col_num in range(1, 14):  # A to M
        max_length = 0
        column_letter = get_column_letter(col_num)
        for row_num in range(1, current_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    output_path = "output/step2_client_analysis/gemini_2.0_flash_001_step2_analysis_SAME_FORMAT_AS_STEP1.xlsx"
    wb.save(output_path)
    
    print(f"✅ Step 2 analysis using same format as Step 1 created: {output_path}")
    return output_path

def main():
    """Main function"""
    
    print("🎯 Creating Step 2 Analysis Using Same Format as Step 1")
    print("=" * 70)
    print("This will create a Step 2 analysis that:")
    print("  ✅ Uses the same format and structure as Step 1")
    print("  ✅ Uses the same formulas and calculations")
    print("  ✅ Uses the same color scheme and layout")
    print("  ✅ Provides the same level of detail and analysis")
    print()
    
    # Create the Step 2 analysis
    result_path = create_step2_same_format_excel()
    
    if result_path:
        print(f"\n🎉 Successfully created Step 2 analysis!")
        print(f"📁 File saved to: {result_path}")
        print("\n📋 The file includes:")
        print("  ✅ Same format as Step 1 analysis files")
        print("  ✅ Same formulas and calculations")
        print("  ✅ Same color scheme and structure")
        print("  ✅ Comprehensive analysis of Step 2 performance")
        print("  ✅ Easy comparison with Step 1 results")
    else:
        print("\n❌ Failed to create Step 2 analysis")

if __name__ == "__main__":
    main()
