#!/usr/bin/env python3
"""
Verify Excel Values for Individual Model Analysis
Helps verify all calculated values in the Excel file against source data
"""

import pandas as pd
import json
import os
from openpyxl import load_workbook

def load_model_data(model_name):
    """Load specific model data from JSON results"""
    
    results_file = "output/phase4_comprehensive_evaluation/detailed_evaluation_results.json"
    
    try:
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Find the specific model
        for result in results:
            if result['model'] == model_name:
                return result
        
        print(f"❌ Model {model_name} not found in results")
        return None
        
    except Exception as e:
        print(f"❌ Error loading results: {e}")
        return None

def verify_excel_values(model_name, excel_file_path):
    """Verify all values in the Excel file"""
    
    print(f"🔍 Verifying Excel values for: {model_name}")
    print("=" * 70)
    
    # Load source data
    source_data = load_model_data(model_name)
    if not source_data:
        return
    
    # Load Excel file
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return
    
    print(f"📊 Source Data Summary:")
    print(f"  - Model: {source_data['model']}")
    print(f"  - Success: {source_data['success']}")
    print(f"  - Duration: {source_data['duration']:.3f}s")
    print(f"  - Refined Clusters: {source_data['num_refined_clusters']}")
    print(f"  - Step 1 Clusters: {source_data['num_step1_clusters']}")
    print(f"  - Cluster Reduction: {source_data['cluster_count_reduction']}")
    print(f"  - Total Operations: {source_data['total_operations']}")
    print(f"  - Avg Combined Similarity: {source_data['avg_combined_similarity']:.4f}")
    print(f"  - Quality Rate: {source_data['quality_rate']:.4f}")
    print(f"  - High Quality Matches: {source_data['high_quality_matches']}")
    print()
    
    # Calculate expected scores
    avg_similarity = source_data['avg_combined_similarity']
    quality_rate = source_data['quality_rate']
    total_operations = source_data['total_operations']
    num_refined_clusters = source_data['num_refined_clusters']
    size_reduction = source_data['size_reduction']
    
    similarity_score = avg_similarity * 100
    quality_score = quality_rate * 100
    operation_efficiency = max(0, min(100, 100 - (total_operations / num_refined_clusters * 10))) if num_refined_clusters > 0 else 100
    size_optimization = max(0, min(100, 100 - abs(size_reduction)))
    
    combined_score = (
        similarity_score * 0.35 +
        quality_score * 0.25 +
        size_optimization * 0.20 +
        operation_efficiency * 0.20
    )
    
    print(f"🎯 Expected Calculated Values:")
    print(f"  - Similarity Score: {similarity_score:.2f}%")
    print(f"  - Quality Score: {quality_score:.2f}%")
    print(f"  - Operation Efficiency: {operation_efficiency:.2f}%")
    print(f"  - Size Optimization: {size_optimization:.2f}%")
    print(f"  - Combined Score: {combined_score:.2f}")
    print()
    
    # Verify key cells in Excel
    print(f"📋 Verifying Excel Cell Values:")
    
    # Find key values in Excel (these are approximate row positions)
    verification_points = [
        (5, 2, "Overall Combined Score", f"{combined_score:.2f}"),
        (6, 2, "Benchmark Similarity", f"{similarity_score:.2f}%"),
        (7, 2, "Quality Rate", f"{quality_score:.1f}%"),
        (8, 2, "High Quality Matches", f"{source_data['high_quality_matches']}"),
        (9, 2, "Clusters Generated", f"{source_data['num_refined_clusters']}"),
        (10, 2, "Cluster Reduction", f"{source_data['cluster_count_reduction']}"),
        (11, 2, "Operations Performed", f"{source_data['total_operations']}"),
        (12, 2, "Average Cluster Size", f"{source_data['avg_cluster_size']:.1f}"),
    ]
    
    for row, col, description, expected_value in verification_points:
        try:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                cell_str = str(cell_value)
                if cell_str.strip() == expected_value.strip():
                    print(f"  ✅ {description}: {cell_value} (Row {row}, Col {col})")
                else:
                    print(f"  ⚠️ {description}: Expected '{expected_value}', Found '{cell_value}' (Row {row}, Col {col})")
            else:
                print(f"  ❌ {description}: No value found (Row {row}, Col {col})")
        except Exception as e:
            print(f"  ❌ {description}: Error reading cell - {e}")
    
    print()
    
    # Verify benchmark evaluation data
    benchmark_eval = source_data.get('benchmark_evaluation', {})
    print(f"📊 Benchmark Evaluation Verification:")
    print(f"  - Number of refined clusters: {benchmark_eval.get('num_refined_clusters', 'N/A')}")
    print(f"  - Number of benchmark topics: {benchmark_eval.get('num_benchmark_topics', 'N/A')}")
    print(f"  - Average title similarity: {benchmark_eval.get('avg_title_similarity', 0):.4f}")
    print(f"  - Average combined similarity: {benchmark_eval.get('avg_combined_similarity', 0):.4f}")
    print(f"  - High quality matches: {benchmark_eval.get('high_quality_matches', 'N/A')}")
    print(f"  - Quality rate: {benchmark_eval.get('quality_rate', 0):.4f}")
    print()
    
    # Verify cluster titles
    cluster_titles = source_data.get('cluster_titles', [])
    print(f"📋 Cluster Titles Verification:")
    print(f"  - Number of cluster titles: {len(cluster_titles)}")
    for i, title in enumerate(cluster_titles, 1):
        if title:
            print(f"    {i:2d}. {title}")
        else:
            print(f"    {i:2d}. [Empty Title]")
    print()
    
    # Verify operations breakdown
    print(f"🔧 Operations Breakdown:")
    print(f"  - Merge Operations: {source_data.get('merge_operations', 0)}")
    print(f"  - Split Operations: {source_data.get('split_operations', 0)}")
    print(f"  - Refine Operations: {source_data.get('refine_operations', 0)}")
    print(f"  - Total Operations: {source_data.get('total_operations', 0)}")
    print()
    
    # Verify size metrics
    print(f"📏 Size Metrics:")
    print(f"  - Average cluster size: {source_data.get('avg_cluster_size', 0):.3f}")
    print(f"  - Average Step 1 size: {source_data.get('avg_step1_size', 0):.3f}")
    print(f"  - Size reduction: {source_data.get('size_reduction', 0):.3f}")
    print(f"  - Size std reduction: {source_data.get('size_std_reduction', 0):.3f}")
    print()
    
    print(f"✅ Verification complete for {model_name}")

def main():
    """Main verification function"""
    
    # Example: Verify gemini-2.0-flash-001
    model_name = "google_gemini-2.0-flash-001"
    excel_file = "output/step2_client_analysis/gemini_2.0_flash_001_step2_updated_analysis_FIXED.xlsx"
    
    if os.path.exists(excel_file):
        verify_excel_values(model_name, excel_file)
    else:
        print(f"❌ Excel file not found: {excel_file}")
        print("Available files:")
        if os.path.exists("output/step2_client_analysis"):
            for file in os.listdir("output/step2_client_analysis"):
                if file.endswith('.xlsx'):
                    print(f"  - {file}")

if __name__ == "__main__":
    main()





















