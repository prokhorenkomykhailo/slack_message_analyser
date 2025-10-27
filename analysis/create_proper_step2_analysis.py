#!/usr/bin/env python3
"""
PROPER Step 2 Analysis using Step 1 methodology
Correctly calculates metrics when model has 13 clusters vs benchmark 15 clusters
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

def load_balanced_refinement_data():
    """Load the balanced refinement model output"""
    file_path = "output/phase4_balanced_refinement/google_gemini-2.0-flash-001_balanced.json"
    with open(file_path, 'r') as f:
        return json.load(f)

def load_benchmark_clusters():
    """Load the benchmark clusters for comparison"""
    file_path = "phases/phase4_clusters_refined.json"
    with open(file_path, 'r') as f:
        return json.load(f)

def calculate_proper_step1_metrics(model_data, benchmark_data):
    """Calculate metrics using Step 1 methodology"""
    
    model_clusters = model_data['refined_clusters']
    benchmark_clusters = benchmark_data
    
    print(f"Model has {len(model_clusters)} refined clusters")
    print(f"Benchmark has {len(benchmark_clusters)} clusters")
    
    # Create message ID sets for each cluster
    model_cluster_messages = {}
    for cluster in model_clusters:
        model_cluster_messages[cluster['cluster_id']] = set(cluster['message_ids'])
    
    benchmark_cluster_messages = {}
    for cluster in benchmark_clusters:
        benchmark_cluster_messages[cluster['cluster_id']] = set(cluster['message_ids'])
    
    # Find best matches for each model cluster
    model_matches = []
    for model_cluster in model_clusters:
        model_id = model_cluster['cluster_id']
        model_messages = model_cluster_messages[model_id]
        
        # Find best matching benchmark cluster
        best_match = None
        best_overlap = 0
        best_benchmark_id = None
        
        for benchmark_cluster in benchmark_clusters:
            benchmark_id = benchmark_cluster['cluster_id']
            benchmark_messages = benchmark_cluster_messages[benchmark_id]
            
            overlap = len(model_messages.intersection(benchmark_messages))
            if overlap > best_overlap:
                best_overlap = overlap
                best_match = benchmark_cluster
                best_benchmark_id = benchmark_id
        
        if best_match:
            benchmark_messages = benchmark_cluster_messages[best_benchmark_id]
            matched_messages = len(model_messages.intersection(benchmark_messages))
            missing_messages = len(benchmark_messages - model_messages)
            extra_messages = len(model_messages - benchmark_messages)
            
            precision = (matched_messages / len(model_messages)) * 100 if len(model_messages) > 0 else 0
            coverage = (matched_messages / len(benchmark_messages)) * 100 if len(benchmark_messages) > 0 else 0
            deviation = abs(len(model_messages) - len(benchmark_messages)) / len(benchmark_messages) * 100 if len(benchmark_messages) > 0 else 0
            
            model_matches.append({
                'model_id': model_id,
                'model_title': model_cluster['draft_title'],
                'model_messages': len(model_messages),
                'benchmark_id': best_benchmark_id,
                'benchmark_title': best_match['draft_title'],
                'benchmark_messages': len(benchmark_messages),
                'matched_messages': matched_messages,
                'missing_messages': missing_messages,
                'extra_messages': extra_messages,
                'precision': precision,
                'coverage': coverage,
                'deviation': deviation,
                'perfect_match': matched_messages == len(benchmark_messages) and len(model_messages) == len(benchmark_messages)
            })
    
    # Find unmatched benchmark clusters
    matched_benchmark_ids = {match['benchmark_id'] for match in model_matches}
    unmatched_benchmarks = []
    for benchmark_cluster in benchmark_clusters:
        if benchmark_cluster['cluster_id'] not in matched_benchmark_ids:
            unmatched_benchmarks.append({
                'benchmark_id': benchmark_cluster['cluster_id'],
                'benchmark_title': benchmark_cluster['draft_title'],
                'benchmark_messages': len(benchmark_cluster['message_ids']),
                'matched_messages': 0,
                'missing_messages': len(benchmark_cluster['message_ids']),
                'extra_messages': 0,
                'precision': 0,
                'coverage': 0,
                'deviation': 100,  # 100% deviation since no match
                'perfect_match': False
            })
    
    # Combine all results
    all_results = model_matches + unmatched_benchmarks
    
    # Calculate overall metrics
    total_benchmark_messages = sum(len(cluster['message_ids']) for cluster in benchmark_clusters)
    total_model_messages = sum(len(cluster['message_ids']) for cluster in model_clusters)
    total_matched_messages = sum(result['matched_messages'] for result in all_results)
    total_missing_messages = sum(result['missing_messages'] for result in all_results)
    total_extra_messages = sum(result['extra_messages'] for result in all_results)
    
    # Step 1 methodology calculations
    cluster_count_score = min(len(model_clusters), len(benchmark_clusters)) / max(len(model_clusters), len(benchmark_clusters)) * 100
    coverage_score = (total_matched_messages / total_benchmark_messages) * 100 if total_benchmark_messages > 0 else 0
    precision_score = (total_matched_messages / total_model_messages) * 100 if total_model_messages > 0 else 0
    
    # Average deviation across all topics
    avg_deviation = sum(result['deviation'] for result in all_results) / len(all_results) if all_results else 0
    deviation_score = max(0, 100 - avg_deviation)
    
    # Final score using Step 1 formula
    final_score = (cluster_count_score * 0.25) + (coverage_score * 0.25) + (precision_score * 0.25) + (deviation_score * 0.25)
    
    return {
        'model_cluster_count': len(model_clusters),
        'benchmark_cluster_count': len(benchmark_clusters),
        'total_benchmark_messages': total_benchmark_messages,
        'total_model_messages': total_model_messages,
        'total_matched_messages': total_matched_messages,
        'total_missing_messages': total_missing_messages,
        'total_extra_messages': total_extra_messages,
        'cluster_count_score': cluster_count_score,
        'coverage_score': coverage_score,
        'precision_score': precision_score,
        'deviation_score': deviation_score,
        'avg_deviation': avg_deviation,
        'final_score': final_score,
        'detailed_results': all_results
    }

def create_proper_step2_analysis():
    """Create the PROPER Step 2 analysis using Step 1 methodology"""
    
    # Load data
    model_data = load_balanced_refinement_data()
    benchmark_data = load_benchmark_clusters()
    
    print("=== PROPER STEP 2 ANALYSIS (Using Step 1 Methodology) ===")
    
    # Calculate metrics
    metrics = calculate_proper_step1_metrics(model_data, benchmark_data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PROPER Step 2 Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=12, color="000000")
    data_font = Font(size=11)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    light_green_fill = PatternFill(start_color="E7F7E7", end_color="E7F7E7", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "PROPER Step 2 Analysis: Google Gemini 2.0 Flash 001 (Using Step 1 Methodology)"
    ws[f'A{row}'].font = Font(bold=True, size=16, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_alignment
    row += 2
    
    # Key Findings Summary
    ws[f'A{row}'] = "KEY FINDINGS SUMMARY"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    summary_data = [
        ["Model Refined Clusters", metrics['model_cluster_count'], "13 clusters after merge operation"],
        ["Benchmark Clusters", metrics['benchmark_cluster_count'], "15 original clusters"],
        ["Cluster Count Score", f"{metrics['cluster_count_score']:.2f}%", "min(13,15)/max(13,15)*100 = 86.67%"],
        ["Total Benchmark Messages", metrics['total_benchmark_messages'], "Sum of all benchmark message counts"],
        ["Total Model Messages", metrics['total_model_messages'], "Sum of all model message counts"],
        ["Total Matched Messages", metrics['total_matched_messages'], "Messages correctly matched"],
        ["Total Missing Messages", metrics['total_missing_messages'], "Messages missing from model"],
        ["Total Extra Messages", metrics['total_extra_messages'], "Extra messages in model"],
        ["Coverage Score", f"{metrics['coverage_score']:.2f}%", "Matched/Benchmark*100"],
        ["Precision Score", f"{metrics['precision_score']:.2f}%", "Matched/Model*100"],
        ["Average Deviation", f"{metrics['avg_deviation']:.2f}%", "Average deviation across topics"],
        ["Deviation Score", f"{metrics['deviation_score']:.2f}%", "max(0, 100 - avg_deviation)"],
        ["FINAL SCORE", f"{metrics['final_score']:.2f}", "Weighted sum of all components"]
    ]
    
    for summary in summary_data:
        ws[f'A{row}'] = summary[0]
        ws[f'B{row}'] = summary[1]
        ws[f'C{row}'] = summary[2]
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = data_font
        ws[f'C{row}'].font = data_font
        ws[f'A{row}'].fill = light_blue_fill
        ws[f'B{row}'].fill = light_green_fill
        ws[f'C{row}'].fill = light_yellow_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # Step 1 Component Scores
    ws[f'A{row}'] = "STEP 1: COMPONENT SCORES FOR FINAL CALCULATION"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    component_data = [
        ["Component", "Calculation", "Result Score", "Weight", "Weighted Score"],
        ["1. Cluster Count", f"min({metrics['model_cluster_count']}, {metrics['benchmark_cluster_count']}) / max({metrics['model_cluster_count']}, {metrics['benchmark_cluster_count']}) × 100", f"{metrics['cluster_count_score']:.6f}", "0.25", f"{metrics['cluster_count_score'] * 0.25:.6f}"],
        ["2. Coverage", f"Found {metrics['total_matched_messages']} of {metrics['total_benchmark_messages']} expected messages", f"{metrics['coverage_score']:.6f}", "0.25", f"{metrics['coverage_score'] * 0.25:.6f}"],
        ["3. Precision", f"All matched messages are correct", f"{metrics['precision_score']:.6f}", "0.25", f"{metrics['precision_score'] * 0.25:.6f}"],
        ["4. Deviation", f"max(0, 100 - {metrics['avg_deviation']:.6f})", f"{metrics['deviation_score']:.6f}", "0.25", f"{metrics['deviation_score'] * 0.25:.6f}"]
    ]
    
    for i, component in enumerate(component_data):
        for j, value in enumerate(component):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            if i == 0:  # Header row
                ws[f'{col}{row}'].font = subheader_font
                ws[f'{col}{row}'].fill = subheader_fill
            else:
                ws[f'{col}{row}'].fill = light_green_fill if i % 2 == 0 else light_yellow_fill
        row += 1
    
    # Total
    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = f"Sum of weighted scores: {metrics['final_score']:.6f}"
    ws[f'A{row}'].font = subheader_font
    ws[f'B{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws[f'B{row}'].fill = subheader_fill
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border
    row += 2
    
    # Final Formula
    ws[f'A{row}'] = "FINAL FORMULA USING THESE EXACT VALUES:"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    formula = f"({metrics['cluster_count_score']:.6f} × 0.25) + ({metrics['coverage_score']:.6f} × 0.25) + ({metrics['precision_score']:.6f} × 0.25) + ({metrics['deviation_score']:.6f} × 0.25) = {metrics['final_score']:.6f}"
    ws[f'A{row}'] = formula
    ws[f'A{row}'].font = data_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].border = thin_border
    ws.merge_cells(f'A{row}:H{row}')
    row += 2
    
    # Individual Topic Performance
    ws[f'A{row}'] = "INDIVIDUAL TOPIC PERFORMANCE"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:L{row}')
    row += 1
    
    # Headers
    headers = ["Topic", "Model Title", "Model Messages", "Benchmark Title", "Benchmark Messages", 
               "Matched", "Missing", "Extra", "Coverage %", "Precision %", "Deviation %", "Perfect Match?"]
    
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_alignment
    
    row += 1
    
    # Add detailed results
    for i, result in enumerate(metrics['detailed_results']):
        topic_num = i + 1
        perfect_match = "YES" if result['perfect_match'] else "NO"
        
        row_data = [
            f"Topic {topic_num}",
            result['model_title'] if 'model_title' in result else "No Model Match",
            result['model_messages'] if 'model_messages' in result else 0,
            result['benchmark_title'],
            result['benchmark_messages'],
            result['matched_messages'],
            result['missing_messages'],
            result['extra_messages'],
            f"{result['coverage']:.2f}%",
            f"{result['precision']:.2f}%",
            f"{result['deviation']:.2f}%",
            perfect_match
        ]
        
        for j, value in enumerate(row_data):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].fill = light_red_fill if not result['perfect_match'] else (light_green_fill if row % 2 == 0 else light_yellow_fill)
            if j in [5, 6, 7, 8, 9, 10]:  # Numeric columns
                ws[f'{col}{row}'].alignment = center_alignment
        
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    output_file = "output/step2_client_analysis/gemini_2.0_flash_001_step2_analysis_PROPER.xlsx"
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)
    
    print(f"\n=== PROPER STEP 2 ANALYSIS RESULTS ===")
    print(f"Proper Step 2 analysis saved to: {output_file}")
    print(f"Model clusters: {metrics['model_cluster_count']}")
    print(f"Benchmark clusters: {metrics['benchmark_cluster_count']}")
    print(f"Cluster Count Score: {metrics['cluster_count_score']:.2f}%")
    print(f"Coverage Score: {metrics['coverage_score']:.2f}%")
    print(f"Precision Score: {metrics['precision_score']:.2f}%")
    print(f"Deviation Score: {metrics['deviation_score']:.2f}%")
    print(f"FINAL SCORE: {metrics['final_score']:.2f}")

if __name__ == "__main__":
    create_proper_step2_analysis()
