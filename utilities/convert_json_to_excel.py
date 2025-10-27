#!/usr/bin/env python3
"""
Convert Phase 5 JSON results to Excel format for client presentation
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def convert_metadata_results_to_excel(json_file, output_excel):
    """Convert phase 5 metadata results JSON to Excel"""
    
    print(f"📊 Converting {json_file} to Excel...")
    
    # Load JSON data
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_excel, engine='openpyxl')
    
    # 1. Summary Sheet
    summary_data = {
        'Provider': [data.get('provider', 'N/A')],
        'Model': [data.get('model', 'N/A')],
        'Phase': [data.get('phase', 'N/A')],
        'Timestamp': [data.get('timestamp', 'N/A')],
        'Success': [data.get('success', False)],
        'Duration (seconds)': [data.get('duration', 0)],
        'Total Tokens': [data.get('usage', {}).get('total_tokens', 0)],
        'Total Cost': [data.get('cost', {}).get('total_cost', 0)],
        'Clusters Processed': [data.get('clusters_processed', 0)]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # 2. All Topics Overview Sheet
    topics_overview = []
    for result in data.get('metadata_results', []):
        if result.get('success'):
            metadata = result.get('metadata', {})
            topics_overview.append({
                'Cluster ID': result.get('cluster_id', ''),
                'Title': metadata.get('title', ''),
                'Summary': metadata.get('summary', ''),
                'Urgency': metadata.get('urgency', ''),
                'Status': metadata.get('status', ''),
                'Deadline': metadata.get('deadline', ''),
                'Action Items Count': len(metadata.get('action_items', [])),
                'Participants': ', '.join(metadata.get('participants', []))
            })
    
    if topics_overview:
        topics_df = pd.DataFrame(topics_overview)
        topics_df.to_excel(writer, sheet_name='Topics Overview', index=False)
    
    # 3. Individual sheets for each topic with action items
    for result in data.get('metadata_results', []):
        if result.get('success'):
            cluster_id = result.get('cluster_id', 'unknown')
            metadata = result.get('metadata', {})
            
            # Create safe sheet name (Excel has 31 char limit)
            sheet_name = cluster_id.replace('_', ' ').title()[:31]
            
            # Topic details
            topic_data = {
                'Field': ['Title', 'Summary', 'Urgency', 'Status', 'Deadline', 'Participants', 'Tags'],
                'Value': [
                    metadata.get('title', ''),
                    metadata.get('summary', ''),
                    metadata.get('urgency', ''),
                    metadata.get('status', ''),
                    metadata.get('deadline', ''),
                    ', '.join(metadata.get('participants', [])),
                    ', '.join(metadata.get('tags', []))
                ]
            }
            
            topic_df = pd.DataFrame(topic_data)
            topic_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            
            # Action items in same sheet, below topic details
            action_items = metadata.get('action_items', [])
            if action_items:
                action_items_data = []
                for item in action_items:
                    action_items_data.append({
                        'Task': item.get('task', ''),
                        'Owner': item.get('owner', ''),
                        'Due Date': item.get('due_date', ''),
                        'Priority': item.get('priority', ''),
                        'Status': item.get('status', '')
                    })
                
                action_items_df = pd.DataFrame(action_items_data)
                action_items_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(topic_data) + 3)
    
    # 4. All Action Items Consolidated
    all_actions = []
    for result in data.get('metadata_results', []):
        if result.get('success'):
            cluster_id = result.get('cluster_id', '')
            metadata = result.get('metadata', {})
            title = metadata.get('title', '')
            
            for item in metadata.get('action_items', []):
                all_actions.append({
                    'Cluster': cluster_id,
                    'Topic': title,
                    'Task': item.get('task', ''),
                    'Owner': item.get('owner', ''),
                    'Due Date': item.get('due_date', ''),
                    'Priority': item.get('priority', ''),
                    'Status': item.get('status', '')
                })
    
    if all_actions:
        all_actions_df = pd.DataFrame(all_actions)
        all_actions_df.to_excel(writer, sheet_name='All Action Items', index=False)
    
    # Save the workbook
    writer.close()
    
    # Apply formatting
    format_excel_workbook(output_excel)
    
    print(f"✅ Excel file created: {output_excel}")
    return output_excel


def convert_benchmark_to_excel(json_file, output_excel):
    """Convert phase 5 benchmark JSON to Excel"""
    
    print(f"📊 Converting {json_file} to Excel...")
    
    # Load JSON data
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_excel, engine='openpyxl')
    
    # 1. All Topics Overview
    topics_overview = []
    for item in data:
        metadata = item.get('metadata', {})
        topics_overview.append({
            'Cluster ID': item.get('cluster_id', ''),
            'Title': metadata.get('title', ''),
            'Summary': metadata.get('summary', ''),
            'Urgency': metadata.get('urgency', ''),
            'Status': metadata.get('status', ''),
            'Deadline': metadata.get('deadline', ''),
            'Action Items Count': len(metadata.get('action_items', [])),
            'Participants': ', '.join(metadata.get('participants', []))
        })
    
    topics_df = pd.DataFrame(topics_overview)
    topics_df.to_excel(writer, sheet_name='Benchmark Overview', index=False)
    
    # 2. Individual sheets for each topic
    for item in data:
        cluster_id = item.get('cluster_id', 'unknown')
        metadata = item.get('metadata', {})
        
        # Create safe sheet name
        sheet_name = cluster_id.replace('_', ' ').title()[:31]
        
        # Topic details
        topic_data = {
            'Field': ['Title', 'Summary', 'Urgency', 'Status', 'Deadline', 'Participants', 'Tags'],
            'Value': [
                metadata.get('title', ''),
                metadata.get('summary', ''),
                metadata.get('urgency', ''),
                metadata.get('status', ''),
                metadata.get('deadline', ''),
                ', '.join(metadata.get('participants', [])),
                ', '.join(metadata.get('tags', []))
            ]
        }
        
        topic_df = pd.DataFrame(topic_data)
        topic_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        
        # Action items
        action_items = metadata.get('action_items', [])
        if action_items:
            action_items_data = []
            for action_item in action_items:
                action_items_data.append({
                    'Task': action_item.get('task', ''),
                    'Owner': action_item.get('owner', ''),
                    'Due Date': action_item.get('due_date', ''),
                    'Priority': action_item.get('priority', ''),
                    'Status': action_item.get('status', '')
                })
            
            action_items_df = pd.DataFrame(action_items_data)
            action_items_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(topic_data) + 3)
    
    # 3. All Action Items
    all_actions = []
    for item in data:
        cluster_id = item.get('cluster_id', '')
        metadata = item.get('metadata', {})
        title = metadata.get('title', '')
        
        for action_item in metadata.get('action_items', []):
            all_actions.append({
                'Cluster': cluster_id,
                'Topic': title,
                'Task': action_item.get('task', ''),
                'Owner': action_item.get('owner', ''),
                'Due Date': action_item.get('due_date', ''),
                'Priority': action_item.get('priority', ''),
                'Status': action_item.get('status', '')
            })
    
    all_actions_df = pd.DataFrame(all_actions)
    all_actions_df.to_excel(writer, sheet_name='All Action Items', index=False)
    
    writer.close()
    
    # Apply formatting
    format_excel_workbook(output_excel)
    
    print(f"✅ Excel file created: {output_excel}")
    return output_excel


def format_excel_workbook(excel_file):
    """Apply professional formatting to Excel workbook"""
    
    wb = load_workbook(excel_file)
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format headers (first row)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_style
        
        # Format all cells with borders and alignment
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border_style
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Freeze top row
        ws.freeze_panes = 'A2'
    
    wb.save(excel_file)
    print(f"✨ Applied formatting to {excel_file}")


def main():
    """Main conversion function"""
    
    print("=" * 70)
    print("📊 JSON to Excel Converter for Client")
    print("=" * 70)
    print()
    
    # Create output directory
    output_dir = "output/excel_for_client"
    os.makedirs(output_dir, exist_ok=True)
    
    # Files to convert
    files_to_convert = [
        {
            'input': 'output/phase5_metadata_generation/google_gemini-2.0-flash.json',
            'output': f'{output_dir}/Google_Gemini_2.0_Flash_Metadata_Results.xlsx',
            'type': 'results'
        },
        {
            'input': 'output/phase5_metadata_generation/google_gemini-2.0-flash-001.json',
            'output': f'{output_dir}/Google_Gemini_2.0_Flash_001_Metadata_Results.xlsx',
            'type': 'results'
        },
        {
            'input': 'phases/phase5_metadata_benchmark.json',
            'output': f'{output_dir}/Phase5_Metadata_Benchmark.xlsx',
            'type': 'benchmark'
        },
        {
            'input': 'output/phase5_metadata_generation/xai_grok-2-1212.json',
            'output': f'{output_dir}/XAI_Grok_2_1212_Metadata_Results.xlsx',
            'type': 'results'
        }
    ]
    
    converted_files = []
    
    for file_info in files_to_convert:
        input_file = file_info['input']
        output_file = file_info['output']
        file_type = file_info['type']
        
        if os.path.exists(input_file):
            try:
                if file_type == 'results':
                    convert_metadata_results_to_excel(input_file, output_file)
                else:
                    convert_benchmark_to_excel(input_file, output_file)
                
                converted_files.append(output_file)
            except Exception as e:
                print(f"❌ Error converting {input_file}: {e}")
        else:
            print(f"⚠️  File not found: {input_file}")
    
    print()
    print("=" * 70)
    print(f"✅ Conversion Complete! {len(converted_files)} files created")
    print("=" * 70)
    print()
    print("📁 Excel files created:")
    for file in converted_files:
        file_size = os.path.getsize(file) / 1024  # KB
        print(f"   • {file} ({file_size:.1f} KB)")
    print()
    print(f"📂 All files saved in: {output_dir}/")
    print()


if __name__ == "__main__":
    main()



