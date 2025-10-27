#!/usr/bin/env python3
"""
FULLY EXPLAINED Analysis - Every value traced to its source with detailed explanations
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

def load_balanced_refinement_data():
    """Load the balanced refinement model output"""
    file_path = "output/phase4_balanced_refinement/google_gemini-2.0-flash-001_balanced.json"
    with open(file_path, 'r') as f:
        return json.load(f)

def load_benchmark_clusters():
    """Load the benchmark clusters for comparison"""
    file_path = "phases/phase4_clusters_refined.json"
    with open(file_path, 'r') as f:
        return json.load(f)

def calculate_fully_explained_metrics(model_data, benchmark_data):
    """Calculate metrics with FULL explanations of where each value comes from"""
    
    model_clusters = model_data['refined_clusters']
    benchmark_clusters = benchmark_data
    
    print(f"\n=== SOURCE DATA ===")
    print(f"Benchmark file: phases/phase4_clusters_refined.json ({len(benchmark_clusters)} clusters)")
    print(f"Model file: output/phase4_balanced_refinement/google_gemini-2.0-flash-001_balanced.json ({len(model_clusters)} clusters)")
    
    # Create message ID sets for each cluster
    model_cluster_messages = {}
    for cluster in model_clusters:
        model_cluster_messages[cluster['cluster_id']] = set(cluster['message_ids'])
    
    benchmark_cluster_messages = {}
    for cluster in benchmark_clusters:
        benchmark_cluster_messages[cluster['cluster_id']] = set(cluster['message_ids'])
    
    # Calculate detailed results for each model cluster
    detailed_results = []
    total_matched = 0
    
    print(f"\n=== PER-CLUSTER MATCHING PROCESS ===")
    for i, model_cluster in enumerate(model_clusters, 1):
        model_id = model_cluster['cluster_id']
        model_messages = model_cluster_messages[model_id]
        
        print(f"\nCluster {i}: {model_id}")
        print(f"  Model messages: {len(model_messages)}")
        
        # Find the BEST matching benchmark cluster
        best_match = None
        best_overlap = 0
        best_benchmark_id = None
        
        for benchmark_cluster in benchmark_clusters:
            benchmark_id = benchmark_cluster['cluster_id']
            benchmark_messages = benchmark_cluster_messages[benchmark_id]
            
            overlap = len(model_messages.intersection(benchmark_messages))
            if overlap > 0:
                print(f"    vs {benchmark_id}: {overlap} messages overlap")
            if overlap > best_overlap:
                best_overlap = overlap
                best_match = benchmark_cluster
                best_benchmark_id = benchmark_id
        
        if best_match:
            print(f"  Best match: {best_benchmark_id} with {best_overlap} messages")
            total_matched += best_overlap
            
            benchmark_messages = benchmark_cluster_messages[best_benchmark_id]
            matched_messages = len(model_messages.intersection(benchmark_messages))
            missing_messages = len(benchmark_messages - model_messages)
            extra_messages = len(model_messages - benchmark_messages)
            
            precision = (matched_messages / len(model_messages)) * 100 if len(model_messages) > 0 else 0
            coverage = (matched_messages / len(benchmark_messages)) * 100 if len(benchmark_messages) > 0 else 0
            deviation = abs(len(model_messages) - len(benchmark_messages)) / len(benchmark_messages) * 100 if len(benchmark_messages) > 0 else 0
            
            detailed_results.append({
                'model_id': model_id,
                'model_title': model_cluster['draft_title'],
                'model_messages': len(model_messages),
                'benchmark_id': best_benchmark_id,
                'benchmark_title': best_match['draft_title'],
                'benchmark_messages': len(benchmark_messages),
                'matched_messages': matched_messages,
                'missing_messages': missing_messages,
                'extra_messages': extra_messages,
                'precision': precision,
                'coverage': coverage,
                'deviation': deviation,
                'perfect_match': matched_messages == len(benchmark_messages) and len(model_messages) == len(benchmark_messages),
                'explanation': f"Matched: {matched_messages}/{len(model_messages)} model msgs, {matched_messages}/{len(benchmark_messages)} benchmark msgs"
            })
    
    # Find unmatched benchmark clusters
    matched_benchmark_ids = {result['benchmark_id'] for result in detailed_results}
    unmatched_benchmarks = []
    for benchmark_cluster in benchmark_clusters:
        if benchmark_cluster['cluster_id'] not in matched_benchmark_ids:
            unmatched_benchmarks.append({
                'model_id': "No Model Match",
                'model_title': "No Model Match",
                'model_messages': 0,
                'benchmark_id': benchmark_cluster['cluster_id'],
                'benchmark_title': benchmark_cluster['draft_title'],
                'benchmark_messages': len(benchmark_cluster['message_ids']),
                'matched_messages': 0,
                'missing_messages': len(benchmark_cluster['message_ids']),
                'extra_messages': 0,
                'precision': 0,
                'coverage': 0,
                'deviation': 100,
                'perfect_match': False,
                'explanation': f"No model cluster matches this benchmark cluster"
            })
    
    # Combine all results
    all_results = detailed_results + unmatched_benchmarks
    
    # Calculate overall metrics
    total_benchmark_messages = sum(len(cluster['message_ids']) for cluster in benchmark_clusters)
    total_model_messages = sum(len(cluster['message_ids']) for cluster in model_clusters)
    total_matched_messages = sum(result['matched_messages'] for result in all_results)
    total_missing_messages = sum(result['missing_messages'] for result in all_results)
    total_extra_messages = sum(result['extra_messages'] for result in all_results)
    
    print(f"\n=== OVERALL TOTALS ===")
    print(f"Total Matched Messages: {total_matched_messages} (sum of all 'matched' from per-cluster analysis)")
    print(f"Total Benchmark Messages: {total_benchmark_messages} (sum of all benchmark cluster sizes)")
    print(f"Total Model Messages: {total_model_messages} (sum of all model cluster sizes)")
    
    # Step 1 methodology calculations
    cluster_count_score = min(len(model_clusters), len(benchmark_clusters)) / max(len(model_clusters), len(benchmark_clusters)) * 100
    coverage_score = (total_matched_messages / total_benchmark_messages) * 100 if total_benchmark_messages > 0 else 0
    precision_score = (total_matched_messages / total_model_messages) * 100 if total_model_messages > 0 else 0
    
    print(f"\n=== FINAL SCORE COMPONENTS ===")
    print(f"1. Coverage: {total_matched_messages}/{total_benchmark_messages} = {coverage_score:.2f}%")
    print(f"2. Precision: {total_matched_messages}/{total_model_messages} = {precision_score:.2f}%")
    
    # Average deviation across all topics
    avg_deviation = sum(result['deviation'] for result in all_results) / len(all_results) if all_results else 0
    deviation_score = max(0, 100 - avg_deviation)
    
    print(f"3. Average Deviation: {avg_deviation:.2f}%")
    print(f"4. Deviation Score: max(0, 100 - {avg_deviation:.2f}) = {deviation_score:.2f}%")
    
    # Final score using Step 1 formula
    final_score = (cluster_count_score * 0.25) + (coverage_score * 0.25) + (precision_score * 0.25) + (deviation_score * 0.25)
    
    print(f"\nFINAL SCORE: ({cluster_count_score:.2f} × 0.25) + ({coverage_score:.2f} × 0.25) + ({precision_score:.2f} × 0.25) + ({deviation_score:.2f} × 0.25) = {final_score:.2f}")
    
    return {
        'model_cluster_count': len(model_clusters),
        'benchmark_cluster_count': len(benchmark_clusters),
        'total_benchmark_messages': total_benchmark_messages,
        'total_model_messages': total_model_messages,
        'total_matched_messages': total_matched_messages,
        'total_missing_messages': total_missing_messages,
        'total_extra_messages': total_extra_messages,
        'cluster_count_score': cluster_count_score,
        'coverage_score': coverage_score,
        'precision_score': precision_score,
        'deviation_score': deviation_score,
        'avg_deviation': avg_deviation,
        'final_score': final_score,
        'detailed_results': all_results
    }

def create_fully_explained_analysis():
    """Create FULLY EXPLAINED analysis with complete traceability"""
    
    # Load data
    model_data = load_balanced_refinement_data()
    benchmark_data = load_benchmark_clusters()
    
    print("=== FULLY EXPLAINED STEP 2 ANALYSIS ===")
    
    # Calculate metrics
    metrics = calculate_fully_explained_metrics(model_data, benchmark_data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fully Explained Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=12, color="000000")
    data_font = Font(size=10)
    small_font = Font(size=9, italic=True)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    light_green_fill = PatternFill(start_color="E7F7E7", end_color="E7F7E7", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFE4CC", end_color="FFE4CC", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "FULLY EXPLAINED Step 2 Analysis: Google Gemini 2.0 Flash 001"
    ws[f'A{row}'].font = Font(bold=True, size=16, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_alignment
    ws.row_dimensions[row].height = 30
    row += 2
    
    # Data Sources
    ws[f'A{row}'] = "DATA SOURCES"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = light_orange_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    source_data = [
        ["Benchmark File", "phases/phase4_clusters_refined.json", "15 clusters, 300 total messages"],
        ["Model Output File", "output/phase4_balanced_refinement/google_gemini-2.0-flash-001_balanced.json", "13 clusters, 300 total messages"],
        ["Evaluation Method", "Each model cluster matched against best benchmark cluster", "Penalties for over-clustering and under-clustering"]
    ]
    
    for source in source_data:
        ws[f'A{row}'] = source[0]
        ws[f'B{row}'] = source[1]
        ws[f'C{row}'] = source[2]
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = data_font
        ws[f'C{row}'].font = data_font
        ws[f'A{row}'].fill = light_blue_fill
        ws[f'B{row}'].fill = light_yellow_fill
        ws[f'C{row}'].fill = light_green_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].alignment = left_alignment
        ws[f'B{row}'].alignment = left_alignment
        ws[f'C{row}'].alignment = left_alignment
        row += 1
    
    row += 1
    
    # WHERE VALUES COME FROM Section
    ws[f'A{row}'] = "WHERE DO THE FINAL SCORE VALUES COME FROM?"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = light_orange_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    value_explanations = [
        ["Value", "Source", "How It's Calculated", "Why It Matters"],
        ["272", "Sum of 'Matched' column in Individual Topic Performance table below", "Each model cluster finds best benchmark match, count overlapping messages, sum all matches", "Shows how many messages were correctly placed"],
        ["300 (benchmark)", "Sum of all benchmark cluster sizes", "13+15+15+15+15+15+23+25+31+29+16+22+22+22+22=300", "Total messages that should be clustered"],
        ["300 (model)", "Sum of all model cluster sizes", "44+15+15+15+15+23+25+31+29+22+22+22+22=300", "Total messages actually clustered by model"],
        ["Coverage 90.67%", "272 / 300 × 100", "Total matched messages / Total benchmark messages × 100", "How well model covers benchmark"],
        ["Precision 90.67%", "272 / 300 × 100", "Total matched messages / Total model messages × 100", "How accurate model clustering is"],
        ["Avg Deviation 25.00%", "Average of 'Deviation %' column from table below", "Sum all per-cluster deviations, divide by cluster count", "Average size difference per cluster"],
        ["Deviation Score 75.00%", "max(0, 100 - 25.00)", "100 minus average deviation", "Reward for smaller deviations"],
        ["Final Score 85.75", "(86.67×0.25)+(90.67×0.25)+(90.67×0.25)+(75.00×0.25)", "Weighted average of 4 components", "Overall model performance"]
    ]
    
    for i, exp in enumerate(value_explanations):
        for j, value in enumerate(exp):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font if i > 0 else subheader_font
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = left_alignment
            if i == 0:
                ws[f'{col}{row}'].fill = subheader_fill
            else:
                ws[f'{col}{row}'].fill = light_yellow_fill if j == 1 else (light_green_fill if j == 2 else light_blue_fill)
        ws.row_dimensions[row].height = 30 if i > 0 else 20
        row += 1
    
    row += 1
    
    # Formulas Section
    ws[f'A{row}'] = "FORMULAS FOR INDIVIDUAL TOPIC PERFORMANCE"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    formulas_data = [
        ["Metric", "Formula", "Example (Topic 1)", "Description"],
        ["Matched", "COUNT(Model Messages ∩ Benchmark Messages)", "16 messages (only topic_011 overlap)", "Number of messages in both model and benchmark cluster"],
        ["Coverage %", "(Matched / Benchmark Messages) × 100", "16/16 × 100 = 100.00%", "What % of benchmark cluster is found in model cluster"],
        ["Precision %", "(Matched / Model Messages) × 100", "16/44 × 100 = 36.36%", "What % of model cluster actually belongs to this benchmark"],
        ["Deviation %", "|Model - Benchmark| / Benchmark × 100", "|44-16|/16 × 100 = 175.00%", "How much model cluster size differs from benchmark"]
    ]
    
    for i, formula in enumerate(formulas_data):
        for j, value in enumerate(formula):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = left_alignment
            if i == 0:
                ws[f'{col}{row}'].font = subheader_font
                ws[f'{col}{row}'].fill = subheader_fill
            else:
                ws[f'{col}{row}'].fill = light_green_fill if i % 2 == 0 else light_yellow_fill
        ws.row_dimensions[row].height = 25
        row += 1
    
    row += 1
    
    # Step 1 Component Scores
    ws[f'A{row}'] = "STEP 1: FINAL SCORE CALCULATION"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    component_data = [
        ["Component", "Calculation", "Result", "Weight", "Weighted Score"],
        ["1. Cluster Count", f"min(13, 15) / max(13, 15) × 100", f"{metrics['cluster_count_score']:.2f}%", "0.25", f"{metrics['cluster_count_score'] * 0.25:.2f}"],
        ["2. Coverage", f"{metrics['total_matched_messages']}/{metrics['total_benchmark_messages']} × 100 (from table below)", f"{metrics['coverage_score']:.2f}%", "0.25", f"{metrics['coverage_score'] * 0.25:.2f}"],
        ["3. Precision", f"{metrics['total_matched_messages']}/{metrics['total_model_messages']} × 100 (from table below)", f"{metrics['precision_score']:.2f}%", "0.25", f"{metrics['precision_score'] * 0.25:.2f}"],
        ["4. Deviation", f"max(0, 100 - {metrics['avg_deviation']:.2f}) (avg from table below)", f"{metrics['deviation_score']:.2f}%", "0.25", f"{metrics['deviation_score'] * 0.25:.2f}"]
    ]
    
    for i, component in enumerate(component_data):
        for j, value in enumerate(component):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = left_alignment
            if i == 0:
                ws[f'{col}{row}'].font = subheader_font
                ws[f'{col}{row}'].fill = subheader_fill
            else:
                ws[f'{col}{row}'].fill = light_green_fill if i % 2 == 0 else light_yellow_fill
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Total
    ws[f'A{row}'] = "TOTAL (FINAL SCORE)"
    ws[f'B{row}'] = f"Sum of weighted scores"
    ws[f'C{row}'] = f"{metrics['final_score']:.2f}"
    ws[f'A{row}'].font = subheader_font
    ws[f'B{row}'].font = subheader_font
    ws[f'C{row}'].font = subheader_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'C{row}'].fill = header_fill
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'B{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'C{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border
    ws[f'C{row}'].border = thin_border
    ws[f'A{row}'].alignment = left_alignment
    ws[f'B{row}'].alignment = left_alignment
    ws[f'C{row}'].alignment = center_alignment
    row += 2
    
    # Individual Topic Performance
    ws[f'A{row}'] = "INDIVIDUAL TOPIC PERFORMANCE (Values used in calculations above)"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:L{row}')
    row += 1
    
    # Headers
    headers = ["Topic", "Model Title", "Model Msgs", "Benchmark Title", "Bench Msgs", 
               "Matched", "Missing", "Extra", "Coverage %", "Precision %", "Deviation %", "Perfect?"]
    
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_alignment
    
    row += 1
    
    # Add detailed results
    for i, result in enumerate(metrics['detailed_results']):
        topic_num = i + 1
        perfect_match = "YES" if result['perfect_match'] else "NO"
        
        row_data = [
            f"Topic {topic_num}",
            result['model_title'] if 'model_title' in result else "No Model Match",
            result['model_messages'] if 'model_messages' in result else 0,
            result['benchmark_title'],
            result['benchmark_messages'],
            result['matched_messages'],
            result['missing_messages'],
            result['extra_messages'],
            f"{result['coverage']:.2f}%",
            f"{result['precision']:.2f}%",
            f"{result['deviation']:.2f}%",
            perfect_match
        ]
        
        for j, value in enumerate(row_data):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].fill = light_red_fill if not result['perfect_match'] else (light_green_fill if row % 2 == 0 else light_yellow_fill)
            if j in [2, 4, 5, 6, 7]:  # Numeric columns
                ws[f'{col}{row}'].alignment = center_alignment
            else:
                ws[f'{col}{row}'].alignment = left_alignment
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 10, 'B': 35, 'C': 12, 'D': 35, 'E': 12, 
        'F': 10, 'G': 10, 'H': 10, 'I': 12, 'J': 12, 'K': 12, 'L': 10
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save the file
    output_file = "output/step2_client_analysis/gemini_2.0_flash_001_step2_analysis_FULLY_EXPLAINED.xlsx"
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)
    
    print(f"\n=== FILE SAVED ===")
    print(f"Fully explained analysis saved to: {output_file}")

if __name__ == "__main__":
    create_fully_explained_analysis()
