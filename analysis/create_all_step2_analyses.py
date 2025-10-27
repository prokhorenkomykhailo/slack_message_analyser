#!/usr/bin/env python3
"""
Create fully explained Step 2 analysis for ALL models and a comparison CSV
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import glob

def load_benchmark_clusters():
    """Load the benchmark clusters for comparison"""
    file_path = "phases/phase4_clusters_refined.json"
    with open(file_path, 'r') as f:
        return json.load(f)

def calculate_model_metrics(model_data, benchmark_data, model_name):
    """Calculate metrics for a single model"""
    
    model_clusters = model_data.get('refined_clusters', [])
    benchmark_clusters = benchmark_data
    
    if not model_clusters:
        print(f"  WARNING: {model_name} has no refined_clusters!")
        return None
    
    # Create message ID sets
    model_cluster_messages = {}
    for cluster in model_clusters:
        model_cluster_messages[cluster['cluster_id']] = set(cluster.get('message_ids', []))
    
    benchmark_cluster_messages = {}
    for cluster in benchmark_clusters:
        benchmark_cluster_messages[cluster['cluster_id']] = set(cluster['message_ids'])
    
    # Calculate detailed results
    detailed_results = []
    total_matched = 0
    
    for model_cluster in model_clusters:
        model_id = model_cluster['cluster_id']
        model_messages = model_cluster_messages[model_id]
        
        # Find best matching benchmark cluster
        best_match = None
        best_overlap = 0
        best_benchmark_id = None
        
        for benchmark_cluster in benchmark_clusters:
            benchmark_id = benchmark_cluster['cluster_id']
            benchmark_messages = benchmark_cluster_messages[benchmark_id]
            
            overlap = len(model_messages.intersection(benchmark_messages))
            if overlap > best_overlap:
                best_overlap = overlap
                best_match = benchmark_cluster
                best_benchmark_id = benchmark_id
        
        if best_match:
            total_matched += best_overlap
            benchmark_messages = benchmark_cluster_messages[best_benchmark_id]
            matched_messages = len(model_messages.intersection(benchmark_messages))
            missing_messages = len(benchmark_messages - model_messages)
            extra_messages = len(model_messages - benchmark_messages)
            
            precision = (matched_messages / len(model_messages)) * 100 if len(model_messages) > 0 else 0
            coverage = (matched_messages / len(benchmark_messages)) * 100 if len(benchmark_messages) > 0 else 0
            deviation = abs(len(model_messages) - len(benchmark_messages)) / len(benchmark_messages) * 100 if len(benchmark_messages) > 0 else 0
            
            detailed_results.append({
                'model_id': model_id,
                'model_title': model_cluster.get('draft_title', 'N/A'),
                'model_messages': len(model_messages),
                'benchmark_id': best_benchmark_id,
                'benchmark_title': best_match['draft_title'],
                'benchmark_messages': len(benchmark_messages),
                'matched_messages': matched_messages,
                'missing_messages': missing_messages,
                'extra_messages': extra_messages,
                'precision': precision,
                'coverage': coverage,
                'deviation': deviation,
                'perfect_match': matched_messages == len(benchmark_messages) and len(model_messages) == len(benchmark_messages)
            })
    
    # Find unmatched benchmark clusters
    matched_benchmark_ids = {result['benchmark_id'] for result in detailed_results}
    for benchmark_cluster in benchmark_clusters:
        if benchmark_cluster['cluster_id'] not in matched_benchmark_ids:
            detailed_results.append({
                'model_id': "No Model Match",
                'model_title': "No Model Match",
                'model_messages': 0,
                'benchmark_id': benchmark_cluster['cluster_id'],
                'benchmark_title': benchmark_cluster['draft_title'],
                'benchmark_messages': len(benchmark_cluster['message_ids']),
                'matched_messages': 0,
                'missing_messages': len(benchmark_cluster['message_ids']),
                'extra_messages': 0,
                'precision': 0,
                'coverage': 0,
                'deviation': 100,
                'perfect_match': False
            })
    
    # Calculate overall metrics
    total_benchmark_messages = sum(len(cluster['message_ids']) for cluster in benchmark_clusters)
    total_model_messages = sum(len(cluster.get('message_ids', [])) for cluster in model_clusters)
    total_matched_messages = sum(result['matched_messages'] for result in detailed_results)
    
    cluster_count_score = min(len(model_clusters), len(benchmark_clusters)) / max(len(model_clusters), len(benchmark_clusters)) * 100
    coverage_score = (total_matched_messages / total_benchmark_messages) * 100 if total_benchmark_messages > 0 else 0
    precision_score = (total_matched_messages / total_model_messages) * 100 if total_model_messages > 0 else 0
    avg_deviation = sum(result['deviation'] for result in detailed_results) / len(detailed_results) if detailed_results else 0
    deviation_score = max(0, 100 - avg_deviation)
    final_score = (cluster_count_score * 0.25) + (coverage_score * 0.25) + (precision_score * 0.25) + (deviation_score * 0.25)
    
    return {
        'model_name': model_name,
        'model_cluster_count': len(model_clusters),
        'benchmark_cluster_count': len(benchmark_clusters),
        'total_benchmark_messages': total_benchmark_messages,
        'total_model_messages': total_model_messages,
        'total_matched_messages': total_matched_messages,
        'cluster_count_score': cluster_count_score,
        'coverage_score': coverage_score,
        'precision_score': precision_score,
        'deviation_score': deviation_score,
        'avg_deviation': avg_deviation,
        'final_score': final_score,
        'detailed_results': detailed_results,
        'success': model_data.get('success', False),
        'duration': model_data.get('duration', 0),
        'merge_operations': len(model_data.get('ai_operations', {}).get('merge_operations', [])),
        'split_operations': len(model_data.get('ai_operations', {}).get('split_operations', []))
    }

def create_excel_for_model(metrics, output_dir):
    """Create Excel file for a single model"""
    
    if not metrics:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Step 2 Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="000000")
    data_font = Font(size=10)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    light_green_fill = PatternFill(start_color="E7F7E7", end_color="E7F7E7", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f"Step 2 Analysis: {metrics['model_name']}"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_alignment
    row += 2
    
    # Key Metrics
    ws[f'A{row}'] = "KEY METRICS SUMMARY"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    summary = [
        ["Model Clusters", metrics['model_cluster_count']],
        ["Benchmark Clusters", metrics['benchmark_cluster_count']],
        ["Coverage Score", f"{metrics['coverage_score']:.2f}%"],
        ["Precision Score", f"{metrics['precision_score']:.2f}%"],
        ["Deviation Score", f"{metrics['deviation_score']:.2f}%"],
        ["FINAL SCORE", f"{metrics['final_score']:.2f}"],
        ["Merge Operations", metrics['merge_operations']],
        ["Split Operations", metrics['split_operations']],
        ["Duration (sec)", f"{metrics['duration']:.2f}"]
    ]
    
    for item in summary:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = data_font
        ws[f'A{row}'].fill = light_blue_fill
        ws[f'B{row}'].fill = light_green_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # Score Components
    ws[f'A{row}'] = "FINAL SCORE CALCULATION"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    components = [
        ["Component", "Result", "Weight", "Weighted Score"],
        ["1. Cluster Count", f"{metrics['cluster_count_score']:.2f}%", "0.25", f"{metrics['cluster_count_score'] * 0.25:.2f}"],
        ["2. Coverage", f"{metrics['coverage_score']:.2f}%", "0.25", f"{metrics['coverage_score'] * 0.25:.2f}"],
        ["3. Precision", f"{metrics['precision_score']:.2f}%", "0.25", f"{metrics['precision_score'] * 0.25:.2f}"],
        ["4. Deviation", f"{metrics['deviation_score']:.2f}%", "0.25", f"{metrics['deviation_score'] * 0.25:.2f}"]
    ]
    
    for i, comp in enumerate(components):
        for j, val in enumerate(comp):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = val
            ws[f'{col}{row}'].font = subheader_font if i == 0 else data_font
            ws[f'{col}{row}'].fill = subheader_fill if i == 0 else (light_green_fill if i % 2 == 0 else light_yellow_fill)
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = center_alignment
        row += 1
    
    # Total
    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = f"{metrics['final_score']:.2f}"
    ws[f'A{row}'].font = subheader_font
    ws[f'B{row}'].font = subheader_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'B{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border
    row += 2
    
    # Individual Performance
    ws[f'A{row}'] = "INDIVIDUAL TOPIC PERFORMANCE"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:L{row}')
    row += 1
    
    headers = ["#", "Model Title", "Model Msgs", "Benchmark Title", "Bench Msgs", 
               "Matched", "Missing", "Extra", "Coverage %", "Precision %", "Deviation %", "Perfect?"]
    
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_alignment
    
    row += 1
    
    for i, result in enumerate(metrics['detailed_results'], 1):
        row_data = [
            i,
            result['model_title'],
            result['model_messages'],
            result['benchmark_title'],
            result['benchmark_messages'],
            result['matched_messages'],
            result['missing_messages'],
            result['extra_messages'],
            f"{result['coverage']:.2f}%",
            f"{result['precision']:.2f}%",
            f"{result['deviation']:.2f}%",
            "YES" if result['perfect_match'] else "NO"
        ]
        
        for j, val in enumerate(row_data):
            col = get_column_letter(j + 1)
            ws[f'{col}{row}'] = val
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].fill = light_green_fill if result['perfect_match'] else (light_red_fill if result['model_messages'] == 0 else light_yellow_fill)
            if j in [0, 2, 4, 5, 6, 7]:
                ws[f'{col}{row}'].alignment = center_alignment
            else:
                ws[f'{col}{row}'].alignment = left_alignment
        row += 1
    
    # Auto-adjust columns
    column_widths = {'A': 8, 'B': 30, 'C': 12, 'D': 30, 'E': 12, 'F': 10, 'G': 10, 'H': 10, 'I': 12, 'J': 12, 'K': 12, 'L': 10}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save
    model_safe_name = metrics['model_name'].replace('/', '_').replace(' ', '_')
    output_file = f"{output_dir}/{model_safe_name}_step2_analysis.xlsx"
    wb.save(output_file)
    print(f"  Created: {output_file}")

def main():
    """Main function to process all models"""
    
    print("=== CREATING STEP 2 ANALYSES FOR ALL MODELS ===\n")
    
    # Load benchmark
    benchmark_data = load_benchmark_clusters()
    print(f"Loaded benchmark: {len(benchmark_data)} clusters\n")
    
    # Find all model files
    model_files = glob.glob("output/phase4_balanced_refinement/*_balanced.json")
    model_files = [f for f in model_files if not f.endswith('comprehensive_results.json')]
    model_files.sort()
    
    print(f"Found {len(model_files)} model files\n")
    
    # Process each model
    all_metrics = []
    output_dir = "output/step2_client_analysis"
    os.makedirs(output_dir, exist_ok=True)
    
    for model_file in model_files:
        model_name = os.path.basename(model_file).replace('_balanced.json', '')
        print(f"Processing: {model_name}")
        
        try:
            with open(model_file, 'r') as f:
                model_data = json.load(f)
            
            metrics = calculate_model_metrics(model_data, benchmark_data, model_name)
            if metrics:
                all_metrics.append(metrics)
                create_excel_for_model(metrics, output_dir)
        except Exception as e:
            print(f"  ERROR: {e}")
    
    # Create comparison CSV
    print(f"\n=== CREATING COMPARISON CSV ===\n")
    
    comparison_data = []
    for m in all_metrics:
        comparison_data.append({
            'Model Name': m['model_name'],
            'Final Score': round(m['final_score'], 2),
            'Model Clusters': m['model_cluster_count'],
            'Coverage Score': round(m['coverage_score'], 2),
            'Precision Score': round(m['precision_score'], 2),
            'Deviation Score': round(m['deviation_score'], 2),
            'Cluster Count Score': round(m['cluster_count_score'], 2),
            'Total Matched Messages': m['total_matched_messages'],
            'Avg Deviation %': round(m['avg_deviation'], 2),
            'Merge Operations': m['merge_operations'],
            'Split Operations': m['split_operations'],
            'Duration (sec)': round(m['duration'], 2),
            'Success': m['success']
        })
    
    # Sort by Final Score (descending)
    comparison_data.sort(key=lambda x: x['Final Score'], reverse=True)
    
    # Add rank
    for i, data in enumerate(comparison_data, 1):
        data['Rank'] = i
    
    # Reorder columns
    df = pd.DataFrame(comparison_data)
    df = df[['Rank', 'Model Name', 'Final Score', 'Model Clusters', 'Coverage Score', 
             'Precision Score', 'Deviation Score', 'Cluster Count Score', 
             'Total Matched Messages', 'Avg Deviation %', 'Merge Operations', 
             'Split Operations', 'Duration (sec)', 'Success']]
    
    # Save CSV
    csv_file = f"{output_dir}/step2_all_models_comparison.csv"
    df.to_csv(csv_file, index=False)
    print(f"Created comparison CSV: {csv_file}\n")
    
    # Print top 5
    print("=== TOP 5 MODELS BY FINAL SCORE ===")
    for i, row in df.head(5).iterrows():
        print(f"{row['Rank']}. {row['Model Name']}: {row['Final Score']}")
    
    print(f"\n=== COMPLETE ===")
    print(f"Generated {len(all_metrics)} Excel files")
    print(f"Generated 1 comparison CSV")

if __name__ == "__main__":
    main()
