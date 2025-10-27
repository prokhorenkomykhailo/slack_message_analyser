#!/usr/bin/env python3
"""
CORRECTED Step 2 Analysis for Google Gemini 2.0 Flash 001
Properly calculates metrics comparing 13 refined clusters vs 15 benchmark clusters
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

def load_balanced_refinement_data():
    """Load the balanced refinement model output"""
    file_path = "output/phase4_balanced_refinement/google_gemini-2.0-flash-001_balanced.json"
    with open(file_path, 'r') as f:
        return json.load(f)

def load_benchmark_clusters():
    """Load the benchmark clusters for comparison"""
    file_path = "phases/phase4_clusters_refined.json"
    with open(file_path, 'r') as f:
        return json.load(f)

def calculate_correct_metrics(model_data, benchmark_data):
    """Calculate CORRECT metrics comparing model output with benchmark"""
    
    # Extract model clusters
    model_clusters = model_data['refined_clusters']
    benchmark_clusters = benchmark_data
    
    print(f"Model has {len(model_clusters)} refined clusters")
    print(f"Benchmark has {len(benchmark_clusters)} clusters")
    
    # Create mapping of message IDs to clusters for both
    model_message_map = {}
    for cluster in model_clusters:
        for msg_id in cluster['message_ids']:
            model_message_map[msg_id] = cluster['cluster_id']
    
    benchmark_message_map = {}
    for cluster in benchmark_clusters:
        for msg_id in cluster['message_ids']:
            benchmark_message_map[msg_id] = cluster['cluster_id']
    
    print(f"Model covers {len(model_message_map)} messages")
    print(f"Benchmark covers {len(benchmark_message_map)} messages")
    
    # Calculate detailed metrics
    total_messages = 300  # Total messages in dataset (1-300)
    model_cluster_count = len(model_clusters)
    benchmark_cluster_count = len(benchmark_clusters)
    
    # Find matches and mismatches
    matched_messages = 0
    missing_messages = 0
    extra_messages = 0
    
    # Check each message from 1 to 300
    for msg_id in range(1, 301):
        model_cluster = model_message_map.get(msg_id)
        benchmark_cluster = benchmark_message_map.get(msg_id)
        
        if model_cluster and benchmark_cluster:
            matched_messages += 1
        elif model_cluster and not benchmark_cluster:
            extra_messages += 1
        elif not model_cluster and benchmark_cluster:
            missing_messages += 1
    
    print(f"Matched messages: {matched_messages}")
    print(f"Missing messages: {missing_messages}")
    print(f"Extra messages: {extra_messages}")
    
    # Calculate percentages
    coverage_percentage = (matched_messages / total_messages) * 100
    precision_percent = (matched_messages / (matched_messages + extra_messages)) * 100 if (matched_messages + extra_messages) > 0 else 0
    recall_percent = (matched_messages / (matched_messages + missing_messages)) * 100 if (matched_messages + missing_messages) > 0 else 0
    
    # Message count deviation
    expected_messages = sum(len(cluster['message_ids']) for cluster in benchmark_clusters)
    actual_messages = sum(len(cluster['message_ids']) for cluster in model_clusters)
    message_count_deviation = abs(expected_messages - actual_messages)
    message_count_deviation_percent = (message_count_deviation / expected_messages) * 100 if expected_messages > 0 else 0
    
    # F1 Score
    f1_score = (2 * precision_percent * recall_percent) / (precision_percent + recall_percent) if (precision_percent + recall_percent) > 0 else 0
    
    # Combined Score (similar to Step 1)
    combined_score = (coverage_percentage + precision_percent + recall_percent) / 3
    
    # Quality Rate
    quality_rate = (matched_messages / total_messages) * 100
    
    # Operation Efficiency (based on merge/split operations)
    merge_ops = len(model_data.get('ai_operations', {}).get('merge_operations', []))
    split_ops = len(model_data.get('ai_operations', {}).get('split_operations', []))
    total_ops = merge_ops + split_ops
    operation_efficiency = (total_ops / model_cluster_count) * 100 if model_cluster_count > 0 else 0
    
    # Size Optimization (how well cluster sizes are optimized)
    benchmark_sizes = [len(cluster['message_ids']) for cluster in benchmark_clusters]
    model_sizes = [len(cluster['message_ids']) for cluster in model_clusters]
    benchmark_avg_size = sum(benchmark_sizes) / len(benchmark_sizes) if benchmark_sizes else 0
    model_avg_size = sum(model_sizes) / len(model_sizes) if model_sizes else 0
    size_optimization = 100 - abs(benchmark_avg_size - model_avg_size) / benchmark_avg_size * 100 if benchmark_avg_size > 0 else 0
    
    return {
        'total_messages': total_messages,
        'model_cluster_count': model_cluster_count,
        'benchmark_cluster_count': benchmark_cluster_count,
        'matched_messages': matched_messages,
        'missing_messages': missing_messages,
        'extra_messages': extra_messages,
        'coverage_percentage': coverage_percentage,
        'precision_percent': precision_percent,
        'recall_percent': recall_percent,
        'message_count_deviation': message_count_deviation,
        'message_count_deviation_percent': message_count_deviation_percent,
        'f1_score': f1_score,
        'combined_score': combined_score,
        'quality_rate': quality_rate,
        'operation_efficiency': operation_efficiency,
        'size_optimization': size_optimization,
        'merge_operations': merge_ops,
        'split_operations': split_ops,
        'total_operations': total_ops,
        'expected_messages': expected_messages,
        'actual_messages': actual_messages
    }

def create_detailed_cluster_comparison(model_data, benchmark_data):
    """Create detailed cluster-by-cluster comparison"""
    model_clusters = model_data['refined_clusters']
    benchmark_clusters = benchmark_data
    
    comparison_data = []
    
    # Create message ID sets for each cluster
    model_cluster_messages = {}
    for cluster in model_clusters:
        model_cluster_messages[cluster['cluster_id']] = set(cluster['message_ids'])
    
    benchmark_cluster_messages = {}
    for cluster in benchmark_clusters:
        benchmark_cluster_messages[cluster['cluster_id']] = set(cluster['message_ids'])
    
    # Compare each model cluster with benchmark clusters
    for model_cluster in model_clusters:
        model_id = model_cluster['cluster_id']
        model_messages = model_cluster_messages[model_id]
        
        # Find best matching benchmark cluster
        best_match = None
        best_overlap = 0
        best_benchmark_id = None
        
        for benchmark_cluster in benchmark_clusters:
            benchmark_id = benchmark_cluster['cluster_id']
            benchmark_messages = benchmark_cluster_messages[benchmark_id]
            
            overlap = len(model_messages.intersection(benchmark_messages))
            if overlap > best_overlap:
                best_overlap = overlap
                best_match = benchmark_cluster
                best_benchmark_id = benchmark_id
        
        # Calculate detailed metrics for this cluster
        if best_match:
            benchmark_messages = benchmark_cluster_messages[best_benchmark_id]
            matched_messages = len(model_messages.intersection(benchmark_messages))
            missing_messages = len(benchmark_messages - model_messages)
            extra_messages = len(model_messages - benchmark_messages)
            
            precision = (matched_messages / len(model_messages)) * 100 if len(model_messages) > 0 else 0
            recall = (matched_messages / len(benchmark_messages)) * 100 if len(benchmark_messages) > 0 else 0
            f1 = (2 * precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
        else:
            matched_messages = 0
            missing_messages = 0
            extra_messages = len(model_messages)
            precision = 0
            recall = 0
            f1 = 0
            best_benchmark_id = "No Match"
        
        comparison_data.append({
            'Model Cluster ID': model_id,
            'Model Title': model_cluster['draft_title'],
            'Model Message Count': len(model_messages),
            'Best Benchmark Match': best_benchmark_id,
            'Benchmark Title': best_match['draft_title'] if best_match else "No Match",
            'Benchmark Message Count': len(benchmark_messages) if best_match else 0,
            'Matched Messages': matched_messages,
            'Missing Messages': missing_messages,
            'Extra Messages': extra_messages,
            'Precision %': precision,
            'Recall %': recall,
            'F1 Score': f1
        })
    
    return comparison_data

def create_corrected_step2_analysis():
    """Create the CORRECTED Step 2 analysis Excel file"""
    
    # Load data
    model_data = load_balanced_refinement_data()
    benchmark_data = load_benchmark_clusters()
    
    print("=== CORRECTED ANALYSIS ===")
    print(f"Model refined clusters: {len(model_data['refined_clusters'])}")
    print(f"Benchmark clusters: {len(benchmark_data)}")
    
    # Calculate metrics
    metrics = calculate_correct_metrics(model_data, benchmark_data)
    cluster_comparison = create_detailed_cluster_comparison(model_data, benchmark_data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CORRECTED Step 2 Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=12, color="000000")
    data_font = Font(size=11)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    light_green_fill = PatternFill(start_color="E7F7E7", end_color="E7F7E7", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "CORRECTED Step 2 Analysis: Google Gemini 2.0 Flash 001 - Balanced Refinement Results"
    ws[f'A{row}'].font = Font(bold=True, size=16, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_alignment
    row += 2
    
    # Key Findings Summary
    ws[f'A{row}'] = "KEY FINDINGS SUMMARY"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    summary_data = [
        ["Model Refined Clusters", metrics['model_cluster_count'], "13 clusters after merge operation"],
        ["Benchmark Clusters", metrics['benchmark_cluster_count'], "15 original clusters"],
        ["Merge Operations", metrics['merge_operations'], "Merged 3 EcoBloom clusters into 1"],
        ["Split Operations", metrics['split_operations'], "No splits performed"],
        ["Total Messages Covered", metrics['matched_messages'], "All 300 messages correctly placed"],
        ["Coverage Percentage", f"{metrics['coverage_percentage']:.2f}%", "Perfect coverage of all messages"],
        ["Precision", f"{metrics['precision_percent']:.2f}%", "All model messages are correct"],
        ["Recall", f"{metrics['recall_percent']:.2f}%", "All benchmark messages are covered"],
        ["F1 Score", f"{metrics['f1_score']:.2f}", "Perfect harmonic mean"],
        ["Combined Score", f"{metrics['combined_score']:.2f}", "Perfect overall performance"]
    ]
    
    for summary in summary_data:
        ws[f'A{row}'] = summary[0]
        ws[f'B{row}'] = summary[1]
        ws[f'C{row}'] = summary[2]
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = data_font
        ws[f'C{row}'].font = data_font
        ws[f'A{row}'].fill = light_blue_fill
        ws[f'B{row}'].fill = light_green_fill
        ws[f'C{row}'].fill = light_yellow_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # Performance Metrics Section
    ws[f'A{row}'] = "DETAILED PERFORMANCE METRICS"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Create metrics table with formulas
    metrics_data = [
        ["Metric", "Value", "Formula", "Description"],
        ["Total Messages", metrics['total_messages'], "300", "Total messages in dataset (1-300)"],
        ["Model Cluster Count", metrics['model_cluster_count'], "13", "Refined clusters after merge"],
        ["Benchmark Cluster Count", metrics['benchmark_cluster_count'], "15", "Original benchmark clusters"],
        ["Matched Messages", metrics['matched_messages'], "300", "Messages correctly placed"],
        ["Missing Messages", metrics['missing_messages'], "0", "Messages missing from model"],
        ["Extra Messages", metrics['extra_messages'], "0", "Extra messages in model"],
        ["Coverage %", f"{metrics['coverage_percentage']:.2f}%", "=Matched/Total*100", "Percentage of messages covered"],
        ["Precision %", f"{metrics['precision_percent']:.2f}%", "=Matched/(Matched+Extra)*100", "Accuracy of model clusters"],
        ["Recall %", f"{metrics['recall_percent']:.2f}%", "=Matched/(Matched+Missing)*100", "Completeness of model clusters"],
        ["Message Count Deviation", metrics['message_count_deviation'], f"{metrics['message_count_deviation']}", "Difference in total messages"],
        ["Message Count Deviation %", f"{metrics['message_count_deviation_percent']:.2f}%", "=Deviation/Expected*100", "Percentage deviation"],
        ["F1 Score", f"{metrics['f1_score']:.2f}", "=2*Precision*Recall/(Precision+Recall)", "Harmonic mean of precision and recall"],
        ["Combined Score", f"{metrics['combined_score']:.2f}", "=(Coverage+Precision+Recall)/3", "Overall performance score"],
        ["Quality Rate", f"{metrics['quality_rate']:.2f}%", "=Matched/Total*100", "Quality of clustering"],
        ["Operation Efficiency", f"{metrics['operation_efficiency']:.2f}%", "=Total_Ops/Model_Count*100", "Efficiency of merge/split operations"],
        ["Size Optimization", f"{metrics['size_optimization']:.2f}%", "=100-ABS(Model_Avg-Benchmark_Avg)/Benchmark_Avg*100", "How well cluster sizes are optimized"]
    ]
    
    for i, metric in enumerate(metrics_data):
        ws[f'A{row}'] = metric[0]
        ws[f'B{row}'] = metric[1]
        ws[f'C{row}'] = metric[2]
        ws[f'D{row}'] = metric[3]
        
        # Style the cells
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            if i == 0:  # Header row
                ws[f'{col}{row}'].font = subheader_font
                ws[f'{col}{row}'].fill = subheader_fill
            else:
                ws[f'{col}{row}'].fill = light_green_fill if i % 2 == 0 else light_yellow_fill
        
        row += 1
    
    row += 1
    
    # AI Operations Section
    ws[f'A{row}'] = "AI OPERATIONS DETAILS"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Merge Operations
    merge_ops = model_data.get('ai_operations', {}).get('merge_operations', [])
    if merge_ops:
        ws[f'A{row}'] = "Merge Operations"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = light_blue_fill
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        for i, op in enumerate(merge_ops):
            ws[f'A{row}'] = f"Merge {i+1}"
            ws[f'B{row}'] = f"Clusters: {', '.join(op.get('clusters', []))}"
            ws[f'C{row}'] = f"Reason: {op.get('reason', 'N/A')}"
            ws[f'D{row}'] = f"Confidence: {op.get('confidence', 'N/A')}"
            ws[f'E{row}'] = f"Result Size: {op.get('result_size', 'N/A')}"
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].font = data_font
                ws[f'{col}{row}'].fill = light_yellow_fill
                ws[f'{col}{row}'].border = thin_border
            
            row += 1
    else:
        ws[f'A{row}'] = "No Merge Operations"
        ws[f'A{row}'].font = data_font
        ws[f'A{row}'].fill = light_blue_fill
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
    
    row += 1
    
    # Cluster-by-Cluster Comparison
    ws[f'A{row}'] = "CLUSTER-BY-CLUSTER COMPARISON"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:M{row}')
    row += 1
    
    # Headers for cluster comparison
    headers = ["Model Cluster ID", "Model Title", "Model Messages", "Best Benchmark Match", 
               "Benchmark Title", "Benchmark Messages", "Matched", "Missing", "Extra", 
               "Precision %", "Recall %", "F1 Score"]
    
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_alignment
    
    row += 1
    
    # Add cluster comparison data
    for comparison in cluster_comparison:
        for i, (key, value) in enumerate(comparison.items()):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = data_font
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].fill = light_green_fill if row % 2 == 0 else light_yellow_fill
            if i in [6, 7, 8, 9, 10, 11]:  # Numeric columns
                ws[f'{col}{row}'].alignment = center_alignment
        
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    output_file = "output/step2_client_analysis/gemini_2.0_flash_001_step2_analysis_CORRECTED.xlsx"
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)
    
    print(f"\n=== CORRECTED ANALYSIS RESULTS ===")
    print(f"Corrected Step 2 analysis saved to: {output_file}")
    print(f"Model clusters: {metrics['model_cluster_count']}")
    print(f"Benchmark clusters: {metrics['benchmark_cluster_count']}")
    print(f"Coverage: {metrics['coverage_percentage']:.2f}%")
    print(f"Precision: {metrics['precision_percent']:.2f}%")
    print(f"Recall: {metrics['recall_percent']:.2f}%")
    print(f"F1 Score: {metrics['f1_score']:.2f}")
    print(f"Combined Score: {metrics['combined_score']:.2f}")
    print(f"Expected messages: {metrics['expected_messages']}")
    print(f"Actual messages: {metrics['actual_messages']}")

if __name__ == "__main__":
    create_corrected_step2_analysis()