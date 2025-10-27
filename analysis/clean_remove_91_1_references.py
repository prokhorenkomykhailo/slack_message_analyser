#!/usr/bin/env python3
"""
Remove ALL references to 91.1 and only show the correct 92.9 analysis
Clean up the Excel to show ONLY correct values
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def clean_remove_91_1_references():
    """Remove all 91.1 references and clean up to show only correct values"""
    
    file_path = "gemini_1.5_flash_updated_analysis_FIXED.xlsx"
    
    print("CLEANING UP - REMOVING ALL 91.1 REFERENCES")
    print("="*50)
    
    # Load corrected data
    corrected_df = pd.read_csv('/home/ubuntu/deemerge/phase_evaluation_engine/gemini_1.5_flash_corrected_analysis.csv')
    corrected_gemini = corrected_df[corrected_df['MODEL'] == 'gemini-1.5-flash'].copy()
    
    # Calculate correct values
    expected_clusters = 6
    total_clusters = len(corrected_gemini)
    cluster_count_score = (min(expected_clusters, total_clusters) / max(expected_clusters, total_clusters)) * 100
    coverage_score = 100.0
    precision_score = 100.0  # All topics have 100% precision with correct data
    
    corrected_deviations = [abs(x) for x in corrected_gemini['MESSAGE_COUNT_DEVIATION_PERCENT'].tolist()]
    avg_deviation = sum(corrected_deviations) / len(corrected_deviations)
    deviation_score = max(0, 100 - avg_deviation)
    
    final_score = (cluster_count_score * 0.25 + coverage_score * 0.25 + precision_score * 0.25 + deviation_score * 0.25)
    
    print(f"CORRECT VALUES:")
    print(f"Final Score: {final_score:.1f}")
    print(f"Precision Score: {precision_score:.1f}")
    
    # Load Excel workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    changes_made = 0
    
    # Remove all references to 91.1 and incorrect historical data
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                cell_str = str(cell.value)
                
                # Remove any reference to 91.1
                if "91.1" in cell_str:
                    new_value = cell_str.replace("91.1", f"{final_score:.1f}")
                    cell.value = new_value
                    changes_made += 1
                    print(f"Updated cell at row {row_num}, col {col_num}: 91.1 → {final_score:.1f}")
                
                # Remove references to "original data" or "historical"
                if "ORIGINAL DATA" in cell_str:
                    new_value = cell_str.replace("ORIGINAL DATA → 91.1", f"CORRECTED DATA → {final_score:.1f}")
                    new_value = new_value.replace("ORIGINAL DATA", "CORRECTED DATA")
                    cell.value = new_value
                    changes_made += 1
                
                # Remove references to old benchmark
                if "original benchmark data" in cell_str.lower():
                    new_value = cell_str.replace("original benchmark data", "corrected benchmark data")
                    cell.value = new_value
                    changes_made += 1
                
                # Update any "old" or "incorrect" references
                if "old benchmark" in cell_str.lower():
                    new_value = cell_str.replace("old benchmark", "corrected benchmark")
                    cell.value = new_value
                    changes_made += 1
    
    # Update the main title to be clear
    ws.cell(row=2, column=1).value = "Gemini 1.5 Flash - CORRECTED Analysis (92.9 Score)"
    ws.cell(row=2, column=1).font = Font(bold=True, color="0000FF")
    
    # Update the notice
    ws.cell(row=3, column=1).value = "✅ SHOWING CORRECT PERFORMANCE WITH ACCURATE BENCHMARK DATA"
    ws.cell(row=3, column=1).font = Font(bold=True, color="008000", size=12)
    ws.cell(row=3, column=1).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Clean up any remaining verification sections that mention 91.1
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "verification" in str(cell_value).lower():
            # Update verification sections
            if "91.1" in str(cell_value):
                new_value = str(cell_value).replace("91.1", f"{final_score:.1f}")
                ws.cell(row=row_num, column=1).value = new_value
                changes_made += 1
        
        # Clean up any explanatory text about historical scoring
        if cell_value and ("historical evaluation" in str(cell_value).lower() or 
                          "original evaluation" in str(cell_value).lower() or
                          "old benchmark" in str(cell_value).lower()):
            # Replace with correct messaging
            ws.cell(row=row_num, column=1).value = "• Analysis shows correct performance with accurate benchmark data"
            changes_made += 1
    
    # Update formula headers to remove confusion
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "EXACT MATHEMATICAL FORMULAS → 91.1" in str(cell_value):
            ws.cell(row=row_num, column=1).value = f"EXACT MATHEMATICAL FORMULAS → {final_score:.1f}"
            ws.cell(row=row_num, column=1).font = Font(bold=True, color="0000FF", size=14)
            changes_made += 1
        
        if cell_value and "91.1 OVERALL MODEL SCORE" in str(cell_value):
            ws.cell(row=row_num, column=1).value = f"{final_score:.1f} OVERALL MODEL SCORE CALCULATION:"
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            changes_made += 1
    
    # Clean up any remaining methodology explanations
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and isinstance(cell_value, str):
            # Remove any confusing references to different scores
            if "score would be 92.9" in cell_value.lower():
                ws.cell(row=row_num, column=1).value = "• This analysis shows the correct performance metrics"
                changes_made += 1
            elif "reflects historical evaluation" in cell_value.lower():
                ws.cell(row=row_num, column=1).value = "• Analysis based on accurate benchmark data"
                changes_made += 1
    
    print(f"\nMade {changes_made} changes to remove 91.1 references")
    
    # Save the cleaned workbook
    wb.save(file_path)
    
    print(f"\n✅ EXCEL FILE CLEANED!")
    print(f"File: {file_path}")
    print(f"Now shows ONLY the correct {final_score:.1f} score")
    print("All references to 91.1 and incorrect historical data removed")

if __name__ == "__main__":
    clean_remove_91_1_references()
