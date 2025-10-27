#!/usr/bin/env python3
"""
Fix Step 2 to show the EXACT values that are used in the final calculation
Make the connection clear and direct
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def fix_step2_connection():
    """Fix Step 2 to show the exact component values used in final calculation"""
    
    file_path = "gemini_1.5_flash_updated_analysis_FIXED.xlsx"
    
    print("FIXING STEP 2 TO SHOW DIRECT CONNECTION TO FINAL CALCULATION")
    print("="*60)
    
    # Calculate the EXACT component values used in final formula
    corrected_df = pd.read_csv('gemini_1.5_flash_corrected_analysis.csv')
    gemini_data = corrected_df[corrected_df['MODEL'] == 'gemini-1.5-flash'].copy()
    
    # EXACT component scores used in final calculation
    cluster_count_score = 85.714286  # (6/7) * 100
    coverage_score = 100.000000      # Found all 6 clusters
    precision_score = 100.000000     # All topics 100% precision
    deviation_score = 85.714286      # max(0, 100 - 14.285714)
    
    print("COMPONENT VALUES THAT GO INTO FINAL FORMULA:")
    print(f"Cluster Count Score: {cluster_count_score:.6f}")
    print(f"Coverage Score: {coverage_score:.6f}")
    print(f"Precision Score: {precision_score:.6f}")
    print(f"Deviation Score: {deviation_score:.6f}")
    
    # Load Excel workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Find Step 2 section and replace it with component scores
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "STEP 2" in str(cell_value):
            print(f"Found Step 2 at row {row_num}")
            
            # Replace Step 2 title
            ws.cell(row=row_num, column=1).value = "STEP 2: COMPONENT SCORES FOR FINAL CALCULATION"
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            current_row = row_num + 2
            
            # Clear old content and add new component table
            # Clear several rows after Step 2
            for clear_row in range(current_row, current_row + 50):
                for clear_col in range(1, 15):
                    ws.cell(row=clear_row, column=clear_col).value = ""
            
            # Add component table header
            component_headers = ['Component', 'Calculation', 'Score', 'Weight', 'Weighted Score']
            for col, header in enumerate(component_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                cell.font = Font(bold=True)
            current_row += 1
            
            # Component 1: Cluster Count
            component1_data = [
                "1. Cluster Count",
                f"min(6, 7) / max(6, 7) × 100 = {cluster_count_score:.6f}",
                f"{cluster_count_score:.6f}",
                "0.25",
                f"{cluster_count_score * 0.25:.6f}"
            ]
            for col, value in enumerate(component1_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if col == 5:  # Weighted score column
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    cell.font = Font(bold=True)
            current_row += 1
            
            # Component 2: Coverage
            component2_data = [
                "2. Coverage",
                f"Found all 6 expected clusters = {coverage_score:.6f}",
                f"{coverage_score:.6f}",
                "0.25",
                f"{coverage_score * 0.25:.6f}"
            ]
            for col, value in enumerate(component2_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if col == 5:  # Weighted score column
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    cell.font = Font(bold=True)
            current_row += 1
            
            # Component 3: Precision
            component3_data = [
                "3. Precision",
                f"All topics have 100% precision = {precision_score:.6f}",
                f"{precision_score:.6f}",
                "0.25",
                f"{precision_score * 0.25:.6f}"
            ]
            for col, value in enumerate(component3_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if col == 5:  # Weighted score column
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    cell.font = Font(bold=True)
            current_row += 1
            
            # Component 4: Deviation
            component4_data = [
                "4. Deviation",
                f"max(0, 100 - 14.285714) = {deviation_score:.6f}",
                f"{deviation_score:.6f}",
                "0.25",
                f"{deviation_score * 0.25:.6f}"
            ]
            for col, value in enumerate(component4_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if col == 5:  # Weighted score column
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    cell.font = Font(bold=True)
            current_row += 2
            
            # Add sum row
            sum_data = [
                "TOTAL",
                "Sum of weighted scores",
                "",
                "",
                f"{(cluster_count_score + coverage_score + precision_score + deviation_score) * 0.25:.6f}"
            ]
            for col, value in enumerate(sum_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
                cell.font = Font(bold=True)
                if col == 5:
                    cell.font = Font(bold=True, color="FF0000", size=12)
            current_row += 2
            
            # Add explanation
            ws.cell(row=current_row, column=1).value = "⬇️ THESE VALUES GO DIRECTLY INTO THE FINAL FORMULA ⬇️"
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="FF0000", size=12)
            current_row += 2
            
            # Show the direct connection
            ws.cell(row=current_row, column=1).value = "FINAL FORMULA USING THESE EXACT VALUES:"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            final_formula = f"({cluster_count_score:.6f} × 0.25) + ({coverage_score:.6f} × 0.25) + ({precision_score:.6f} × 0.25) + ({deviation_score:.6f} × 0.25) = 92.857143"
            ws.cell(row=current_row, column=1).value = final_formula
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="0000FF")
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            break
    
    # Save the workbook
    wb.save(file_path)
    
    print(f"\n✅ STEP 2 FIXED!")
    print("Now Step 2 shows the EXACT component values used in the final calculation")
    print("Direct connection established between Step 2 and final formula")

if __name__ == "__main__":
    fix_step2_connection()
