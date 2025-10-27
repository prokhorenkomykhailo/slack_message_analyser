#!/usr/bin/env python3
"""
Fix the precision mismatch in Step 1 and format all percentages as 100.00%
Modify the existing FIXED file directly
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def fix_step1_precision():
    """Fix Step 1 precision values to match the original data that produced 91.1"""
    
    file_path = "gemini_1.5_flash_updated_analysis_FIXED.xlsx"
    
    print(f"Fixing Step 1 precision values in: {file_path}")
    
    # Load original data to get correct values
    original_df = pd.read_csv('/home/ubuntu/deemerge/phase_evaluation_engine/llm_analysis_with_improved_scores.csv')
    original_gemini = original_df[original_df['MODEL'] == 'gemini-1.5-flash'].copy()
    
    print("Original precision values from data that produced 91.1:")
    for _, row in original_gemini.iterrows():
        print(f"  Topic {row['CLUSTER_ID']}: {row['PRECISION_PERCENT']:.2f}%")
    
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Find Step 1 section and fix precision values
    step1_found = False
    precision_col = None
    
    for row_num in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_num, column=1).value
        
        if cell_value and "STEP 1" in str(cell_value):
            print(f"Found Step 1 at row {row_num}")
            step1_found = True
            continue
            
        if step1_found and cell_value and str(cell_value).strip() == "Topic":
            # Found the header row
            print(f"Found Step 1 header row at {row_num}")
            
            # Find precision column
            for col_num in range(1, ws.max_column + 1):
                header_value = ws.cell(row=row_num, column=col_num).value
                if header_value and "Precision" in str(header_value):
                    precision_col = col_num
                    print(f"Found Precision column at {col_num}")
                    break
            
            # Update the precision values for each topic
            if precision_col:
                for i, (_, orig_row) in enumerate(original_gemini.iterrows()):
                    topic_row = row_num + 1 + i
                    if topic_row <= ws.max_row:
                        # Get the original precision value
                        original_precision = orig_row['PRECISION_PERCENT']
                        
                        # Update the cell with correct precision
                        precision_cell = ws.cell(row=topic_row, column=precision_col)
                        precision_cell.value = f"{original_precision:.2f}%"
                        
                        print(f"Updated Topic {orig_row['CLUSTER_ID']} precision: {original_precision:.2f}%")
                        
                        # Also update coverage to match format
                        coverage_col = precision_col - 1  # Assuming coverage is before precision
                        if coverage_col >= 1:
                            coverage_cell = ws.cell(row=topic_row, column=coverage_col)
                            original_coverage = orig_row['COVERAGE_PERCENTAGE']
                            coverage_cell.value = f"{original_coverage:.2f}%"
                        
                        # Also update recall to match format  
                        recall_col = precision_col + 1  # Assuming recall is after precision
                        if recall_col <= ws.max_column:
                            recall_cell = ws.cell(row=topic_row, column=recall_col)
                            original_recall = orig_row['RECALL_PERCENT']
                            recall_cell.value = f"{original_recall:.2f}%"
                        
                        # Also update deviation to match format
                        deviation_col = precision_col + 2  # Assuming deviation is after recall
                        if deviation_col <= ws.max_column:
                            deviation_cell = ws.cell(row=topic_row, column=deviation_col)
                            original_deviation = orig_row['MESSAGE_COUNT_DEVIATION_PERCENT']
                            deviation_cell.value = f"{original_deviation:.2f}%"
            break
    
    # Now fix ALL percentage formatting throughout the sheet to use .2f format
    print("\nFixing percentage formatting throughout the sheet...")
    
    percentage_fixes = 0
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value and isinstance(cell.value, str):
                cell_str = str(cell.value).strip()
                
                # Check if it's a percentage that needs reformatting
                if cell_str.endswith('%') and cell_str != '':
                    try:
                        # Extract the number part
                        number_part = cell_str[:-1].strip()
                        
                        # Check if it's a simple number that can be reformatted
                        if '.' in number_part:
                            try:
                                number_value = float(number_part)
                                # Reformat to 2 decimal places
                                new_value = f"{number_value:.2f}%"
                                if new_value != cell_str:
                                    cell.value = new_value
                                    percentage_fixes += 1
                            except ValueError:
                                pass
                        elif number_part.replace('-', '').isdigit():
                            # It's a whole number, add .00
                            try:
                                number_value = float(number_part)
                                new_value = f"{number_value:.2f}%"
                                if new_value != cell_str:
                                    cell.value = new_value
                                    percentage_fixes += 1
                            except ValueError:
                                pass
                    except:
                        pass
    
    print(f"Fixed {percentage_fixes} percentage formatting issues")
    
    # Save the updated workbook
    wb.save(file_path)
    
    print(f"\n✅ Fixed file saved: {file_path}")
    print("Step 1 precision values now match the original data that produced 91.1")
    print("All percentages now formatted as XX.XX%")

if __name__ == "__main__":
    fix_step1_precision()
